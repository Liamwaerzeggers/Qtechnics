from fastapi import FastAPI, APIRouter, Depends, HTTPException, UploadFile, File, Response, Cookie, Header, Request, Query, Form, Body
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import json
import base64
import asyncio
import resend
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Any
import uuid
import shutil
from datetime import datetime, timezone, timedelta
import hashlib
import secrets
import pandas as pd
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
import re
from bs4 import BeautifulSoup

# AI imports for floor plan analysis
from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContent, ImageContent
import httpx
import hmac
import tempfile
from invoice_parser import InvoiceParser

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Import simple auth module
from auth_simple import router as auth2_router

# Resend Email Setup
resend.api_key = os.environ.get('RESEND_API_KEY')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'noreply@qtechnics.be')

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# ============= MODELS =============

# Auth Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore", populate_by_name=True)
    id: str = Field(alias="_id")
    email: str
    username: Optional[str] = None  # For username/password admins
    name: str
    picture: Optional[str] = None
    role: str = "admin"  # admin, worker, subcontractor, realtor, investor
    password_hash: Optional[str] = None  # For username/password admins
    # Multi-tenant profile links
    subcontractor_id: Optional[str] = None  # Link to Subcontractor profile
    realtor_id: Optional[str] = None  # Link to RealtorProfile
    investor_id: Optional[str] = None  # Link to InvestorProfile
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    email: str
    name: str
    picture: Optional[str] = None
    role: str = "admin"

class AdminCreate(BaseModel):
    username: str
    name: str
    email: str
    password: str

class Worker(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"WORKER-{str(uuid.uuid4())[:8].upper()}")
    username: str  # Changed from email to username
    name: str
    password_hash: str
    created_by: str  # admin user_id
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_active: bool = True

class WorkerCreate(BaseModel):
    username: str  # Changed from email to username
    name: str
    password: str

class UserSession(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    session_token: str
    expires_at: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SessionRequest(BaseModel):
    session_id: str

class SessionResponse(BaseModel):
    user: User
    session_token: str

# Lead Models
class Lead(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"LEAD-{str(uuid.uuid4())[:8].upper()}")
    name: str
    email: str
    phone: str
    address: str
    project_type: str
    description: str
    vat_number: Optional[str] = None  # BTW nummer voor bedrijven (optioneel voor particulieren)
    is_business: bool = False  # True als het een bedrijf is
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    user_id: str

class LeadCreate(BaseModel):
    name: str
    email: Optional[str] = "geen-email@example.com"
    phone: Optional[str] = "0000000000"
    address: Optional[str] = "Geen adres opgegeven"
    project_type: Optional[str] = "Renovatie"
    description: Optional[str] = ""
    vat_number: Optional[str] = None
    is_business: bool = False

class LeadUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    project_type: Optional[str] = None
    description: Optional[str] = None
    vat_number: Optional[str] = None
    is_business: Optional[bool] = None

# Login models for JSON body (more reliable than URL params on mobile)
class AdminLoginRequest(BaseModel):
    username: str
    password: str

class WorkerLoginRequest(BaseModel):
    username: str
    password: str

# Webhook model for external lead submissions (from website)
class WebsiteLeadWebhook(BaseModel):
    """Model for leads submitted from external website forms"""
    # Required fields
    name: str
    email: str
    phone: str
    
    # Optional fields - all extra form data
    address: Optional[str] = None
    postal_code: Optional[str] = None
    city: Optional[str] = None
    project_type: Optional[str] = "Website aanvraag"
    description: Optional[str] = None
    message: Optional[str] = None  # Contact form message
    
    # Business fields
    company_name: Optional[str] = None
    vat_number: Optional[str] = None
    is_business: bool = False
    
    # Extra metadata
    source: Optional[str] = "website"  # Where the lead came from
    form_name: Optional[str] = None  # Which form was submitted
    page_url: Optional[str] = None  # Which page the form was on
    
    # Any extra fields from the form (flexible)
    extra_data: Optional[dict] = None

# Quote Models
class Quote(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"OFF-{datetime.now(timezone.utc).year}-{str(uuid.uuid4())[:6].upper()}")
    lead_id: str
    quote_number: str = ""
    room: Optional[str] = None  # Kamer (bijv. "Badkamer", "Keuken", "Woonkamer")
    date: Optional[datetime] = Field(default_factory=lambda: datetime.now(timezone.utc))
    status: str = "concept"
    line_items: List[dict] = []  # Line items (materials + work)
    subtotal_labor: float = 0.0
    subtotal_material: float = 0.0
    total_excl_vat: float = 0.0
    vat_breakdown: dict = {}  # {"21": 100.50, "6": 25.30}
    total_vat: float = 0.0
    total_incl_vat: float = 0.0
    total_price: float = 0.0  # Backwards compatibility
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    user_id: str

class QuoteCreate(BaseModel):
    lead_id: str
    room: Optional[str] = None

class QuoteUpdate(BaseModel):
    status: Optional[str] = None
    room: Optional[str] = None
    is_sold: Optional[bool] = None  # Mark quote as sold/won

# Line Item Models
class LineItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    quote_id: str
    description: str
    quantity: float
    unit_price: float
    item_type: str  # "arbeid", "materiaal", "overig"
    vat_rate: float = 21.0  # BTW percentage (0, 6, 9, 21)
    total_excl_vat: float = 0.0
    vat_amount: float = 0.0
    total_incl_vat: float = 0.0
    total: float = 0.0  # For backwards compatibility
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LineItemCreate(BaseModel):
    description: str
    quantity: float
    unit_price: float
    item_type: str
    vat_rate: float = 21.0

class LineItemUpdate(BaseModel):
    description: Optional[str] = None
    quantity: Optional[float] = None
    unit_price: Optional[float] = None
    item_type: Optional[str] = None
    vat_rate: Optional[float] = None

# Material Models
class Material(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sku: str
    name: str
    description: Optional[str] = None
    category: Optional[str] = None  # e.g., "Vloer", "Muur", "Meubel"
    subcategory: Optional[str] = None  # e.g., "Tegels", "Parket", "Kast"
    brand: Optional[str] = None
    price: float
    unit: Optional[str] = "stuk"  # m², stuk, lm
    image_url: Optional[str] = None  # URL to product image
    colors: Optional[List[str]] = None  # Available colors for this material
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    user_id: str

class WorkItem(BaseModel):
    """Work/Labor items for quotes (stucwerk, schilderwerk, etc)"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str  # Titel van het werk
    unit: str  # Eenheid: m², lm, stuks, uur
    price: float  # Verkoopprijs ex BTW
    category: Optional[str] = None  # e.g., "Vloer", "Muur", "Plafond"
    # Multi-tenant: component labels voor renovatiecalculator
    component_label: Optional[str] = None  # "vloer" | "muur" | "plafond" | "elektriciteit" | "sanitair" | "verwarming" | "isolatie" | "overig"
    room_types: List[str] = Field(default_factory=lambda: ["all"])  # ["all"] of ["bathroom", "kitchen"] etc.
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    user_id: str

# Project Models
class Project(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"PROJ-{str(uuid.uuid4())[:8].upper()}")
    lead_id: Optional[str] = None  # NEW: Optional link to lead
    quote_id: Optional[str] = None  # BLIJFT: Backward compatible (was verplicht)
    name: str
    status: str = "gepland"  # eerste bezoek, offerte in opmaak, gepland, in uitvoering, voltooid, niet verkocht
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    notes: Optional[str] = None
    color: str = "#1E40AF"  # Default blue color
    
    # Sales tracking
    is_sold: bool = False  # Whether the project has been sold
    not_sold_reason: Optional[str] = None  # Reason if marked as "niet verkocht"
    not_sold_date: Optional[datetime] = None  # When marked as niet verkocht
    
    # NIEUWE SECTIE: Eerste Bezoek
    # Support both old format (List[str]) and new format (List[dict] with room)
    first_visit_photos: List[Any] = []  # Photo URLs/paths OR photo objects with room
    first_visit_notes: str = ""  # Notes from first visit
    first_visit_date: Optional[datetime] = None
    
    # NIEUWE SECTIE: Algemene Project Notities (doorlopend)
    # [{id, text, created_at, created_by, is_task, assigned_to, task_completed, task_completed_at}]
    project_notes: List[dict] = []
    
    # NIEUWE SECTIE: Metingen & Werk Items (voor offerte generatie)
    measurements: List[dict] = []  # [{work_item_id, title, quantity, unit, price, vat_rate}]
    room_measurements: List[dict] = []  # LEGACY - replaced by rooms
    
    # NIEUWE SECTIE: Kamers & Renovatiecalculator
    rooms: List[dict] = []  # Same format as Property rooms [{id, name, room_type, length, width, height, floor_area, wall_area, ceiling_area}]
    floor_plan_url: Optional[str] = None
    renovation_calculation_id: Optional[str] = None
    
    # LEGACY: Grondplan Analyses (AI) - replaced by new system
    floor_plan_analyses: List[dict] = []
    
    # NIEUWE SECTIE: 3D Ontwerpen
    design_3d_files: List[dict] = []  # [{filename, url, upload_date}]
    
    # Cost tracking
    labor_cost_per_hour: float = 0.0
    labor_hours: float = 0.0
    material_costs: float = 0.0
    material_costs_incl_vat: float = 0.0
    other_costs: float = 0.0
    invoice_uploads: List[dict] = []  # [{filename, total_excl_vat, total_incl_vat, vat_amount, upload_date}]
    total_costs: float = 0.0
    total_costs_incl_vat: float = 0.0
    sales_price: float = 0.0  # Total from SOLD quotes only (incl VAT)
    potential_sales: float = 0.0  # Total from approved but not sold quotes
    profit: Optional[float] = 0.0
    profit_margin: float = 0.0
    is_archived: bool = False  # Soft delete - hidden from workers when True
    visible_to_workers: bool = False  # Toggle: show/hide project for workers
    
    # NIEUWE SECTIE: Planning
    # scheduled_days format: [{id, start_date, end_date, description, team_name}]
    scheduled_days: List[dict] = []  # Work periods with team assignment
    required_materials: str = ""  # Manual text for additional materials needed
    material_reminder_sent: bool = False  # Track if reminder was sent
    
    # NIEUWE SECTIE: Klantportaal
    customer_access_token: Optional[str] = None  # Unique token for customer portal access
    customer_messages: List[dict] = []  # [{id, message, sender, timestamp, is_from_customer}]
    customer_rating: Optional[int] = None  # 1-5 stars
    customer_rating_comment: Optional[str] = None  # Optional comment with rating
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    user_id: str
    lead_address: Optional[str] = None  # Enriched: address from lead for worker view

class ProjectCreate(BaseModel):
    lead_id: Optional[str] = None  # NEW: Optional
    quote_id: Optional[str] = None  # BLIJFT: Backward compatible
    name: str
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    notes: Optional[str] = None

class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    status: Optional[str] = None
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    notes: Optional[str] = None
    color: Optional[str] = None
    labor_cost_per_hour: Optional[float] = None
    labor_hours: Optional[float] = None
    material_costs: Optional[float] = None
    material_costs_incl_vat: Optional[float] = None
    other_costs: Optional[float] = None
    sales_price: Optional[float] = None  # Total from approved quotes
    lead_id: Optional[str] = None  # Allow updating lead_id
    scheduled_days: Optional[List[dict]] = None  # Planning days with notes
    required_materials: Optional[str] = None  # Manual materials text
    room_measurements: Optional[List[dict]] = None  # Room-based measurements
    floor_plan_analyses: Optional[List[dict]] = None  # AI floor plan analyses

class InvoiceUpload(BaseModel):
    filename: str
    total_excl_vat: float
    total_incl_vat: float
    vat_amount: float
    notes: Optional[str] = None

# Legacy Document Models (voor oude PDF's uit vorig systeem)
class LegacyDocument(BaseModel):
    id: str = Field(default_factory=lambda: f"DOC-{str(uuid.uuid4())[:8].upper()}")
    project_id: str
    document_type: str  # "offerte", "factuur", "anders"
    filename: str
    original_filename: str
    document_date: Optional[str] = None  # Datum van het originele document
    description: Optional[str] = None
    total_price: Optional[float] = None  # Totaalprijs van het document (incl BTW)
    visible_to_customer: bool = False  # Zichtbaar in klantenportaal
    file_size: int = 0
    uploaded_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    uploaded_by: str = ""

class LegacyDocumentCreate(BaseModel):
    document_type: str  # "offerte", "factuur", "anders"
    document_date: Optional[str] = None
    description: Optional[str] = None
    total_price: Optional[float] = None  # Totaalprijs voor financieel overzicht

class LegacyDocumentUpdate(BaseModel):
    document_type: Optional[str] = None
    document_date: Optional[str] = None
    description: Optional[str] = None
    total_price: Optional[float] = None
    visible_to_customer: Optional[bool] = None
    is_sold: Optional[bool] = None  # Mark legacy quote as sold/won

# Calendar Models
class CalendarEvent(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    title: str
    start_date: datetime
    end_date: datetime
    description: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    user_id: str

# Werkbon (Daily Report) Models
class MaterialUsed(BaseModel):
    """Material from quote that was used"""
    material_id: str  # Line item ID from quote
    description_nl: str
    description_uk: str
    quantity_used: Optional[float] = None
    notes: Optional[str] = None

class ExtraMaterial(BaseModel):
    """Extra material not in original quote"""
    description_nl: str
    description_uk: str
    quantity: Optional[str] = None
    photo_url: Optional[str] = None

class DailyReport(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    
    # Hours worked and labor tracking
    hours_worked: Optional[float] = None
    number_of_workers: int = 1  # Number of workers on this day
    hourly_rate: float = 34.0  # €34 per hour default
    labor_cost: Optional[float] = None  # Calculated: hours_worked * number_of_workers * hourly_rate
    
    # Materials tracking (NO PRICES)
    materials_used: List[MaterialUsed] = []  # From quote
    extra_materials: List[ExtraMaterial] = []  # Added by workers
    
    # Work description (bilingual)
    work_description_nl: Optional[str] = None
    work_description_uk: Optional[str] = None
    
    # Office feedback (only visible to office)
    office_feedback_nl: Optional[str] = None
    office_feedback_uk: Optional[str] = None
    
    # Photos
    photos: List[str] = []
    
    # Customer visibility - NEW
    visible_to_customer: bool = False  # Toggle: show/hide work slip for customer portal
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    user_id: str

class DailyReportCreate(BaseModel):
    project_id: str
    date: Optional[datetime] = None
    hours_worked: Optional[float] = None
    number_of_workers: int = 1
    hourly_rate: float = 34.0
    materials_used: Optional[List[MaterialUsed]] = []
    extra_materials: Optional[List[ExtraMaterial]] = []
    work_description_nl: Optional[str] = None
    work_description_uk: Optional[str] = None

class DailyReportUpdate(BaseModel):
    hours_worked: Optional[float] = None
    number_of_workers: Optional[int] = None
    hourly_rate: Optional[float] = None
    materials_used: Optional[List[MaterialUsed]] = None
    extra_materials: Optional[List[ExtraMaterial]] = None
    work_description_nl: Optional[str] = None
    work_description_uk: Optional[str] = None
    office_feedback_nl: Optional[str] = None
    office_feedback_uk: Optional[str] = None
    visible_to_customer: Optional[bool] = None  # NEW: Toggle customer visibility

# Invoice Models
class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: str  # FACT-2025-001
    project_id: str
    quote_id: str
    milestone: str  # "10_approval", "40_before_start", "40_completion", "10_satisfaction"
    milestone_percentage: int  # 10, 40, 40, 10
    
    # Line items (copied from quote)
    line_items: List[dict] = []
    
    # Totals
    subtotal_labor: float = 0.0
    subtotal_material: float = 0.0
    total_excl_vat: float = 0.0
    vat_breakdown: dict = {}  # {"6": 123.45, "21": 456.78}
    total_vat: float = 0.0
    total_incl_vat: float = 0.0
    
    # Payment info
    payment_status: str = "unpaid"  # unpaid, paid, overdue
    payment_term_days: int = 7
    payment_reference: Optional[str] = None  # OGM structured reference (+++123/4567/89012+++)
    due_date: datetime
    paid_date: Optional[datetime] = None
    
    # Peppol/Billit integration
    peppol_status: str = "not_sent"  # not_sent, sending, sent, delivered, failed
    billit_invoice_id: Optional[str] = None
    peppol_message_id: Optional[str] = None
    peppol_sent_at: Optional[datetime] = None
    peppol_delivered_at: Optional[datetime] = None
    peppol_error: Optional[str] = None
    
    # Dates
    invoice_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    user_id: str

class InvoiceCreate(BaseModel):
    milestone: str
    milestone_percentage: int

class InvoiceUpdate(BaseModel):
    payment_status: Optional[str] = None
    paid_date: Optional[datetime] = None

# Manual Invoice Entry - for phased invoicing without automatic invoice generation
class ManualInvoiceEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"MAN-{str(uuid.uuid4())[:8].upper()}")
    project_id: str
    amount: float  # Gefactureerd bedrag (incl. BTW)
    description: str = ""  # Bijv. "Fase 1", "Deelbetaling maart"
    invoice_date: datetime  # Datum voor maandelijkse rapportage
    send_via_billit: bool = False  # Of er een echte factuur moet worden verstuurd
    billit_invoice_id: Optional[str] = None  # Reference to created invoice if sent via Billit
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    user_id: str

class ManualInvoiceEntryCreate(BaseModel):
    amount: float
    description: str = ""
    invoice_date: str  # YYYY-MM-DD format
    send_via_billit: bool = False

# Quick Tasks - standalone tasks not tied to a project
class QuickTask(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: Optional[str] = None
    start_date: Optional[str] = None  # YYYY-MM-DD, optional for small tasks
    end_date: Optional[str] = None  # YYYY-MM-DD, optional for small tasks
    team_name: Optional[str] = None  # Assigned team
    completed: bool = False  # Task completion status
    completed_at: Optional[str] = None  # ISO datetime when completed
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    user_id: str

class QuickTaskCreate(BaseModel):
    title: str
    description: Optional[str] = None
    start_date: Optional[str] = None  # Optional - can be set later by dragging to calendar
    end_date: Optional[str] = None  # Optional - can be set later by dragging to calendar
    team_name: Optional[str] = None  # Optional - can assign team at creation

class QuickTaskUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    team_name: Optional[str] = None
    completed: Optional[bool] = None
    completed_at: Optional[str] = None

# ============= WORKER TASK MODELS =============

class WorkerTask(BaseModel):
    """Task assigned to a worker from a project note"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"TASK-{str(uuid.uuid4())[:8].upper()}")
    project_id: str
    project_name: str
    note_id: str  # Reference to the project note
    text: str  # The task description
    assigned_to: str  # Worker ID
    assigned_to_name: str  # Worker name for display
    assigned_by: str  # Admin ID who assigned
    assigned_by_name: str  # Admin name
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    completed: bool = False
    completed_at: Optional[str] = None
    seen: bool = False  # Whether worker has seen the notification

class WorkerTaskCreate(BaseModel):
    project_id: str
    note_id: str
    text: str
    assigned_to: str

# ============= MULTI-TENANT MODELS =============

# Room binnen een Property
class PropertyRoom(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str  # "Woonkamer", "Badkamer 1", etc.
    room_type: str = "other"  # "living" | "bedroom" | "bathroom" | "kitchen" | "hallway" | "other"
    length: float = 0.0  # meter
    width: float = 0.0  # meter
    height: float = 2.7  # meter (standaard)
    floor_area: float = 0.0  # Berekend: length * width
    ceiling_area: float = 0.0  # Berekend: length * width
    wall_area: float = 0.0  # Berekend: 2*(length + width) * height
    windows: int = 0
    doors: int = 0
    notes: str = ""

class PropertyRoomCreate(BaseModel):
    name: str
    room_type: str = "other"
    length: float
    width: float
    height: float = 2.7
    windows: int = 0
    doors: int = 0
    notes: str = ""

# Property (Pand) - voor makelaars en investeerders
class Property(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"PROP-{str(uuid.uuid4())[:8].upper()}")
    owner_type: str  # "realtor" | "investor" | "admin"
    owner_id: str  # ID van eigenaar
    
    # Bron info
    source_url: Optional[str] = None
    source_platform: str = "manual"  # "immoweb" | "zimmo" | "immoscoop" | "manual"
    
    # Adres
    address: str = ""
    postal_code: str = ""
    city: str = ""
    
    # Kenmerken
    living_area: float = 0.0  # Bewoonbare opp. (m²)
    plot_area: float = 0.0  # Grondoppervlakte (m²)
    bedrooms: int = 0
    bathrooms: int = 0
    construction_year: Optional[int] = None
    epc_score: Optional[str] = None  # A, B, C, D, E, F, G
    epc_value: Optional[float] = None  # kWh/m²/jaar
    
    # Kamers met afmetingen
    rooms: List[PropertyRoom] = []
    
    # Prijzen
    asking_price: float = 0.0
    estimated_value: float = 0.0  # Geschatte waarde na renovatie
    
    # Foto's
    photos: List[str] = []
    
    # Grondplan
    floor_plan_url: Optional[str] = None
    
    # Status
    status: str = "imported"  # "imported" | "analyzing" | "calculated" | "shared"
    
    # Renovatie link
    renovation_calculation_id: Optional[str] = None
    
    # Sharing - welke users mogen dit pand zien
    shared_with: List[str] = []
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PropertyCreate(BaseModel):
    source_url: Optional[str] = None
    address: str = ""
    postal_code: str = ""
    city: str = ""
    living_area: float = 0.0
    plot_area: float = 0.0
    bedrooms: int = 0
    bathrooms: int = 0
    construction_year: Optional[int] = None
    epc_score: Optional[str] = None
    epc_value: Optional[float] = None
    asking_price: float = 0.0
    rooms: List[PropertyRoomCreate] = []

class PropertyUpdate(BaseModel):
    address: Optional[str] = None
    postal_code: Optional[str] = None
    city: Optional[str] = None
    living_area: Optional[float] = None
    plot_area: Optional[float] = None
    bedrooms: Optional[int] = None
    bathrooms: Optional[int] = None
    construction_year: Optional[int] = None
    epc_score: Optional[str] = None
    epc_value: Optional[float] = None
    asking_price: Optional[float] = None
    estimated_value: Optional[float] = None
    status: Optional[str] = None

# Calculation Item - één werkpost in een berekening
class CalculationItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    work_item_id: Optional[str] = None  # Referentie naar WorkItem
    title: str
    quantity: float
    unit: str
    unit_price: float
    total: float
    included: bool = True  # True = meegenomen in totaal
    is_subcontractor: bool = False
    subcontractor_id: Optional[str] = None
    category: str = ""  # "vloer_optie", "muur_scenario", "plafond", "elektriciteit", etc.
    is_selected: bool = True  # True = dit is de gekozen optie in een category
    option_group: Optional[str] = None  # Group alternatieven samen (bijv. "vloer_afwerking")

# Room Calculation - berekening per kamer
class RoomCalculation(BaseModel):
    room_id: str
    room_name: str
    room_type: str = "other"  # "living", "bedroom", "bathroom", "kitchen", "hallway", "other"
    floor_area: float = 0.0
    wall_area: float = 0.0
    ceiling_area: float = 0.0
    room_height: float = 2.55  # Hoogte van de kamer
    height_source: str = "standaard"  # "standaard" of "opgegeven"
    floor_items: List[CalculationItem] = []
    wall_items: List[CalculationItem] = []
    ceiling_items: List[CalculationItem] = []
    other_items: List[CalculationItem] = []
    subtotal: float = 0.0
    # Selected scenarios per category
    selected_floor_option: str = "tegels"  # tegels, parket, vinyl, laminaat
    selected_wall_scenario: str = "nieuw_pleisterwerk"  # nieuw_pleisterwerk, egaliseren, gyproc

# Renovation Calculation - volledige renovatieberekening
class RenovationCalculation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"CALC-{str(uuid.uuid4())[:8].upper()}")
    property_id: Optional[str] = None  # For properties
    project_id: Optional[str] = None   # For projects
    calculated_by: str  # user_id
    
    room_calculations: List[RoomCalculation] = []
    
    total_min: float = 0.0
    total_realistic: float = 0.0
    total_max: float = 0.0
    
    estimated_duration_weeks: int = 0
    estimated_epc_improvement: str = ""
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# MaterialRequest - Werkman Materiaal Aanvraag
class MaterialRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"MATREQ-{str(uuid.uuid4())[:8].upper()}")
    
    # Request details
    title: str  # Naam van het materiaal
    quantity: str  # Hoeveelheid (als string voor flexibiliteit: "5 zakken", "10m²", etc.)
    needed_by: str  # Datum wanneer nodig op werf
    photo_url: Optional[str] = None  # Foto van het materiaal
    notes: Optional[str] = None  # Extra notities
    
    # Project info (optional)
    project_id: Optional[str] = None
    project_name: Optional[str] = None
    
    # Requester info
    requested_by: str  # Worker user_id
    requested_by_name: str  # Worker name
    
    # Status tracking
    status: str = "pending"  # "pending" | "ordered" | "delivered"
    is_ordered: bool = False
    is_delivered: bool = False
    ordered_at: Optional[datetime] = None
    ordered_by: Optional[str] = None
    delivered_at: Optional[datetime] = None
    delivered_by: Optional[str] = None
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MaterialRequestCreate(BaseModel):
    title: str
    quantity: str
    needed_by: str
    photo_url: Optional[str] = None
    notes: Optional[str] = None
    project_id: Optional[str] = None
    project_name: Optional[str] = None

# Material Catalog (Beheerder materialenlijst voor werkmannen)
class MaterialCategory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"MCTG-{str(uuid.uuid4())[:8].upper()}")
    name: str
    name_ua: Optional[str] = None
    sort_order: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MaterialCatalogItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"MCAT-{str(uuid.uuid4())[:8].upper()}")
    category_id: Optional[str] = None
    title: str
    title_ua: Optional[str] = None
    description: Optional[str] = None
    image_url: Optional[str] = None
    sizes: List[str] = []  # e.g. ["60x60", "30x60", "80x80"]
    active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MaterialCatalogItemCreate(BaseModel):
    title: str
    title_ua: Optional[str] = None
    category_id: Optional[str] = None
    description: Optional[str] = None
    sizes: List[str] = []

class MaterialCatalogItemUpdate(BaseModel):
    title: Optional[str] = None
    title_ua: Optional[str] = None
    category_id: Optional[str] = None
    description: Optional[str] = None
    sizes: Optional[List[str]] = None
    active: Optional[bool] = None

class MaterialOrderCreate(BaseModel):
    items: List[dict]  # [{catalog_item_id, title, selected_size, quantity, image_url}]
    project_id: str
    project_name: str
    notes: Optional[str] = None
    delivery_date: Optional[str] = None

# Subcontractor (Onderaannemer)
class Subcontractor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"SUB-{str(uuid.uuid4())[:8].upper()}")
    user_id: Optional[str] = None  # Gekoppelde user voor login
    
    company_name: str
    contact_name: str
    email: str
    phone: str = ""
    vat_number: str = ""
    
    category: str  # "dak" | "ramen" | "metselwerk" | "gevel" | "isolatie" | "elektriciteit" | "sanitair"
    
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SubcontractorCreate(BaseModel):
    company_name: str
    contact_name: str
    email: str
    phone: str = ""
    vat_number: str = ""
    category: str
    password: Optional[str] = None  # Als login nodig is

# Subcontractor Price
class SubcontractorPrice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    subcontractor_id: str
    
    title: str
    category: str
    
    price_type: str  # "forfait" | "per_m2" | "per_lm" | "per_stuk"
    price: float
    price_min: Optional[float] = None
    price_max: Optional[float] = None
    
    valid_from: Optional[str] = None
    valid_until: Optional[str] = None
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SubcontractorPriceCreate(BaseModel):
    title: str
    category: str
    price_type: str
    price: float
    price_min: Optional[float] = None
    price_max: Optional[float] = None

# ============= MAINTENANCE MODELS =============

# Onderhoudscontract/dossier
class MaintenanceContract(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"MAINT-{str(uuid.uuid4())[:8].upper()}")
    
    # Klantgegevens (direct, geen lead nodig)
    client_name: str
    client_email: str = ""
    client_phone: str = ""
    client_address: str = ""
    client_postal_code: str = ""
    client_city: str = ""
    
    # Type onderhoud
    maintenance_type: str  # "verwarming" | "ventilatie" | "waterfilter"
    
    # Omschrijving
    description: str = ""
    
    # Planning
    scheduled_date: Optional[str] = None  # Geplande onderhoudsdatum
    last_maintenance_date: Optional[str] = None
    next_maintenance_date: Optional[str] = None
    
    # Frequentie (voor terugkerend onderhoud)
    frequency_months: int = 12  # Elke X maanden
    
    # Status
    status: str = "gepland"  # "gepland" | "uitgevoerd" | "gefactureerd" | "geannuleerd"
    
    # Prijzen
    service_price: float = 0.0  # Onderhoudsprijs
    materials_cost: float = 0.0  # Materiaalkost (berekend uit purchases)
    
    # Notities
    notes: str = ""
    technician_notes: str = ""  # Technische opmerkingen na onderhoud
    
    # Gekoppelde facturen
    invoice_ids: List[str] = []
    
    created_by: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MaintenanceContractCreate(BaseModel):
    client_name: str
    client_email: str = ""
    client_phone: str = ""
    client_address: str = ""
    client_postal_code: str = ""
    client_city: str = ""
    maintenance_type: str
    description: str = ""
    scheduled_date: Optional[str] = None
    frequency_months: int = 12
    service_price: float = 0.0
    notes: str = ""

class MaintenanceContractUpdate(BaseModel):
    client_name: Optional[str] = None
    client_email: Optional[str] = None
    client_phone: Optional[str] = None
    client_address: Optional[str] = None
    client_postal_code: Optional[str] = None
    client_city: Optional[str] = None
    maintenance_type: Optional[str] = None
    description: Optional[str] = None
    scheduled_date: Optional[str] = None
    last_maintenance_date: Optional[str] = None
    next_maintenance_date: Optional[str] = None
    frequency_months: Optional[int] = None
    status: Optional[str] = None
    service_price: Optional[float] = None
    notes: Optional[str] = None
    technician_notes: Optional[str] = None

# Aankoopfactuur voor onderhoud (klein materiaal)
class MaintenancePurchase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    maintenance_id: str  # Gekoppeld aan onderhoudsdossier
    
    supplier: str  # Leverancier
    invoice_number: str = ""
    invoice_date: str = ""
    
    description: str
    amount: float  # Bedrag excl BTW
    vat_amount: float = 0.0  # BTW bedrag
    total_amount: float = 0.0  # Totaal incl BTW
    
    # Optioneel: foto van factuur
    invoice_photo: Optional[str] = None  # Base64 of URL
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MaintenancePurchaseCreate(BaseModel):
    supplier: str
    invoice_number: str = ""
    invoice_date: str = ""
    description: str
    amount: float
    vat_amount: float = 0.0

# Onderhoudsfactuur (verkoopfactuur)
class MaintenanceInvoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"MINV-{str(uuid.uuid4())[:8].upper()}")
    maintenance_id: str
    
    invoice_number: str = ""
    invoice_date: str = ""
    due_date: str = ""
    
    # Bedragen
    service_amount: float = 0.0  # Onderhoudskost
    materials_amount: float = 0.0  # Doorgerekende materialen
    subtotal: float = 0.0
    vat_rate: float = 21.0
    vat_amount: float = 0.0
    total_amount: float = 0.0
    
    # Status
    status: str = "concept"  # "concept" | "verstuurd" | "betaald"
    paid_date: Optional[str] = None
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MaintenanceInvoiceCreate(BaseModel):
    service_amount: float
    materials_amount: float = 0.0
    vat_rate: float = 21.0
    invoice_date: Optional[str] = None
    due_date: Optional[str] = None

# Realtor Profile (Makelaar)
class RealtorProfile(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"REALTOR-{str(uuid.uuid4())[:8].upper()}")
    user_id: str
    
    company_name: str
    contact_name: str
    email: str
    phone: str = ""
    
    property_limit: int = 5  # Gratis tier
    properties_used: int = 0
    subscription_tier: str = "free"  # "free" | "basic" | "pro"
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RealtorCreate(BaseModel):
    company_name: str
    contact_name: str
    email: str
    phone: str = ""
    username: str
    password: str

# Investor Profile
class InvestorProfile(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"INVESTOR-{str(uuid.uuid4())[:8].upper()}")
    user_id: str
    
    name: str
    email: str
    phone: str = ""
    
    target_roi: float = 10.0  # Gewenst rendement %
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InvestorCreate(BaseModel):
    name: str
    email: str
    phone: str = ""
    username: str
    password: str
    target_roi: float = 10.0

# WorkItem uitbreiding met component labels
class WorkItemLabel(BaseModel):
    component_label: str  # "vloer" | "muur" | "plafond" | "elektriciteit" | "sanitair" | "verwarming" | "isolatie" | "overig"
    room_types: List[str] = ["all"]  # ["all"] of ["bathroom", "kitchen"] etc.

# ============= AUTH DEPENDENCIES =============

# Hardcoded admin users (same as auth_simple.py)
HARDCODED_ADMINS = {
    "ADMIN-LIAM": {
        "id": "ADMIN-LIAM",
        "username": "Liam",
        "email": "liam.waerzeggers@qtechnics.be",
        "name": "Liam",
        "role": "admin"
    }
}

async def get_current_user(session_token: Optional[str] = Cookie(None), authorization: Optional[str] = Header(None)) -> User:
    """Get current user from session token (cookie or header)"""
    token = session_token
    
    if not token and authorization:
        parts = authorization.split()
        if len(parts) == 2 and parts[0].lower() == "bearer":
            token = parts[1]
    
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Find session - check both user_sessions (for admins) and sessions (for workers)
    session = await db.user_sessions.find_one({"session_token": token})
    if not session:
        session = await db.sessions.find_one({"session_token": token})
    
    logger.info(f"Looking for session with token: {token[:10]}... Found: {session is not None}")
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    # Check expiry - handle both string and datetime
    expires_at = session["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    
    # Ensure expires_at is timezone-aware
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    
    if expires_at < datetime.now(timezone.utc):
        # Delete from both collections
        await db.user_sessions.delete_one({"session_token": token})
        await db.sessions.delete_one({"session_token": token})
        raise HTTPException(status_code=401, detail="Session expired")
    
    # Find user
    user_id = session["user_id"]
    
    # Check hardcoded admins first
    if user_id in HARDCODED_ADMINS:
        admin = HARDCODED_ADMINS[user_id]
        user_dict = {
            "_id": admin["id"],
            "email": admin["email"],
            "username": admin["username"],
            "name": admin["name"],
            "role": admin["role"]
        }
        return User(**user_dict)
    
    # Try to find in database - first by id field
    user_doc = await db.users.find_one({"id": user_id})
    
    if not user_doc:
        # Try by _id
        user_doc = await db.users.find_one({"_id": user_id})
    
    if not user_doc:
        # Check if it's a worker
        worker_doc = await db.workers.find_one({"id": user_id}, {"_id": 0})
        if worker_doc:
            user_doc = {
                "id": worker_doc["id"],
                "username": worker_doc.get("username"),
                "email": worker_doc.get("username") + "@worker.local",
                "name": worker_doc.get("name", ""),
                "role": "worker",
                "created_at": worker_doc.get("created_at", datetime.now(timezone.utc).isoformat())
            }
        else:
            raise HTTPException(status_code=404, detail="User not found")
    
    return User(**user_doc)

# ============= AUTH ROUTES =============

@api_router.post("/auth/session", response_model=SessionResponse)
async def create_session(request: SessionRequest, response: Response):
    """Process session_id from Emergent Auth and create user session"""
    
    # Call Emergent Auth API to get user data
    async with httpx.AsyncClient() as client:
        try:
            auth_response = await client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": request.session_id}
            )
            auth_response.raise_for_status()
            user_data = auth_response.json()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to validate session: {str(e)}")
    
    # Check if user exists
    existing_user = await db.users.find_one({"_id": user_data["email"]})
    
    if not existing_user:
        # Create new user
        user_doc = {
            "_id": user_data["email"],
            "email": user_data["email"],
            "name": user_data.get("name", ""),
            "picture": user_data.get("picture", ""),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(user_doc)
        user = User(**user_doc)
    else:
        user = User(**existing_user)
    
    # Create session
    session_token = user_data.get("session_token") or str(uuid.uuid4())
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    session_doc = {
        "user_id": user.id,
        "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.user_sessions.insert_one(session_doc)
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=7*24*60*60,
        path="/"
    )
    
    return SessionResponse(user=user, session_token=session_token)

@api_router.get("/auth/me")
async def get_me(current_user: User = Depends(get_current_user)):
    """Get current authenticated user"""
    # Return user data with 'id' instead of '_id' for frontend compatibility
    return {
        "id": current_user.id,
        "email": current_user.email,
        "username": current_user.username,
        "name": current_user.name,
        "picture": current_user.picture,
        "role": current_user.role,
        "created_at": current_user.created_at.isoformat() if isinstance(current_user.created_at, datetime) else current_user.created_at
    }

@api_router.post("/auth/logout")
async def logout(response: Response, current_user: User = Depends(get_current_user), session_token: Optional[str] = Cookie(None)):
    """Logout user and delete session"""
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    
    response.delete_cookie(key="session_token", path="/")
    return {"message": "Logged out successfully"}

# ============= WEBHOOK ROUTES (PUBLIC - API KEY AUTH) =============

# API Key for webhook authentication (stored in environment or database)
WEBHOOK_API_KEY = os.environ.get('WEBHOOK_API_KEY', 'qtechnics-webhook-2024-secure-key')

async def verify_webhook_api_key(x_api_key: str = Header(None, alias="X-API-Key")):
    """Verify the API key for webhook requests"""
    if not x_api_key:
        raise HTTPException(status_code=401, detail="API Key ontbreekt. Voeg X-API-Key header toe.")
    if x_api_key != WEBHOOK_API_KEY:
        raise HTTPException(status_code=403, detail="Ongeldige API Key")
    return True

@api_router.post("/webhook/lead")
async def webhook_create_lead(
    lead_data: WebsiteLeadWebhook,
    api_key_valid: bool = Depends(verify_webhook_api_key)
):
    """
    PUBLIC WEBHOOK: Create a lead from external website form submission.
    
    Authentication: X-API-Key header required
    
    This endpoint:
    1. Creates a new lead with all form data
    2. Automatically creates a project
    3. Stores all extra form data in the lead description
    """
    try:
        # Build full address if components provided
        full_address = lead_data.address or ""
        if lead_data.postal_code or lead_data.city:
            address_parts = [lead_data.address or ""]
            if lead_data.postal_code:
                address_parts.append(lead_data.postal_code)
            if lead_data.city:
                address_parts.append(lead_data.city)
            full_address = ", ".join(filter(None, address_parts))
        
        # Build comprehensive description with all form data
        description_parts = []
        
        if lead_data.message:
            description_parts.append(f"📝 Bericht:\n{lead_data.message}")
        
        if lead_data.description:
            description_parts.append(f"📋 Omschrijving:\n{lead_data.description}")
        
        if lead_data.company_name:
            description_parts.append(f"🏢 Bedrijf: {lead_data.company_name}")
        
        if lead_data.source:
            description_parts.append(f"🌐 Bron: {lead_data.source}")
        
        if lead_data.form_name:
            description_parts.append(f"📄 Formulier: {lead_data.form_name}")
        
        if lead_data.page_url:
            description_parts.append(f"🔗 Pagina: {lead_data.page_url}")
        
        # Add any extra data fields
        if lead_data.extra_data:
            extra_info = []
            for key, value in lead_data.extra_data.items():
                if value:
                    extra_info.append(f"• {key}: {value}")
            if extra_info:
                description_parts.append(f"📎 Extra info:\n" + "\n".join(extra_info))
        
        full_description = "\n\n".join(description_parts)
        
        # Determine if business
        is_business = lead_data.is_business or bool(lead_data.vat_number) or bool(lead_data.company_name)
        
        # Create the lead
        lead_id = f"LEAD-{str(uuid.uuid4())[:8].upper()}"
        lead_doc = {
            "id": lead_id,
            "name": lead_data.company_name if lead_data.company_name else lead_data.name,
            "email": lead_data.email,
            "phone": lead_data.phone,
            "address": full_address or "Via website formulier",
            "project_type": lead_data.project_type or "Website aanvraag",
            "description": full_description,
            "vat_number": lead_data.vat_number,
            "is_business": is_business,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "user_id": "website-webhook",  # System user for webhook leads
            "source": lead_data.source or "website",
            "contact_person": lead_data.name if lead_data.company_name else None,  # Store contact if company
        }
        
        await db.leads.insert_one(lead_doc)
        logger.info(f"Webhook: Created lead {lead_id} from {lead_data.source}")
        
        # AUTOMATICALLY CREATE PROJECT
        project_id = f"PROJ-{str(uuid.uuid4())[:8].upper()}"
        project_name = f"Project - {lead_data.company_name or lead_data.name}"
        if lead_data.city:
            project_name += f" ({lead_data.city})"
        
        project_doc = {
            "id": project_id,
            "lead_id": lead_id,
            "name": project_name,
            "status": "eerste bezoek",
            "first_visit_date": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "user_id": "website-webhook",
            "notes": f"Automatisch aangemaakt via website formulier.\n\n{full_description}"
        }
        
        await db.projects.insert_one(project_doc)
        logger.info(f"Webhook: Created project {project_id} for lead {lead_id}")
        
        # CREATE CELEBRATION for new website lead
        celebration_id = f"CELEB-{str(uuid.uuid4())[:8].upper()}"
        celebration = {
            "id": celebration_id,
            "type": "new_website_lead",  # Different type for website leads
            "lead_id": lead_id,
            "project_id": project_id,
            "lead_name": lead_data.company_name or lead_data.name,
            "project_name": project_name,
            "project_type": lead_data.project_type or "Website aanvraag",
            "city": lead_data.city,
            "source": lead_data.source or "website",
            "form_name": lead_data.form_name,
            "message_preview": (lead_data.message or lead_data.description or "")[:100],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "seen_by": []  # Track which admins have seen this
        }
        await db.celebrations.insert_one(celebration)
        logger.info(f"Webhook: Created celebration {celebration_id} for new lead")
        
        return {
            "success": True,
            "message": "Lead en project succesvol aangemaakt",
            "lead_id": lead_id,
            "project_id": project_id
        }
        
    except Exception as e:
        logger.error(f"Webhook error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Fout bij aanmaken lead: {str(e)}")

@api_router.get("/webhook/test")
async def webhook_test(api_key_valid: bool = Depends(verify_webhook_api_key)):
    """Test endpoint to verify API key is working"""
    return {
        "success": True,
        "message": "API Key is geldig! Webhook verbinding werkt.",
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

# ============= LEAD ROUTES =============

@api_router.post("/leads", response_model=Lead)
async def create_lead(lead: LeadCreate, current_user: User = Depends(get_current_user)):
    """Create a new lead and automatically create a project"""
    lead_obj = Lead(**lead.model_dump(), user_id=current_user.id)
    lead_doc = lead_obj.model_dump()
    lead_doc["created_at"] = lead_doc["created_at"].isoformat()
    
    await db.leads.insert_one(lead_doc)
    
    # AUTOMATISCH PROJECT AANMAKEN
    project = Project(
        lead_id=lead_obj.id,
        name=f"Project - {lead_obj.name}",
        status="eerste bezoek",
        first_visit_date=datetime.now(timezone.utc),
        user_id=current_user.id
    )
    
    project_doc = project.model_dump()
    project_doc["created_at"] = project_doc["created_at"].isoformat()
    if project_doc.get("first_visit_date"):
        project_doc["first_visit_date"] = project_doc["first_visit_date"].isoformat()
    
    await db.projects.insert_one(project_doc)
    
    return lead_obj

@api_router.get("/leads", response_model=List[Lead])
async def get_leads(current_user: User = Depends(get_current_user)):
    """Get all leads (all admins see all data)"""
    # All admins see all leads, workers see nothing
    if current_user.role == "admin":
        leads = await db.leads.find({}, {"_id": 0}).to_list(1000)
    else:
        leads = []
    
    for lead in leads:
        if isinstance(lead["created_at"], str):
            lead["created_at"] = datetime.fromisoformat(lead["created_at"])
    
    return leads

@api_router.get("/leads/{lead_id}", response_model=Lead)
async def get_lead(lead_id: str, current_user: User = Depends(get_current_user)):
    """Get a specific lead (all admins can access)"""
    # All admins can see all leads
    if current_user.role == "admin":
        lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    else:
        lead = None
    
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    if isinstance(lead["created_at"], str):
        lead["created_at"] = datetime.fromisoformat(lead["created_at"])
    
    return Lead(**lead)

@api_router.put("/leads/{lead_id}", response_model=Lead)
async def update_lead(lead_id: str, lead_update: LeadUpdate, current_user: User = Depends(get_current_user)):
    """Update a lead (all admins can edit)"""
    # All admins can update any lead
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update leads")
    
    existing_lead = await db.leads.find_one({"id": lead_id})
    if not existing_lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    update_data = {k: v for k, v in lead_update.model_dump().items() if v is not None}
    
    if update_data:
        await db.leads.update_one({"id": lead_id}, {"$set": update_data})
        existing_lead.update(update_data)
    
    if isinstance(existing_lead["created_at"], str):
        existing_lead["created_at"] = datetime.fromisoformat(existing_lead["created_at"])
    
    return Lead(**existing_lead)

@api_router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str, current_user: User = Depends(get_current_user)):
    """Delete a lead (all admins can delete)"""
    # All admins can delete any lead
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete leads")
    
    result = await db.leads.delete_one({"id": lead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    return {"message": "Lead deleted successfully"}

# ============= QUOTE ROUTES =============

@api_router.post("/quotes", response_model=Quote)
async def create_quote(quote_create: QuoteCreate, current_user: User = Depends(get_current_user)):
    """Create a new quote from a lead"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can create quotes")
    
    # All admins can create quotes for any lead
    lead = await db.leads.find_one({"id": quote_create.lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    quote_obj = Quote(**quote_create.model_dump(), user_id=current_user.id)
    quote_obj.quote_number = quote_obj.id
    
    quote_doc = quote_obj.model_dump()
    quote_doc["date"] = quote_doc["date"].isoformat()
    quote_doc["created_at"] = quote_doc["created_at"].isoformat()
    
    await db.quotes.insert_one(quote_doc)
    return quote_obj

@api_router.get("/quotes", response_model=List[Quote])
async def get_quotes(current_user: User = Depends(get_current_user)):
    """Get all quotes (all admins see all data)"""
    # All admins see all quotes, workers see nothing
    if current_user.role == "admin":
        quotes = await db.quotes.find({}, {"_id": 0}).to_list(1000)
    else:
        quotes = []
    
    for quote in quotes:
        if quote.get("date") and isinstance(quote["date"], str):
            quote["date"] = datetime.fromisoformat(quote["date"])
        if quote.get("created_at") and isinstance(quote["created_at"], str):
            quote["created_at"] = datetime.fromisoformat(quote["created_at"])
    
    return quotes

@api_router.get("/quotes/{quote_id}", response_model=Quote)
async def get_quote(quote_id: str, current_user: User = Depends(get_current_user)):
    """Get a specific quote (all admins can access)"""
    # All admins can see all quotes
    if current_user.role == "admin":
        quote = await db.quotes.find_one({"id": quote_id}, {"_id": 0})
    else:
        quote = None
    
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    if isinstance(quote["date"], str):
        quote["date"] = datetime.fromisoformat(quote["date"])
    if isinstance(quote["created_at"], str):
        quote["created_at"] = datetime.fromisoformat(quote["created_at"])
    
    return Quote(**quote)

@api_router.put("/quotes/{quote_id}", response_model=Quote)
async def update_quote(quote_id: str, quote_update: QuoteUpdate, current_user: User = Depends(get_current_user)):
    """Update a quote (all admins can edit)"""
    # All admins can update any quote
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update quotes")
    
    existing_quote = await db.quotes.find_one({"id": quote_id})
    if not existing_quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    update_data = {k: v for k, v in quote_update.model_dump(exclude_unset=True).items()}
    
    if update_data:
        await db.quotes.update_one({"id": quote_id}, {"$set": update_data})
        existing_quote.update(update_data)
        
        lead_id = existing_quote.get("lead_id")
        project = None
        if lead_id:
            project = await db.projects.find_one({"lead_id": lead_id})
        
        # If quote is marked as sold, update project status and create celebration
        if update_data.get("is_sold") == True and project:
            # Update project status to "in uitvoering"
            await db.projects.update_one(
                {"id": project["id"]},
                {"$set": {"status": "in uitvoering"}}
            )
            
            # Create celebration record for all admins
            celebration = {
                "id": f"CELEB-{str(uuid.uuid4())[:8].upper()}",
                "project_id": project["id"],
                "project_name": project.get("name", "Onbekend project"),
                "quote_id": quote_id,
                "quote_number": existing_quote.get("quote_number", ""),
                "amount": existing_quote.get("total_incl_vat", 0),
                "sold_at": datetime.now(timezone.utc).isoformat(),
                "sold_by": current_user.id,
                "seen_by": []  # Track which users have seen this celebration
            }
            await db.celebrations.insert_one(celebration)
            logger.info(f"Quote {quote_id} marked as sold, project {project['id']} status set to 'in uitvoering'")
        
        # If status changed to "goedgekeurd", update project financials
        if update_data.get("status") == "goedgekeurd" and project:
            # Calculate total sales from all approved quotes for this lead
            approved_quotes = await db.quotes.find({
                "lead_id": lead_id,
                "status": "goedgekeurd"
            }, {"_id": 0, "total_incl_vat": 1}).to_list(100)
            
            total_sales = sum(q.get("total_incl_vat", 0) for q in approved_quotes)
            total_costs = project.get("total_costs", 0)
            profit = total_sales - total_costs
            
            # Update project with new sales total
            await db.projects.update_one(
                {"id": project["id"]},
                {"$set": {
                    "sales_price": total_sales,
                    "profit": profit
                }}
            )
            logger.info(f"Updated project {project['id']} financials: sales={total_sales}, profit={profit}")
    
    if isinstance(existing_quote["date"], str):
        existing_quote["date"] = datetime.fromisoformat(existing_quote["date"])
    if isinstance(existing_quote["created_at"], str):
        existing_quote["created_at"] = datetime.fromisoformat(existing_quote["created_at"])
    
    return Quote(**existing_quote)

@api_router.delete("/quotes/{quote_id}")
async def delete_quote(quote_id: str, current_user: User = Depends(get_current_user)):
    """Delete a quote (all admins can delete)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete quotes")
    
    result = await db.quotes.delete_one({"id": quote_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    # Also delete associated line items
    await db.line_items.delete_many({"quote_id": quote_id})
    
    return {"message": "Quote deleted successfully"}

@api_router.post("/quotes/{quote_id}/split")
async def split_quote_labor_materials(quote_id: str, current_user: User = Depends(get_current_user)):
    """Split a quote into two separate quotes: one for labor (arbeid) and one for materials (materialen)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can split quotes")
    
    # Get original quote
    original_quote = await db.quotes.find_one({"id": quote_id})
    if not original_quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    # Get all line items
    all_items = await db.line_items.find({"quote_id": quote_id}, {"_id": 0}).to_list(1000)
    
    labor_items = [item for item in all_items if item.get('item_type') == 'arbeid']
    material_items = [item for item in all_items if item.get('item_type') != 'arbeid']
    
    if not labor_items and not material_items:
        raise HTTPException(status_code=400, detail="Offerte heeft geen items om te splitsen")
    
    if not labor_items:
        raise HTTPException(status_code=400, detail="Offerte heeft geen arbeid items - splitsen niet nodig")
    
    if not material_items:
        raise HTTPException(status_code=400, detail="Offerte heeft geen materiaal items - splitsen niet nodig")
    
    created_quotes = []
    
    # Create LABOR quote (Arbeid)
    labor_quote_id = f"OFF-{datetime.now(timezone.utc).year}-{str(uuid.uuid4())[:6].upper()}-ARB"
    
    labor_total_excl = sum(item.get('total_excl_vat', 0) for item in labor_items)
    labor_vat_6 = sum(item.get('vat_amount', 0) for item in labor_items if item.get('vat_rate', 6) == 6)
    labor_vat_21 = sum(item.get('vat_amount', 0) for item in labor_items if item.get('vat_rate', 21) == 21)
    labor_total_vat = labor_vat_6 + labor_vat_21
    labor_total_incl = labor_total_excl + labor_total_vat
    
    labor_vat_breakdown = {}
    if labor_vat_6 > 0:
        labor_vat_breakdown["6"] = labor_vat_6
    if labor_vat_21 > 0:
        labor_vat_breakdown["21"] = labor_vat_21
    
    labor_quote = {
        "id": labor_quote_id,
        "lead_id": original_quote["lead_id"],
        "project_id": original_quote.get("project_id"),
        "quote_number": labor_quote_id,
        "date": datetime.now(timezone.utc).isoformat(),
        "status": "concept",
        "description": f"Arbeid - Afgesplitst van {original_quote['quote_number']}",
        "subtotal_labor": labor_total_excl,
        "subtotal_material": 0,
        "total_excl_vat": labor_total_excl,
        "vat_breakdown": labor_vat_breakdown,
        "total_vat": labor_total_vat,
        "total_incl_vat": labor_total_incl,
        "total_price": labor_total_incl,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "user_id": current_user.id,
        "split_from": quote_id,
        "quote_type": "arbeid"
    }
    
    await db.quotes.insert_one(labor_quote)
    
    # Copy labor line items to new quote
    for item in labor_items:
        new_item = {**item}
        new_item["id"] = str(uuid.uuid4())
        new_item["quote_id"] = labor_quote_id
        new_item["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.line_items.insert_one(new_item)
    
    created_quotes.append({
        "id": labor_quote_id,
        "type": "arbeid",
        "total_incl_vat": labor_total_incl,
        "items_count": len(labor_items)
    })
    
    # Create MATERIALS quote (Materialen)
    material_quote_id = f"OFF-{datetime.now(timezone.utc).year}-{str(uuid.uuid4())[:6].upper()}-MAT"
    
    material_total_excl = sum(item.get('total_excl_vat', 0) for item in material_items)
    material_vat_6 = sum(item.get('vat_amount', 0) for item in material_items if item.get('vat_rate', 6) == 6)
    material_vat_21 = sum(item.get('vat_amount', 0) for item in material_items if item.get('vat_rate', 21) == 21)
    material_total_vat = material_vat_6 + material_vat_21
    material_total_incl = material_total_excl + material_total_vat
    
    material_vat_breakdown = {}
    if material_vat_6 > 0:
        material_vat_breakdown["6"] = material_vat_6
    if material_vat_21 > 0:
        material_vat_breakdown["21"] = material_vat_21
    
    material_quote = {
        "id": material_quote_id,
        "lead_id": original_quote["lead_id"],
        "project_id": original_quote.get("project_id"),
        "quote_number": material_quote_id,
        "date": datetime.now(timezone.utc).isoformat(),
        "status": "concept",
        "description": f"Materialen - Afgesplitst van {original_quote['quote_number']}",
        "subtotal_labor": 0,
        "subtotal_material": material_total_excl,
        "total_excl_vat": material_total_excl,
        "vat_breakdown": material_vat_breakdown,
        "total_vat": material_total_vat,
        "total_incl_vat": material_total_incl,
        "total_price": material_total_incl,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "user_id": current_user.id,
        "split_from": quote_id,
        "quote_type": "materialen"
    }
    
    await db.quotes.insert_one(material_quote)
    
    # Copy material line items to new quote
    for item in material_items:
        new_item = {**item}
        new_item["id"] = str(uuid.uuid4())
        new_item["quote_id"] = material_quote_id
        new_item["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.line_items.insert_one(new_item)
    
    created_quotes.append({
        "id": material_quote_id,
        "type": "materialen",
        "total_incl_vat": material_total_incl,
        "items_count": len(material_items)
    })
    
    # Mark original quote as split
    await db.quotes.update_one(
        {"id": quote_id},
        {"$set": {
            "status": "gesplitst",
            "split_into": [labor_quote_id, material_quote_id]
        }}
    )
    
    return {
        "message": "Offerte succesvol gesplitst in Arbeid en Materialen",
        "original_quote_id": quote_id,
        "created_quotes": created_quotes
    }

# ============= LINE ITEM ROUTES =============

@api_router.post("/quotes/{quote_id}/items", response_model=LineItem)
async def add_line_item(quote_id: str, item: LineItemCreate, current_user: User = Depends(get_current_user)):
    """Add a line item to a quote (all admins can add)"""
    # All admins can add items to any quote
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can add line items")
    
    quote = await db.quotes.find_one({"id": quote_id})
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    # Create line item with VAT calculations
    item_obj = LineItem(**item.model_dump(), quote_id=quote_id)
    item_obj.total_excl_vat = item_obj.quantity * item_obj.unit_price
    item_obj.vat_amount = item_obj.total_excl_vat * (item_obj.vat_rate / 100)
    item_obj.total_incl_vat = item_obj.total_excl_vat + item_obj.vat_amount
    item_obj.total = item_obj.total_incl_vat  # Backwards compatibility
    
    item_doc = item_obj.model_dump()
    item_doc["created_at"] = item_doc["created_at"].isoformat()
    
    await db.line_items.insert_one(item_doc)
    
    # Recalculate quote totals
    await recalculate_quote_totals(quote_id)
    
    return item_obj

@api_router.get("/quotes/{quote_id}/items", response_model=List[LineItem])
async def get_line_items(quote_id: str, current_user: User = Depends(get_current_user)):
    """Get all line items for a quote (all admins can access)"""
    # All admins can see all quote items
    if current_user.role == "admin":
        quote = await db.quotes.find_one({"id": quote_id})
    else:
        quote = None
    
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    items = await db.line_items.find({"quote_id": quote_id}, {"_id": 0}).to_list(1000)
    
    for item in items:
        if isinstance(item["created_at"], str):
            item["created_at"] = datetime.fromisoformat(item["created_at"])
    
    return items

@api_router.put("/quotes/{quote_id}/items/{item_id}", response_model=LineItem)
async def update_line_item(quote_id: str, item_id: str, item_update: LineItemUpdate, current_user: User = Depends(get_current_user)):
    """Update a line item (all admins can edit)"""
    # All admins can update any quote item
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update line items")
    
    quote = await db.quotes.find_one({"id": quote_id})
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    existing_item = await db.line_items.find_one({"id": item_id, "quote_id": quote_id})
    if not existing_item:
        raise HTTPException(status_code=404, detail="Line item not found")
    
    update_data = {k: v for k, v in item_update.model_dump().items() if v is not None}
    
    if update_data:
        # Recalculate totals if quantity, price, or vat_rate changed
        if "quantity" in update_data or "unit_price" in update_data or "vat_rate" in update_data:
            quantity = update_data.get("quantity", existing_item["quantity"])
            unit_price = update_data.get("unit_price", existing_item["unit_price"])
            vat_rate = update_data.get("vat_rate", existing_item.get("vat_rate", 21))
            
            # Calculate all totals correctly
            total_excl_vat = quantity * unit_price
            vat_amount = total_excl_vat * (vat_rate / 100)
            total_incl_vat = total_excl_vat + vat_amount
            
            update_data["total"] = total_excl_vat  # Base total (excl VAT)
            update_data["total_excl_vat"] = total_excl_vat
            update_data["vat_amount"] = vat_amount
            update_data["total_incl_vat"] = total_incl_vat
        
        await db.line_items.update_one({"id": item_id}, {"$set": update_data})
        existing_item.update(update_data)
        
        # Recalculate quote totals
        await recalculate_quote_totals(quote_id)
    
    if isinstance(existing_item["created_at"], str):
        existing_item["created_at"] = datetime.fromisoformat(existing_item["created_at"])
    
    return LineItem(**existing_item)

@api_router.delete("/quotes/{quote_id}/items/{item_id}")
async def delete_line_item(quote_id: str, item_id: str, current_user: User = Depends(get_current_user)):
    """Delete a line item (all admins can delete)"""
    # All admins can delete any quote item
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete line items")
    
    quote = await db.quotes.find_one({"id": quote_id})
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    result = await db.line_items.delete_one({"id": item_id, "quote_id": quote_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Line item not found")
    
    # Recalculate quote totals
    await recalculate_quote_totals(quote_id)
    
    return {"message": "Line item deleted successfully"}

async def recalculate_quote_totals(quote_id: str):
    """Recalculate quote subtotals, VAT breakdown and total price"""
    items = await db.line_items.find({"quote_id": quote_id}, {"_id": 0}).to_list(1000)
    
    # Calculate subtotals by type (excl VAT)
    subtotal_labor = sum(item.get("total_excl_vat", item.get("total", 0)) for item in items if item["item_type"] == "arbeid")
    subtotal_material = sum(item.get("total_excl_vat", item.get("total", 0)) for item in items if item["item_type"] == "materiaal")
    other_total = sum(item.get("total_excl_vat", item.get("total", 0)) for item in items if item["item_type"] == "overig")
    
    total_excl_vat = subtotal_labor + subtotal_material + other_total
    
    # Calculate VAT breakdown by rate
    vat_breakdown = {}
    total_vat = 0.0
    
    for item in items:
        vat_rate = str(item.get("vat_rate", 21))
        vat_amount = item.get("vat_amount", 0)
        
        if vat_rate not in vat_breakdown:
            vat_breakdown[vat_rate] = 0.0
        vat_breakdown[vat_rate] += vat_amount
        total_vat += vat_amount
    
    total_incl_vat = total_excl_vat + total_vat
    
    await db.quotes.update_one(
        {"id": quote_id},
        {"$set": {
            "subtotal_labor": subtotal_labor,
            "subtotal_material": subtotal_material,
            "total_excl_vat": total_excl_vat,
            "vat_breakdown": vat_breakdown,
            "total_vat": total_vat,
            "total_incl_vat": total_incl_vat,
            "total_price": total_incl_vat  # For backwards compatibility
        }}
    )

# ============= MATERIAL ROUTES =============

@api_router.post("/materials/upload")
async def upload_materials_csv(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    """Upload materials CSV and replace existing catalog"""
    logger.info(f"Uploading materials CSV: {file.filename} for user {current_user.id}")
    
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Alleen CSV bestanden zijn toegestaan")
    
    try:
        # Read CSV
        content = await file.read()
        logger.info(f"CSV file size: {len(content)} bytes")
        
        df = pd.read_csv(io.BytesIO(content))
        logger.info(f"CSV loaded with {len(df)} rows and columns: {list(df.columns)}")
        
        # Map common CSV column names to expected format
        column_mapping = {
            # ECK format
            'Artikelcode': 'sku',
            'artikelcode': 'sku',
            'ARTIKELCODE': 'sku',
            'artikelomschrijving': 'name',
            'Artikelomschrijving': 'name',
            'ARTIKELOMSCHRIJVING': 'name',
            'Tariefprijs': 'price',
            'tariefprijs': 'price',
            'TARIEFPRIJS': 'price',
            'Prijs': 'price',
            'prijs': 'price',
            'PRIJS': 'price',
            # Standard format
            'sku': 'sku',
            'SKU': 'sku',
            'name': 'name',
            'Name': 'name',
            'NAME': 'name',
            'naam': 'name',
            'Naam': 'name',
            'NAAM': 'name',
            'price': 'price',
            'Price': 'price',
            'PRICE': 'price',
            # Optional fields
            'EAN': 'ean',
            'ean': 'ean',
            'description': 'description',
            'Description': 'description',
            'beschrijving': 'description',
            'Beschrijving': 'description',
            'category': 'category',
            'Category': 'category',
            'categorie': 'category',
            'Categorie': 'category',
            'brand': 'brand',
            'Brand': 'brand',
            'merk': 'brand',
            'Merk': 'brand'
        }
        
        # Rename columns
        df = df.rename(columns=column_mapping)
        logger.info(f"Columns after mapping: {list(df.columns)}")
        
        # Validate required columns after mapping
        required_columns = ['sku', 'name', 'price']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            error_msg = f"Ontbrekende kolommen in CSV. Gevonden kolommen: {list(df.columns)}. Vereiste: artikelcode/sku, artikelomschrijving/name, tariefprijs/price"
            logger.error(error_msg)
            raise HTTPException(status_code=400, detail=error_msg)
        
        # Delete existing materials for this user
        delete_result = await db.materials.delete_many({"user_id": current_user.id})
        logger.info(f"Deleted {delete_result.deleted_count} existing materials")
        
        # Insert new materials in batches for better performance
        materials = []
        skipped = 0
        
        for idx, row in df.iterrows():
            try:
                # Validate price
                price = float(row['price'])
                if price < 0:
                    logger.warning(f"Skipping row {idx}: negative price")
                    skipped += 1
                    continue
                
                material = Material(
                    sku=str(row['sku']).strip(),
                    name=str(row['name']).strip(),
                    description=str(row.get('description', '')).strip() if pd.notna(row.get('description')) else '',
                    category=str(row.get('category', '')).strip() if pd.notna(row.get('category')) else '',
                    brand=str(row.get('brand', '')).strip() if pd.notna(row.get('brand')) else '',
                    price=price,
                    user_id=current_user.id
                )
                material_doc = material.model_dump()
                material_doc["created_at"] = material_doc["created_at"].isoformat()
                materials.append(material_doc)
                
                # Insert in batches of 1000
                if len(materials) >= 1000:
                    await db.materials.insert_many(materials)
                    logger.info(f"Inserted batch of {len(materials)} materials")
                    materials = []
                    
            except Exception as e:
                logger.warning(f"Skipping row {idx}: {str(e)}")
                skipped += 1
                continue
        
        # Insert remaining materials
        if materials:
            await db.materials.insert_many(materials)
            logger.info(f"Inserted final batch of {len(materials)} materials")
        
        total_inserted = len(df) - skipped
        logger.info(f"Upload complete: {total_inserted} materials inserted, {skipped} skipped")
        
        return {
            "message": f"Succesvol {total_inserted} materialen geüpload", 
            "count": total_inserted,
            "skipped": skipped
        }
    
    except HTTPException:
        raise
    except Exception as e:
        error_msg = f"Fout bij verwerken CSV: {str(e)}"
        logger.error(error_msg)
        raise HTTPException(status_code=400, detail=error_msg)

@api_router.get("/materials/search")
async def search_materials(q: str, current_user: User = Depends(get_current_user)):
    """Search materials by text query (all admins see all materials)"""
    if not q:
        return {"results": [], "count": 0}
    
    # Case-insensitive search on multiple fields
    # All admins see all materials
    if current_user.role == "admin":
        query = {
            "$or": [
                {"sku": {"$regex": q, "$options": "i"}},
                {"name": {"$regex": q, "$options": "i"}},
                {"description": {"$regex": q, "$options": "i"}},
                {"category": {"$regex": q, "$options": "i"}},
                {"brand": {"$regex": q, "$options": "i"}}
            ]
        }
    else:
        # Workers see nothing
        return {"results": [], "count": 0}
    
    materials = await db.materials.find(query, {"_id": 0}).limit(50).to_list(50)
    
    for material in materials:
        if isinstance(material["created_at"], str):
            material["created_at"] = datetime.fromisoformat(material["created_at"])
    
    return {"results": materials, "count": len(materials)}

@api_router.get("/materials")
async def get_materials(skip: int = 0, limit: int = 100, current_user: User = Depends(get_current_user)):
    """Get all materials with pagination (all admins see all materials)"""
    # All admins see all materials, workers see nothing
    if current_user.role == "admin":
        materials = await db.materials.find({}, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
        total = await db.materials.count_documents({})
        
        # Convert datetime strings if needed
        for material in materials:
            if isinstance(material.get("created_at"), str):
                try:
                    material["created_at"] = datetime.fromisoformat(material["created_at"])
                except:
                    pass
    else:
        materials = []
        total = 0
    
    return {"materials": materials, "total": total}


@api_router.post("/materials/auto-add")
async def auto_add_material(
    name: str,
    price: float,
    unit: str = "stuk",
    current_user: User = Depends(get_current_user)
):
    """Auto-add a material to catalog if it doesn't exist, or return existing one"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can add materials")
    
    # Check if material with same name already exists
    existing = await db.materials.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
    if existing:
        return {"material": existing, "created": False}
    
    # Create new material
    material = Material(
        name=name,
        sku=f"MAT-{uuid.uuid4().hex[:6].upper()}",
        price=price,
        unit=unit,
        user_id=current_user.id
    )
    
    await db.materials.insert_one(material.model_dump())
    
    return {"material": material.model_dump(), "created": True}


@api_router.post("/materials/create-with-image")
async def create_material_with_image(
    name: str = Form(...),
    price: float = Form(...),
    unit: str = Form("stuk"),
    file: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_user)
):
    """Create a new material with optional image upload"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can add materials")
    
    # Check if material with same name already exists
    existing = await db.materials.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
    if existing:
        # If existing material and new image provided, update the image
        if file and file.filename:
            materials_dir = ROOT_DIR / "uploads" / "materials"
            materials_dir.mkdir(parents=True, exist_ok=True)
            
            unique_filename = f"{existing['id']}_{uuid.uuid4().hex[:8]}_{file.filename}"
            file_path = materials_dir / unique_filename
            
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            image_url = f"/api/static/materials/{unique_filename}"
            await db.materials.update_one(
                {"id": existing['id']},
                {"$set": {"image_url": image_url}}
            )
            existing['image_url'] = image_url
        
        return {"material": existing, "created": False}
    
    # Create new material
    material_id = str(uuid.uuid4())
    image_url = None
    
    # Handle image upload if provided
    if file and file.filename:
        if not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="Alleen afbeeldingen zijn toegestaan")
        
        materials_dir = ROOT_DIR / "uploads" / "materials"
        materials_dir.mkdir(parents=True, exist_ok=True)
        
        unique_filename = f"{material_id}_{uuid.uuid4().hex[:8]}_{file.filename}"
        file_path = materials_dir / unique_filename
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        image_url = f"/api/static/materials/{unique_filename}"
    
    material = Material(
        id=material_id,
        name=name,
        sku=f"MAT-{uuid.uuid4().hex[:6].upper()}",
        price=price,
        unit=unit,
        image_url=image_url,
        user_id=current_user.id
    )
    
    await db.materials.insert_one(material.model_dump())
    
    return {"material": material.model_dump(), "created": True}

# ===== WORK ITEMS ENDPOINTS =====
@api_router.post("/work-items/upload")
async def upload_work_items_csv(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    """Upload work items CSV (titel, eenheid, verkoopprijs)"""
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Alleen CSV bestanden zijn toegestaan")
    
    try:
        content = await file.read()
        df = pd.read_csv(io.BytesIO(content))
        
        # Map column names
        column_mapping = {
            'titel': 'title',
            'Titel': 'title',
            'TITEL': 'title',
            'Title': 'title',
            'eenheid': 'unit',
            'Eenheid': 'unit',
            'EENHEID': 'unit',
            'Unit': 'unit',
            'verkoopprijs': 'price',
            'Verkoopprijs': 'price',
            'VERKOOPPRIJS': 'price',
            'Price': 'price',
            'Prijs': 'price',
            'prijs': 'price'
        }
        
        df = df.rename(columns=column_mapping)
        
        # Validate required columns
        required_cols = ['title', 'unit', 'price']
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            raise HTTPException(status_code=400, detail=f"Ontbrekende kolommen: {', '.join(missing)}")
        
        # Delete existing work items for this user
        await db.work_items.delete_many({"user_id": current_user.id})
        
        # Insert new work items
        work_items_to_insert = []
        for _, row in df.iterrows():
            try:
                price = float(row['price'])
                # Validate price - skip if NaN or Infinity
                import math
                if math.isnan(price) or math.isinf(price) or price < 0:
                    logger.warning(f"Skipping row with invalid price: {price}")
                    continue
                    
                work_item = WorkItem(
                    title=str(row['title']).strip(),
                    unit=str(row['unit']).strip(),
                    price=price,
                    user_id=current_user.id
                )
                work_item_doc = work_item.model_dump()
                work_item_doc['created_at'] = work_item_doc['created_at'].isoformat()
                work_items_to_insert.append(work_item_doc)
            except Exception as e:
                logger.warning(f"Skipping invalid row: {e}")
                continue
        
        if work_items_to_insert:
            await db.work_items.insert_many(work_items_to_insert)
        
        return {"message": f"{len(work_items_to_insert)} werk items geüpload", "count": len(work_items_to_insert)}
    
    except Exception as e:
        logger.error(f"Error uploading work items CSV: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/work-items/search")
async def search_work_items(q: str = "", current_user: User = Depends(get_current_user)):
    """Search work items by title"""
    if len(q) < 2:
        return []
    
    # All admins see all work items
    search_filter = {
        "$or": [
            {"title": {"$regex": q, "$options": "i"}},
            {"unit": {"$regex": q, "$options": "i"}}
        ]
    }
    
    results = await db.work_items.find(search_filter, {"_id": 0}).limit(20).to_list(20)
    
    # Clean results - remove NaN/Infinity values
    import math
    for item in results:
        if 'price' in item and (math.isnan(item['price']) or math.isinf(item['price'])):
            item['price'] = 0.0
    
    return results

@api_router.get("/work-items")
async def get_work_items(skip: int = 0, limit: int = 100, current_user: User = Depends(get_current_user)):
    """Get all work items"""
    if current_user.role == "admin":
        work_items = await db.work_items.find({}, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
        total = await db.work_items.count_documents({})
        
        # Clean work items - remove NaN/Infinity values
        import math
        for item in work_items:
            if 'price' in item and (math.isnan(item['price']) or math.isinf(item['price'])):
                item['price'] = 0.0
            if isinstance(item.get("created_at"), str):
                try:
                    item["created_at"] = datetime.fromisoformat(item["created_at"])
                except:
                    pass
    else:
        work_items = []
        total = 0
    
    return {"work_items": work_items, "total": total}

@api_router.get("/work-items/all")
async def get_all_work_items(current_user: User = Depends(get_current_user)):
    """Get ALL work items without pagination for full list view"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view all work items")
    
    work_items = await db.work_items.find({}, {"_id": 0}).to_list(10000)
    
    # Clean work items - remove NaN/Infinity values
    import math
    for item in work_items:
        if 'price' in item and (math.isnan(item['price']) or math.isinf(item['price'])):
            item['price'] = 0.0
    
    return {"work_items": work_items, "total": len(work_items)}

@api_router.post("/work-items")
async def create_work_item(
    title: str,
    unit: str,
    price: float,
    category: str = None,
    current_user: User = Depends(get_current_user)
):
    """Create a new work item manually"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can create work items")
    
    # Check if work item with same title already exists
    existing = await db.work_items.find_one({"title": title})
    if existing:
        raise HTTPException(status_code=400, detail="Werk item met deze titel bestaat al")
    
    work_item = {
        "id": str(uuid.uuid4()),
        "title": title,
        "unit": unit,
        "price": price,
        "category": category,
        "user_id": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.work_items.insert_one(work_item)
    
    # Remove _id before returning
    work_item.pop("_id", None)
    return work_item

@api_router.put("/work-items/{work_item_id}")
async def update_work_item(
    work_item_id: str,
    title: str = None,
    unit: str = None,
    price: float = None,
    current_user: User = Depends(get_current_user)
):
    """Update a work item"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update work items")
    
    # Find work item
    work_item = await db.work_items.find_one({"id": work_item_id})
    if not work_item:
        raise HTTPException(status_code=404, detail="Werk item niet gevonden")
    
    # Build update
    update_data = {}
    if title is not None:
        update_data["title"] = title
    if unit is not None:
        update_data["unit"] = unit
    if price is not None:
        update_data["price"] = price
    
    if update_data:
        await db.work_items.update_one({"id": work_item_id}, {"$set": update_data})
    
    # Return updated item
    updated = await db.work_items.find_one({"id": work_item_id}, {"_id": 0})
    return updated

@api_router.delete("/work-items/{work_item_id}")
async def delete_work_item(work_item_id: str, current_user: User = Depends(get_current_user)):
    """Delete a work item"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete work items")
    
    result = await db.work_items.delete_one({"id": work_item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Werk item niet gevonden")
    
    return {"message": "Werk item verwijderd"}

# ===== CONFIGURATOR ENDPOINTS =====

@api_router.get("/configurator/work-items")
async def get_configurator_work_items(category: str = None, current_user: User = Depends(get_current_user)):
    """Get work items for configurator, optionally filtered by category"""
    query = {}
    if category:
        query["category"] = {"$regex": category, "$options": "i"}
    
    work_items = await db.work_items.find(query, {"_id": 0}).to_list(1000)
    
    # Clean NaN/Infinity values
    import math
    for item in work_items:
        if 'price' in item and (math.isnan(item['price']) or math.isinf(item['price'])):
            item['price'] = 0.0
    
    return {"work_items": work_items}

@api_router.get("/configurator/materials")
async def get_configurator_materials(category: str = None, current_user: User = Depends(get_current_user)):
    """Get materials for configurator (products/furniture), optionally filtered by category"""
    query = {}
    if category:
        query["category"] = {"$regex": category, "$options": "i"}
    
    materials = await db.materials.find(query, {"_id": 0}).to_list(1000)
    return {"materials": materials}

@api_router.put("/materials/{material_id}")
async def update_material(
    material_id: str,
    name: str = None,
    category: str = None,
    subcategory: str = None,
    price: float = None,
    unit: str = None,
    image_url: str = None,
    colors: str = None,  # Comma-separated colors
    current_user: User = Depends(get_current_user)
):
    """Update a material"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update materials")
    
    material = await db.materials.find_one({"id": material_id})
    if not material:
        raise HTTPException(status_code=404, detail="Materiaal niet gevonden")
    
    update_data = {}
    if name is not None:
        update_data["name"] = name
    if category is not None:
        update_data["category"] = category
    if subcategory is not None:
        update_data["subcategory"] = subcategory
    if price is not None:
        update_data["price"] = price
    if unit is not None:
        update_data["unit"] = unit
    if image_url is not None:
        update_data["image_url"] = image_url
    if colors is not None:
        update_data["colors"] = [c.strip() for c in colors.split(",") if c.strip()]
    
    if update_data:
        await db.materials.update_one({"id": material_id}, {"$set": update_data})
    
    updated = await db.materials.find_one({"id": material_id}, {"_id": 0})
    return updated

@api_router.post("/materials/{material_id}/upload-image")
async def upload_material_image(
    material_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload an image for a material/product"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can upload images")
    
    material = await db.materials.find_one({"id": material_id})
    if not material:
        raise HTTPException(status_code=404, detail="Materiaal niet gevonden")
    
    # Validate file type
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Alleen afbeeldingen zijn toegestaan")
    
    # Save image
    materials_dir = ROOT_DIR / "uploads" / "materials"
    materials_dir.mkdir(parents=True, exist_ok=True)
    
    unique_filename = f"{material_id}_{uuid.uuid4().hex[:8]}_{file.filename}"
    file_path = materials_dir / unique_filename
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    image_url = f"/api/static/materials/{unique_filename}"
    
    # Update material with image URL
    await db.materials.update_one(
        {"id": material_id},
        {"$set": {"image_url": image_url}}
    )
    
    return {"image_url": image_url, "message": "Afbeelding geüpload"}

@api_router.put("/work-items/{work_item_id}/category")
async def update_work_item_category(
    work_item_id: str,
    category: str,
    current_user: User = Depends(get_current_user)
):
    """Update work item category"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update work items")
    
    result = await db.work_items.update_one(
        {"id": work_item_id},
        {"$set": {"category": category}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Werk item niet gevonden")
    
    return {"message": "Categorie bijgewerkt"}

@api_router.post("/work-items/auto-add")
async def auto_add_work_item(
    title: str,
    unit: str,
    price: float,
    current_user: User = Depends(get_current_user)
):
    """Auto-add a work item if it doesn't exist (called when creating quotes)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can add work items")
    
    # Check if work item with same title already exists (case-insensitive)
    existing = await db.work_items.find_one({"title": {"$regex": f"^{title}$", "$options": "i"}})
    if existing:
        # Return the existing item
        existing.pop("_id", None)
        return {"work_item": existing, "created": False}
    
    # Create new work item
    work_item = {
        "id": str(uuid.uuid4()),
        "title": title,
        "unit": unit,
        "price": price,
        "user_id": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "auto_added": True  # Flag to indicate this was auto-added from quote
    }
    
    await db.work_items.insert_one(work_item)
    work_item.pop("_id", None)
    
    return {"work_item": work_item, "created": True}

# ===== PROJECTS ENDPOINTS =====
@api_router.post("/projects", response_model=Project)
async def create_project(project_create: ProjectCreate, current_user: User = Depends(get_current_user)):
    """Create a new project (from quote OR lead)"""
    # Backward compatible: verify quote if quote_id provided
    if project_create.quote_id:
        quote = await db.quotes.find_one({"id": project_create.quote_id, "user_id": current_user.id})
        if not quote:
            raise HTTPException(status_code=404, detail="Quote not found")
    
    # NEW: verify lead if lead_id provided
    if project_create.lead_id:
        lead = await db.leads.find_one({"id": project_create.lead_id, "user_id": current_user.id})
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
    
    project_obj = Project(**project_create.model_dump(), user_id=current_user.id)
    
    project_doc = project_obj.model_dump()
    project_doc["created_at"] = project_doc["created_at"].isoformat()
    if project_doc.get("start_date"):
        project_doc["start_date"] = project_doc["start_date"].isoformat()
    if project_doc.get("end_date"):
        project_doc["end_date"] = project_doc["end_date"].isoformat()
    if project_doc.get("first_visit_date"):
        project_doc["first_visit_date"] = project_doc["first_visit_date"].isoformat()
    
    await db.projects.insert_one(project_doc)
    return project_obj

@api_router.get("/projects", response_model=List[Project])
async def get_projects(current_user: User = Depends(get_current_user)):
    """Get all projects (workers see only visible, all admins see all projects)"""
    # Workers see only projects with visible_to_workers=True, all admins see ALL projects
    if current_user.role == "worker":
        projects = await db.projects.find({"visible_to_workers": True}, {"_id": 0}).to_list(1000)
        # Enrich with lead address for workers
        lead_ids = [p["lead_id"] for p in projects if p.get("lead_id")]
        if lead_ids:
            leads = await db.leads.find({"id": {"$in": lead_ids}}, {"_id": 0, "id": 1, "address": 1}).to_list(1000)
            leads_map = {l["id"]: l.get("address") for l in leads}
            for project in projects:
                project["lead_address"] = leads_map.get(project.get("lead_id"), None)
    else:
        # All admins see all projects
        projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    
    for project in projects:
        if isinstance(project["created_at"], str):
            project["created_at"] = datetime.fromisoformat(project["created_at"])
        if project.get("start_date") and isinstance(project["start_date"], str):
            project["start_date"] = datetime.fromisoformat(project["start_date"])
        if project.get("end_date") and isinstance(project["end_date"], str):
            project["end_date"] = datetime.fromisoformat(project["end_date"])
        if project.get("first_visit_date") and isinstance(project["first_visit_date"], str):
            project["first_visit_date"] = datetime.fromisoformat(project["first_visit_date"])
        
        # Calculate profit from sales_price (which includes approved quotes + legacy documents)
        if current_user.role != "worker":
            # Use stored sales_price (already updated by approved quotes and legacy documents)
            total_sales = project.get("sales_price", 0) or 0
            total_costs = project.get("total_costs", 0) or 0
            
            # Calculate from SOLD quotes only (not just approved)
            lead_id = project.get("lead_id")
            if lead_id:
                # Get SOLD quotes only for actual sales
                sold_quotes = await db.quotes.find({
                    "lead_id": lead_id,
                    "is_sold": True
                }, {"_id": 0, "total_incl_vat": 1}).to_list(100)
                total_sales = sum(q.get("total_incl_vat", 0) for q in sold_quotes)
                
                # Get SOLD legacy documents - check both lead_id and project_id
                sold_legacy = await db.legacy_documents.find({
                    "$or": [
                        {"lead_id": lead_id, "is_sold": True},
                        {"project_id": project["id"], "is_sold": True}
                    ]
                }, {"_id": 0, "total_price": 1}).to_list(100)
                total_sales += sum(d.get("total_price", 0) for d in sold_legacy)
                
                # Calculate potential sales (approved but not sold)
                potential_quotes = await db.quotes.find({
                    "lead_id": lead_id,
                    "status": "goedgekeurd",
                    "$or": [{"is_sold": False}, {"is_sold": {"$exists": False}}]
                }, {"_id": 0, "total_incl_vat": 1}).to_list(100)
                project["potential_sales"] = sum(q.get("total_incl_vat", 0) for q in potential_quotes)
            else:
                # No lead_id - check legacy docs by project_id only
                sold_legacy = await db.legacy_documents.find({
                    "project_id": project["id"],
                    "is_sold": True
                }, {"_id": 0, "total_price": 1}).to_list(100)
                total_sales = sum(d.get("total_price", 0) for d in sold_legacy)
            
            if total_sales > 0:
                project["profit"] = total_sales - total_costs
                project["sales_price"] = total_sales
            else:
                project["profit"] = None
                project["sales_price"] = 0
    
    return projects

@api_router.get("/projects/{project_id}", response_model=Project)
async def get_project(project_id: str, current_user: User = Depends(get_current_user)):
    """Get a specific project (all users can see all projects)"""
    # All users (admins and workers) can see all projects
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    if isinstance(project["created_at"], str):
        project["created_at"] = datetime.fromisoformat(project["created_at"])
    if project.get("start_date") and isinstance(project["start_date"], str):
        project["start_date"] = datetime.fromisoformat(project["start_date"])
    if project.get("end_date") and isinstance(project["end_date"], str):
        project["end_date"] = datetime.fromisoformat(project["end_date"])
    
    return Project(**project)

@api_router.put("/projects/{project_id}", response_model=Project)
async def update_project(project_id: str, project_update: ProjectUpdate, current_user: User = Depends(get_current_user)):
    """Update a project (all admins can update)"""
    # All admins can update any project
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update projects")
    
    existing_project = await db.projects.find_one({"id": project_id})
    if not existing_project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    update_data = {k: v for k, v in project_update.model_dump().items() if v is not None}
    
    if update_data:
        # Convert datetime fields to ISO strings
        if "start_date" in update_data:
            update_data["start_date"] = update_data["start_date"].isoformat()
        if "end_date" in update_data:
            update_data["end_date"] = update_data["end_date"].isoformat()
        
        # Calculate costs if cost-related fields are updated
        cost_fields_updated = any(k in update_data for k in ["labor_cost_per_hour", "labor_hours", "material_costs", "other_costs"])
        
        if cost_fields_updated:
            # Get current values
            labor_cost_per_hour = update_data.get("labor_cost_per_hour", existing_project.get("labor_cost_per_hour", 0.0))
            labor_hours = update_data.get("labor_hours", existing_project.get("labor_hours", 0.0))
            material_costs = update_data.get("material_costs", existing_project.get("material_costs", 0.0))
            other_costs = update_data.get("other_costs", existing_project.get("other_costs", 0.0))
            
            # Calculate total costs
            total_labor_cost = labor_cost_per_hour * labor_hours
            total_costs = total_labor_cost + material_costs + other_costs
            update_data["total_costs"] = total_costs
            
            # Get quote to calculate profit
            quote = await db.quotes.find_one({"id": existing_project["quote_id"]})
            if quote:
                # Use total_incl_vat if available, otherwise fall back to total_price
                revenue = quote.get("total_incl_vat", quote.get("total_price", 0.0))
                
                # Use total_costs_incl_vat for comparison if material_costs_incl_vat exists
                material_costs_incl_vat = existing_project.get("material_costs_incl_vat", material_costs)
                if "material_costs" in update_data:
                    # Assume same VAT ratio if not specified
                    if material_costs > 0:
                        vat_ratio = material_costs_incl_vat / material_costs
                    else:
                        vat_ratio = 1.21  # Default 21% VAT
                    material_costs_incl_vat = material_costs * vat_ratio
                
                total_costs_incl_vat = total_labor_cost + material_costs_incl_vat + other_costs
                update_data["total_costs_incl_vat"] = total_costs_incl_vat
                update_data["material_costs_incl_vat"] = material_costs_incl_vat
                
                profit = revenue - total_costs_incl_vat
                profit_margin = (profit / revenue * 100) if revenue > 0 else 0.0
                
                update_data["profit"] = profit
                update_data["profit_margin"] = profit_margin
        
        await db.projects.update_one({"id": project_id}, {"$set": update_data})
        existing_project.update(update_data)
    
    if isinstance(existing_project["created_at"], str):
        existing_project["created_at"] = datetime.fromisoformat(existing_project["created_at"])
    if existing_project.get("start_date") and isinstance(existing_project["start_date"], str):
        existing_project["start_date"] = datetime.fromisoformat(existing_project["start_date"])
    if existing_project.get("end_date") and isinstance(existing_project["end_date"], str):
        existing_project["end_date"] = datetime.fromisoformat(existing_project["end_date"])
    
    return Project(**existing_project)

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, current_user: User = Depends(get_current_user)):
    """Permanently delete a project (all admins can delete)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete projects")
    
    result = await db.projects.delete_one({"id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    
    return {"message": "Project permanently deleted"}

@api_router.put("/projects/{project_id}/toggle-archive")
async def toggle_project_archive(project_id: str, current_user: User = Depends(get_current_user)):
    """Toggle project archive status (all admins can archive)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can archive projects")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    new_status = not project.get("is_archived", False)
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"is_archived": new_status}}
    )
    
    return {"is_archived": new_status, "message": f"Project {'archived' if new_status else 'activated'}"}

@api_router.put("/projects/{project_id}/toggle-worker-visibility")
async def toggle_worker_visibility(project_id: str, current_user: User = Depends(get_current_user)):
    """Toggle project visibility for workers (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can change visibility")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    new_visibility = not project.get("visible_to_workers", False)
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"visible_to_workers": new_visibility}}
    )
    
    return {
        "visible_to_workers": new_visibility, 
        "message": f"Project is nu {'zichtbaar' if new_visibility else 'verborgen'} voor werkmannen"
    }

@api_router.put("/projects/{project_id}/mark-not-sold")
async def mark_project_not_sold(
    project_id: str, 
    reason: str = Query(..., description="Reden waarom niet verkocht"),
    current_user: User = Depends(get_current_user)
):
    """Mark a project as not sold with a reason"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can mark projects as not sold")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {
            "status": "niet verkocht",
            "is_sold": False,
            "not_sold_reason": reason,
            "not_sold_date": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    logger.info(f"Project {project_id} marked as not sold. Reason: {reason}")
    
    return {
        "message": "Project gemarkeerd als niet verkocht",
        "reason": reason
    }

@api_router.put("/projects/{project_id}/reactivate")
async def reactivate_project(project_id: str, current_user: User = Depends(get_current_user)):
    """Reactivate a project that was marked as not sold"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can reactivate projects")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {
            "status": "offerte in opmaak",
            "not_sold_reason": None,
            "not_sold_date": None
        }}
    )
    
    return {"message": "Project opnieuw geactiveerd"}

@api_router.post("/projects/{project_id}/invoices")
async def add_invoice_to_project(project_id: str, invoice: InvoiceUpload, current_user: User = Depends(get_current_user)):
    """Add an invoice to project costs (all admins can add)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can add invoices")
    
    existing_project = await db.projects.find_one({"id": project_id})
    if not existing_project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Add invoice to list
    invoice_doc = invoice.model_dump()
    invoice_doc["upload_date"] = datetime.now(timezone.utc).isoformat()
    
    invoices = existing_project.get("invoice_uploads", [])
    invoices.append(invoice_doc)
    
    # Update material costs with invoice totals
    material_costs_excl = existing_project.get("material_costs", 0) + invoice.total_excl_vat
    material_costs_incl = existing_project.get("material_costs_incl_vat", 0) + invoice.total_incl_vat
    
    # Recalculate total costs
    labor_costs = existing_project.get("labor_cost_per_hour", 0) * existing_project.get("labor_hours", 0)
    other_costs = existing_project.get("other_costs", 0)
    
    total_costs_excl = labor_costs + material_costs_excl + other_costs
    total_costs_incl = labor_costs + material_costs_incl + other_costs
    
    # Get quote to calculate profit
    quote = await db.quotes.find_one({"id": existing_project["quote_id"]})
    if quote:
        revenue = quote.get("total_incl_vat", quote.get("total_price", 0))
        profit = revenue - total_costs_incl
        profit_margin = (profit / revenue * 100) if revenue > 0 else 0.0
    else:
        profit = 0
        profit_margin = 0
    
    # Update project
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {
            "invoice_uploads": invoices,
            "material_costs": material_costs_excl,
            "material_costs_incl_vat": material_costs_incl,
            "total_costs": total_costs_incl,
            "total_costs_incl_vat": total_costs_incl,
            "profit": profit,
            "profit_margin": profit_margin
        }}
    )
    
    return {"message": "Invoice added successfully", "total_invoices": len(invoices)}

@api_router.get("/projects/{project_id}/invoices")
async def get_project_invoices(project_id: str, current_user: User = Depends(get_current_user)):
    """Get all invoices for a project (all admins can view)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view project invoices")
    
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    return {"invoices": project.get("invoice_uploads", [])}

# ============= LEGACY DOCUMENTS (Oude PDF's) =============

# Create uploads directory if it doesn't exist
UPLOADS_DIR = ROOT_DIR / "uploads" / "legacy_documents"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

@api_router.post("/projects/{project_id}/legacy-documents")
async def upload_legacy_document(
    project_id: str,
    document_type: str,
    file: UploadFile = File(...),
    document_date: Optional[str] = None,
    description: Optional[str] = None,
    total_price: Optional[float] = None,
    current_user: User = Depends(get_current_user)
):
    """Upload een oud document (PDF) uit het vorige systeem"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen documenten uploaden")
    
    # Verify project exists
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    # Validate file type
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Alleen PDF bestanden zijn toegestaan")
    
    # Check file size (max 10MB)
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Bestand is te groot (max 10MB)")
    
    # Generate unique filename
    doc_id = f"DOC-{str(uuid.uuid4())[:8].upper()}"
    safe_filename = f"{doc_id}_{file.filename.replace(' ', '_')}"
    file_path = UPLOADS_DIR / safe_filename
    
    # Save file
    with open(file_path, "wb") as f:
        f.write(contents)
    
    # Create document record
    doc_record = {
        "id": doc_id,
        "project_id": project_id,
        "document_type": document_type,
        "filename": safe_filename,
        "original_filename": file.filename,
        "document_date": document_date,
        "description": description,
        "total_price": total_price,
        "visible_to_customer": False,
        "file_size": len(contents),
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "uploaded_by": current_user.id or current_user.username
    }
    
    # Store in database
    await db.legacy_documents.insert_one(doc_record)
    
    # Update project sales_price if this is an approved quote with total_price
    if document_type == "offerte" and total_price and total_price > 0:
        current_sales = project.get("sales_price", 0) or 0
        new_sales = current_sales + total_price
        total_costs = project.get("total_costs", 0) or 0
        new_profit = new_sales - total_costs
        
        await db.projects.update_one(
            {"id": project_id},
            {"$set": {"sales_price": new_sales, "profit": new_profit}}
        )
        logger.info(f"Updated project {project_id} sales_price with legacy document: +{total_price}")
    
    return {
        "success": True,
        "message": "Document succesvol geüpload",
        "document": {**doc_record, "_id": None}
    }

@api_router.get("/projects/{project_id}/legacy-documents")
async def get_project_legacy_documents(project_id: str, current_user: User = Depends(get_current_user)):
    """Haal alle legacy documenten op voor een project"""
    # Verify project exists
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    # Get documents
    documents = await db.legacy_documents.find(
        {"project_id": project_id},
        {"_id": 0}
    ).sort("uploaded_at", -1).to_list(1000)
    
    return documents

@api_router.get("/legacy-documents/{document_id}/download")
async def download_legacy_document(document_id: str, current_user: User = Depends(get_current_user)):
    """Download een legacy document"""
    # All admins can download any document
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen documenten downloaden")
    
    # Get document record
    doc = await db.legacy_documents.find_one({"id": document_id})
    if not doc:
        logger.error(f"Legacy document not found: {document_id}")
        raise HTTPException(status_code=404, detail="Document niet gevonden")
    
    file_path = UPLOADS_DIR / doc["filename"]
    logger.info(f"Attempting to download: {file_path}")
    
    if not file_path.exists():
        logger.error(f"File not found on disk: {file_path}")
        raise HTTPException(status_code=404, detail="Bestand niet gevonden op server")
    
    return FileResponse(
        path=str(file_path),
        filename=doc["original_filename"],
        media_type="application/pdf"
    )

@api_router.delete("/legacy-documents/{document_id}")
async def delete_legacy_document(document_id: str, current_user: User = Depends(get_current_user)):
    """Verwijder een legacy document"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen documenten verwijderen")
    
    # Get document record
    doc = await db.legacy_documents.find_one({"id": document_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Document niet gevonden")
    
    # Delete file
    file_path = UPLOADS_DIR / doc["filename"]
    if file_path.exists():
        file_path.unlink()
    
    # Delete record
    await db.legacy_documents.delete_one({"id": document_id})
    
    return {"success": True, "message": "Document verwijderd"}


@api_router.put("/legacy-documents/{document_id}")
async def update_legacy_document(
    document_id: str,
    update_data: LegacyDocumentUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update een legacy document (bijv. zichtbaarheid of totaalprijs)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen documenten bewerken")
    
    # Get existing document
    doc = await db.legacy_documents.find_one({"id": document_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Document niet gevonden")
    
    # Track if visibility is being set to true (for notification)
    was_visible = doc.get("visible_to_customer", False)
    was_sold = doc.get("is_sold", False)
    
    # Build update dict
    updates = {k: v for k, v in update_data.model_dump(exclude_unset=True).items() if v is not None}
    
    if updates:
        project = await db.projects.find_one({"id": doc["project_id"]})
        
        # Check if total_price changed and update project sales_price accordingly
        if "total_price" in updates and doc.get("document_type") == "offerte":
            old_price = doc.get("total_price", 0) or 0
            new_price = updates["total_price"] or 0
            price_diff = new_price - old_price
            
            if price_diff != 0 and project:
                current_sales = project.get("sales_price", 0) or 0
                new_sales = current_sales + price_diff
                total_costs = project.get("total_costs", 0) or 0
                new_profit = new_sales - total_costs
                
                await db.projects.update_one(
                    {"id": doc["project_id"]},
                    {"$set": {"sales_price": new_sales, "profit": new_profit}}
                )
        
        # If legacy document is marked as sold, update project status and create celebration
        if updates.get("is_sold") == True and not was_sold and project:
            # Update project status to "in uitvoering"
            await db.projects.update_one(
                {"id": project["id"]},
                {"$set": {"status": "in uitvoering"}}
            )
            
            # Create celebration record for all admins
            celebration = {
                "id": f"CELEB-{str(uuid.uuid4())[:8].upper()}",
                "project_id": project["id"],
                "project_name": project.get("name", "Onbekend project"),
                "legacy_doc_id": document_id,
                "document_name": doc.get("original_filename", ""),
                "amount": doc.get("total_price", 0) or updates.get("total_price", 0),
                "sold_at": datetime.now(timezone.utc).isoformat(),
                "sold_by": current_user.id,
                "seen_by": []  # Track which users have seen this celebration
            }
            await db.celebrations.insert_one(celebration)
            logger.info(f"Legacy doc {document_id} marked as sold, project {project['id']} status set to 'in uitvoering'")
        
        await db.legacy_documents.update_one({"id": document_id}, {"$set": updates})
        
        # Send notification if visibility is being set to true for the first time
        if updates.get("visible_to_customer") and not was_visible:
            doc_type = doc.get("document_type", "document")
            doc_name = doc.get("filename", "Document")
            await send_customer_notification(
                project_id=doc["project_id"],
                subject=f"Nieuw document beschikbaar - Q-Technics",
                content_description=f"Er is een nieuw document ({doc_name}) toegevoegd aan uw project. U kunt dit nu bekijken in uw klantenportaal."
            )
    
    # Return updated document
    updated_doc = await db.legacy_documents.find_one({"id": document_id}, {"_id": 0})
    return updated_doc

@api_router.get("/customer-portal/{access_token}/legacy-documents")
async def get_customer_legacy_documents(access_token: str):
    """Haal legacy documenten op voor het klantenportaal"""
    # Find project by access token
    project = await db.projects.find_one(
        {"customer_access_token": access_token},
        {"_id": 0, "id": 1, "name": 1}
    )
    if not project:
        raise HTTPException(status_code=404, detail="Ongeldige toegangslink")
    
    # Get only visible documents for this project
    documents = await db.legacy_documents.find(
        {"project_id": project["id"], "visible_to_customer": True},
        {"_id": 0, "id": 1, "document_type": 1, "original_filename": 1, "document_date": 1, "description": 1, "total_price": 1, "uploaded_at": 1}
    ).sort("uploaded_at", -1).to_list(1000)
    
    return documents

@api_router.get("/customer-portal/{access_token}/legacy-documents/{document_id}/download")
async def download_customer_legacy_document(access_token: str, document_id: str):
    """Download een legacy document via het klantenportaal"""
    # Verify access token
    project = await db.projects.find_one({"customer_access_token": access_token})
    if not project:
        raise HTTPException(status_code=404, detail="Ongeldige toegangslink")
    
    # Get document and verify it belongs to this project AND is visible
    doc = await db.legacy_documents.find_one({
        "id": document_id, 
        "project_id": project["id"],
        "visible_to_customer": True
    })
    if not doc:
        logger.error(f"Customer portal: Document not found or not visible: {document_id}")
        raise HTTPException(status_code=404, detail="Document niet gevonden")
    
    file_path = UPLOADS_DIR / doc["filename"]
    logger.info(f"Customer portal download: {file_path}")
    
    if not file_path.exists():
        logger.error(f"Customer portal: File not found on disk: {file_path}")
        raise HTTPException(status_code=404, detail="Bestand niet gevonden")
    
    return FileResponse(
        path=str(file_path),
        filename=doc["original_filename"],
        media_type="application/pdf"
    )

# ============= EXPORT ROUTES =============

@api_router.get("/quotes/{quote_id}/export/pdf")
async def export_quote_pdf(quote_id: str, current_user: User = Depends(get_current_user)):
    """Export quote as PDF with logo and VAT details (all admins can export)"""
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Image as ReportLabImage
    
    # All admins can export any quote
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can export quotes")
    
    # Get quote
    quote = await db.quotes.find_one({"id": quote_id})
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    # Get lead
    lead = await db.leads.find_one({"id": quote["lead_id"]})
    
    # Get line items
    items = await db.line_items.find({"quote_id": quote_id}, {"_id": 0}).to_list(1000)
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch, leftMargin=0.75*inch, rightMargin=0.75*inch)
    
    story = []
    styles = getSampleStyleSheet()
    
    # Add logo if exists
    logo_path = Path(__file__).parent / 'qtechnics_logo.png'
    if logo_path.exists():
        logo = ReportLabImage(str(logo_path), width=2.5*inch, height=1.1*inch)
        story.append(logo)
        story.append(Spacer(1, 0.2*inch))
    
    # Title
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=20, textColor=colors.HexColor('#1E40AF'))
    story.append(Paragraph(f"Offerte {quote['quote_number']}", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Lead info
    if lead:
        story.append(Paragraph(f"<b>Klant:</b> {lead['name']}", styles['Normal']))
        story.append(Paragraph(f"<b>Email:</b> {lead['email']}", styles['Normal']))
        story.append(Paragraph(f"<b>Telefoon:</b> {lead['phone']}", styles['Normal']))
        story.append(Paragraph(f"<b>Adres:</b> {lead['address']}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
    
    # Separate labor and material items
    labor_items = [item for item in items if item['item_type'] == 'arbeid']
    material_items = [item for item in items if item['item_type'] != 'arbeid']
    
    # Calculate bundled labor total
    labor_total_excl = sum(item.get('total_excl_vat', item.get('quantity', 0) * item.get('unit_price', 0)) for item in labor_items)
    labor_vat_rate = 6  # Renovatie tarief
    labor_vat = labor_total_excl * (labor_vat_rate / 100)
    labor_total_incl = labor_total_excl + labor_vat
    
    # === ARBEID SECTIE MET DETAILS MAAR ZONDER EENHEIDSPRIJZEN ===
    if labor_items:
        # Arbeid header
        labor_header_style = ParagraphStyle('LaborHeader', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#1E40AF'))
        story.append(Paragraph("Arbeid", labor_header_style))
        story.append(Spacer(1, 0.1*inch))
        
        # Style for wrapping description text
        desc_style = ParagraphStyle('DescStyle', parent=styles['Normal'], fontSize=9, leading=11)
        
        # Create labor table showing items with quantity and unit (NO unit price)
        # Same column layout as materials for consistency
        labor_table_data = [['Omschrijving', 'Aantal', 'Eenheid', '', 'Subtotaal', '']]
        
        for item in labor_items:
            unit = item.get('unit', 'm²')
            # Clean up unit display
            if unit in ['m2', 'vierkante meter']:
                unit = 'm²'
            elif unit in ['lm', 'lopende meter']:
                unit = 'm'
            
            # Use Paragraph for description to allow text wrapping
            desc_para = Paragraph(item['description'], desc_style)
            
            labor_table_data.append([
                desc_para,
                f"{item['quantity']:.2f}" if item['quantity'] else '-',
                unit,
                '',  # Empty column for alignment with materials
                '',  # No individual subtotal shown
                ''
            ])
        
        # Add labor total rows
        labor_table_data.append([
            Paragraph('<b>Subtotaal Arbeid</b>', desc_style),
            '', '', '', 
            f"€{labor_total_excl:.2f}",
            'excl. BTW'
        ])
        labor_table_data.append([
            Paragraph(f'<b>BTW {labor_vat_rate}%</b>', desc_style),
            '', '', '',
            f"€{labor_vat:.2f}",
            ''
        ])
        labor_table_data.append([
            Paragraph('<b>Totaal Arbeid incl. BTW</b>', desc_style),
            '', '', '',
            f"€{labor_total_incl:.2f}",
            ''
        ])
        
        labor_table = Table(labor_table_data, colWidths=[2.4*inch, 0.6*inch, 0.6*inch, 0.5*inch, 0.9*inch, 0.7*inch])
        labor_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E5E7EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            # Highlight totals
            ('BACKGROUND', (0, -3), (-1, -1), colors.HexColor('#D1FAE5')),
            ('FONTNAME', (4, -3), (4, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB'))
        ]))
        story.append(labor_table)
        story.append(Spacer(1, 0.3*inch))
    
    # === MATERIALEN SECTIE (met prijzen) ===
    if material_items:
        material_header_style = ParagraphStyle('MaterialHeader', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#1E40AF'))
        story.append(Paragraph("Materialen", material_header_style))
        story.append(Spacer(1, 0.1*inch))
        
        # Style for wrapping description text
        desc_style = ParagraphStyle('DescStyle', parent=styles['Normal'], fontSize=9, leading=11)
        
        # Line items table with VAT for materials
        table_data = [['Omschrijving', 'Aantal', 'Eenheid', 'Prijs', 'Subtotaal', 'BTW%']]
        
        # Add individual material items
        for item in material_items:
            excl_vat = item.get('total_excl_vat', item.get('quantity', 0) * item.get('unit_price', 0))
            vat_rate = item.get('vat_rate', 21)
            unit = item.get('unit', 'stuk')
            
            # Use Paragraph for description to allow text wrapping
            desc_para = Paragraph(item['description'], desc_style)
            
            table_data.append([
                desc_para,
                f"{item['quantity']:.2f}" if item['quantity'] else '-',
                unit,
                f"€{item['unit_price']:.2f}",
                f"€{excl_vat:.2f}",
                f"{vat_rate}%"
            ])
        
        # Calculate material totals
        material_total_excl = sum(item.get('total_excl_vat', 0) for item in material_items)
        material_vat = sum(item.get('vat_amount', 0) for item in material_items)
        material_total_incl = material_total_excl + material_vat
        
        # Add material total rows
        table_data.append([
            Paragraph('<b>Subtotaal Materialen</b>', desc_style),
            '', '', '',
            f"€{material_total_excl:.2f}",
            'excl.'
        ])
        table_data.append([
            Paragraph('<b>BTW Materialen</b>', desc_style),
            '', '', '',
            f"€{material_vat:.2f}",
            ''
        ])
        table_data.append([
            Paragraph('<b>Totaal Materialen incl. BTW</b>', desc_style),
            '', '', '',
            f"€{material_total_incl:.2f}",
            ''
        ])
        
        table = Table(table_data, colWidths=[2.4*inch, 0.6*inch, 0.6*inch, 0.7*inch, 0.9*inch, 0.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E5E7EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, -3), (-1, -1), colors.HexColor('#DBEAFE')),
            ('FONTNAME', (4, -3), (4, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB'))
        ]))
        story.append(table)
        story.append(Spacer(1, 0.2*inch))
    
    # Totals with VAT breakdown
    total_excl = quote.get('total_excl_vat', quote.get('total_price', 0))
    vat_breakdown = quote.get('vat_breakdown', {})
    total_vat = quote.get('total_vat', 0)
    total_incl = quote.get('total_incl_vat', quote.get('total_price', 0))
    
    story.append(Paragraph(f"<b>Totaal excl. BTW:</b> €{total_excl:.2f}", styles['Normal']))
    
    # Show VAT breakdown
    if vat_breakdown:
        story.append(Spacer(1, 0.1*inch))
        for vat_rate, vat_amount in sorted(vat_breakdown.items()):
            story.append(Paragraph(f"<b>BTW {vat_rate}%:</b> €{vat_amount:.2f}", styles['Normal']))
    
    story.append(Paragraph(f"<b>Totaal BTW:</b> €{total_vat:.2f}", styles['Normal']))
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(f"<b>Totaal incl. BTW:</b> €{total_incl:.2f}", title_style))
    
    # === VISUELE MATERIAALLIJST ===
    # Get unique material names from line items and look up their images + prices
    material_info = {}  # name -> {price, quantity, total}
    for item in material_items:
        name = item['description']
        if name not in material_info:
            material_info[name] = {
                'unit_price': item.get('unit_price', 0),
                'quantity': item.get('quantity', 0),
                'total': item.get('total', item.get('unit_price', 0) * item.get('quantity', 0))
            }
    
    # Find materials with images and add price info
    materials_with_images = []
    for name, info in material_info.items():
        material = await db.materials.find_one({
            "name": {"$regex": f"^{name}$", "$options": "i"},
            "image_url": {"$ne": None, "$exists": True}
        })
        if material and material.get('image_url'):
            material['quote_price'] = info['unit_price']
            material['quote_quantity'] = info['quantity']
            material['quote_total'] = info['total']
            materials_with_images.append(material)
    
    # Add visual material list if there are materials with images
    if materials_with_images:
        # Page break before visual list
        from reportlab.platypus import PageBreak
        story.append(PageBreak())
        
        # Header for visual material list
        visual_header_style = ParagraphStyle('VisualHeader', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1E40AF'))
        story.append(Paragraph("Visuele Materiaallijst", visual_header_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Create grid of materials (2 columns)
        material_cells = []
        for material in materials_with_images:
            image_url = material.get('image_url', '')
            image_path = None
            
            # Get the actual file path from the URL
            if image_url.startswith('/api/static/materials/'):
                filename = image_url.split('/')[-1]
                potential_path = ROOT_DIR / "uploads" / "materials" / filename
                if potential_path.exists():
                    image_path = str(potential_path)
            
            # Create cell content
            cell_content = []
            
            if image_path:
                try:
                    # Add image (max 2.5 inch width, maintain aspect ratio)
                    img = ReportLabImage(image_path, width=2.5*inch, height=2.5*inch)
                    img.hAlign = 'CENTER'
                    cell_content.append(img)
                except:
                    pass
            
            # Add material name
            name_style = ParagraphStyle('MaterialName', parent=styles['Normal'], fontSize=11, alignment=1, spaceBefore=6)
            cell_content.append(Paragraph(f"<b>{material['name']}</b>", name_style))
            
            # Add price from quote
            price_style = ParagraphStyle('MaterialPrice', parent=styles['Normal'], fontSize=10, alignment=1, textColor=colors.HexColor('#1E40AF'))
            quote_price = material.get('quote_price', 0)
            cell_content.append(Paragraph(f"€{quote_price:.2f}", price_style))
            
            material_cells.append(cell_content)
        
        # Create table with 2 columns
        if material_cells:
            # Pad to even number
            if len(material_cells) % 2 != 0:
                material_cells.append([''])
            
            # Create rows of 2
            for i in range(0, len(material_cells), 2):
                row_table_data = [[material_cells[i], material_cells[i+1] if i+1 < len(material_cells) else ['']]]
                row_table = Table(row_table_data, colWidths=[3.25*inch, 3.25*inch])
                row_table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('TOPPADDING', (0, 0), (-1, -1), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ]))
                story.append(row_table)
                story.append(Spacer(1, 0.2*inch))
    
    doc.build(story)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=offerte_{quote['quote_number']}.pdf"}
    )

@api_router.get("/quotes/{quote_id}/export/excel")
async def export_quote_excel(quote_id: str, current_user: User = Depends(get_current_user)):
    """Export quote as Excel (all admins can export)"""
    # All admins can export any quote
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can export quotes")
    
    # Get quote
    quote = await db.quotes.find_one({"id": quote_id})
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    # Get lead
    lead = await db.leads.find_one({"id": quote["lead_id"]})
    
    # Get line items
    items = await db.line_items.find({"quote_id": quote_id}, {"_id": 0}).to_list(1000)
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Offerte"
    
    # Header
    ws['A1'] = f"Offerte {quote['quote_number']}"
    ws['A1'].font = ws['A1'].font.copy(bold=True, size=16)
    
    # Lead info
    if lead:
        ws['A3'] = 'Klant:'
        ws['B3'] = lead['name']
        ws['A4'] = 'Email:'
        ws['B4'] = lead['email']
        ws['A5'] = 'Telefoon:'
        ws['B5'] = lead['phone']
        ws['A6'] = 'Adres:'
        ws['B6'] = lead['address']
    
    # Line items
    ws['A8'] = 'Omschrijving'
    ws['B8'] = 'Aantal'
    ws['C8'] = 'Eenheidsprijs'
    ws['D8'] = 'Type'
    ws['E8'] = 'Totaal'
    
    for i, item in enumerate(items, start=9):
        ws[f'A{i}'] = item['description']
        ws[f'B{i}'] = item['quantity']
        ws[f'C{i}'] = item['unit_price']
        ws[f'D{i}'] = item['item_type']
        ws[f'E{i}'] = item['total']
    
    # Totals
    row = len(items) + 10
    ws[f'A{row}'] = 'Subtotaal Arbeid:'
    ws[f'E{row}'] = quote['subtotal_labor']
    ws[f'A{row+1}'] = 'Subtotaal Materiaal:'
    ws[f'E{row+1}'] = quote['subtotal_material']
    ws[f'A{row+2}'] = 'Totaalprijs:'
    ws[f'E{row+2}'] = quote['total_price']
    ws[f'A{row+2}'].font = ws[f'A{row+2}'].font.copy(bold=True)
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=offerte_{quote['quote_number']}.xlsx"}
    )

# ============= DASHBOARD STATS =============

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    """Get dashboard statistics (all admins see all data)"""
    # All admins see all data, workers see nothing
    if current_user.role == "admin":
        total_leads = await db.leads.count_documents({})
        total_quotes = await db.quotes.count_documents({})
        total_projects = await db.projects.count_documents({})
        total_materials = await db.materials.count_documents({})
        
        # Calculate ACTUAL sales (only sold quotes AND sold legacy documents)
        sold_quotes = await db.quotes.find({"is_sold": True}, {"_id": 0, "total_incl_vat": 1}).to_list(1000)
        total_actual_sales = sum(q.get("total_incl_vat", 0) for q in sold_quotes)
        
        # Add legacy documents that are marked as sold
        sold_legacy = await db.legacy_documents.find({"is_sold": True}, {"_id": 0, "total_price": 1}).to_list(1000)
        total_actual_sales += sum(d.get("total_price", 0) for d in sold_legacy)
        
        # Calculate POTENTIAL sales (approved but not sold)
        potential_quotes = await db.quotes.find({
            "status": "goedgekeurd",
            "$or": [{"is_sold": False}, {"is_sold": {"$exists": False}}]
        }, {"_id": 0, "total_incl_vat": 1}).to_list(1000)
        total_potential_sales = sum(q.get("total_incl_vat", 0) for q in potential_quotes)
        
        # Count sold vs potential quotes
        sold_quotes_count = await db.quotes.count_documents({"is_sold": True})
        sold_legacy_count = await db.legacy_documents.count_documents({"is_sold": True})
        total_sold_count = sold_quotes_count + sold_legacy_count
        
        potential_quotes_count = await db.quotes.count_documents({
            "status": "goedgekeurd",
            "$or": [{"is_sold": False}, {"is_sold": {"$exists": False}}]
        })
        
        # Get recent items - exclude MongoDB _id field
        recent_leads = await db.leads.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
        recent_quotes = await db.quotes.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
        
        # Get material reminders - projects starting within 30 days
        reminder_date = datetime.now(timezone.utc) + timedelta(days=30)
        today = datetime.now(timezone.utc)
        
        projects_with_reminders = await db.projects.find({
            "start_date": {"$gte": today.isoformat(), "$lte": reminder_date.isoformat()}
        }, {"_id": 0}).to_list(100)
        
        material_reminders = []
        for project in projects_with_reminders:
            # Get materials from approved quotes
            lead_id = project.get("lead_id")
            quote_materials = []
            if lead_id:
                approved_quotes = await db.quotes.find({
                    "lead_id": lead_id,
                    "status": "goedgekeurd"
                }, {"_id": 0}).to_list(100)
                
                for quote in approved_quotes:
                    items = await db.line_items.find({
                        "quote_id": quote["id"],
                        "item_type": "materiaal"
                    }, {"_id": 0, "description": 1, "quantity": 1, "unit": 1}).to_list(100)
                    quote_materials.extend(items)
            
            start_date = project.get("start_date")
            if isinstance(start_date, str):
                start_date = datetime.fromisoformat(start_date)
            
            days_until_start = (start_date - today).days if start_date else None
            
            material_reminders.append({
                "project_id": project["id"],
                "project_name": project.get("name", "Naamloos"),
                "start_date": start_date.isoformat() if start_date else None,
                "days_until_start": days_until_start,
                "quote_materials": quote_materials,
                "required_materials": project.get("required_materials", ""),
            })
    else:
        total_leads = 0
        total_quotes = 0
        total_projects = 0
        total_materials = 0
        total_actual_sales = 0
        total_potential_sales = 0
        sold_quotes_count = 0
        potential_quotes_count = 0
        recent_leads = []
        recent_quotes = []
        material_reminders = []
    
    return {
        "total_leads": total_leads,
        "total_quotes": total_quotes,
        "total_projects": total_projects,
        "total_materials": total_materials,
        "total_actual_sales": total_actual_sales,
        "total_potential_sales": total_potential_sales,
        "sold_quotes_count": sold_quotes_count,
        "potential_quotes_count": potential_quotes_count,
        "recent_leads": recent_leads,
        "recent_quotes": recent_quotes,
        "material_reminders": material_reminders
    }

# ============= CALENDAR ROUTES =============

@api_router.get("/calendar/events")
async def get_calendar_events(current_user: User = Depends(get_current_user)):
    """Get all calendar events from projects and work slips (all admins see all)"""
    events = []
    
    # 1. Get project events - all admins see all projects
    if current_user.role == "admin":
        projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    else:
        projects = await db.projects.find({"visible_to_workers": True}, {"_id": 0}).to_list(1000)
    
    for project in projects:
        # First collect scheduled work periods - these should be displayed IN FRONT
        scheduled_days = project.get("scheduled_days", [])
        work_events = []
        
        for period in scheduled_days:
            start_date = period.get("start_date")
            end_date = period.get("end_date")
            
            # Support both old format (single date) and new format (date range)
            if not end_date:
                end_date = period.get("date", start_date)
                start_date = period.get("date", start_date)
            
            if start_date and end_date:
                try:
                    # Parse dates
                    if 'T' in str(start_date):
                        start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                    else:
                        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                    
                    if 'T' in str(end_date):
                        end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                    else:
                        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                    
                    # Add 1 day to end_date for proper calendar display (calendars use exclusive end dates)
                    end_dt_display = end_dt + timedelta(days=1)
                    
                    description = period.get("description", period.get("notes", "Gepland werk"))
                    
                    work_event = {
                        "id": f"work_{project['id']}_{start_date}",
                        "title": f"🔧 {description}",
                        "start": start_dt.isoformat(),
                        "end": end_dt_display.isoformat(),
                        "project_id": project["id"],
                        "project_name": project.get("name", "Project"),
                        "status": project.get("status", "actief"),
                        "color": "#F59E0B",
                        "type": "scheduled_work",
                        "description": description
                    }
                    work_events.append(work_event)
                except Exception as e:
                    logger.error(f"Error parsing scheduled work dates: {e}")
        
        # Add work events FIRST (they appear in front)
        events.extend(work_events)
        
        # Then add project event SECOND (it appears behind/below)
        if project.get("start_date") or project.get("end_date"):
            start = project.get("start_date")
            end = project.get("end_date")
            
            if isinstance(start, str):
                start = datetime.fromisoformat(start)
            if isinstance(end, str):
                end = datetime.fromisoformat(end)
            
            project_event = {
                "id": f"project_{project['id']}",
                "title": project.get("name", "Naamloos Project"),
                "start": start.isoformat() if start else None,
                "end": end.isoformat() if end else None,
                "project_id": project["id"],
                "status": project.get("status", "actief"),
                "color": project.get("color", "#1E40AF"),
                "type": "project",
                "scheduled_work": [{"start": w["start"], "end": w["end"], "description": w["description"]} for w in work_events]
            }
            
            if project_event["start"]:
                events.append(project_event)
    
    # 2. Get work slip events - all admins see all work slips
    if current_user.role == "admin":
        work_slips = await db.work_slips.find({}, {"_id": 0}).to_list(1000)
    else:
        # Workers only see work slips for visible projects
        visible_project_ids = [p["id"] for p in projects]
        work_slips = await db.work_slips.find({"project_id": {"$in": visible_project_ids}}, {"_id": 0}).to_list(1000)
    
    for slip in work_slips:
        slip_date = slip.get("date")
        if isinstance(slip_date, str):
            slip_date = datetime.fromisoformat(slip_date)
        
        if slip_date:
            # Get project name
            project = await db.projects.find_one({"id": slip["project_id"]}, {"_id": 0})
            project_name = project.get("name", "Project") if project else "Project"
            project_color = project.get("color", "#10B981") if project else "#10B981"
            
            event = {
                "id": f"workslip_{slip['id']}",
                "title": f"📋 {project_name}",
                "start": slip_date.isoformat(),
                "end": slip_date.isoformat(),
                "project_id": slip["project_id"],
                "work_slip_id": slip["id"],
                "color": project_color,
                "type": "workslip"
            }
            events.append(event)
    
    return events

# ============= INVOICE UPLOAD ROUTES =============

@api_router.post("/projects/{project_id}/invoices/upload")
async def upload_invoice(
    project_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Upload and parse a purchase invoice PDF for a project.
    Automatically extracts total amounts and adds to project costs.
    """
    # Verify project exists
    project = await db.projects.find_one({"id": project_id, "user_id": current_user.id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Validate file type
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are supported")
    
    # Save file temporarily
    temp_file = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
            temp_file = tmp.name
            content = await file.read()
            tmp.write(content)
        
        # Parse invoice
        parser = InvoiceParser()
        amounts = parser.parse_invoice(temp_file)
        
        # Store invoice data
        invoice_data = {
            "filename": file.filename,
            "total_excl_vat": float(amounts['total_excl_vat']),
            "total_incl_vat": float(amounts['total_incl_vat']),
            "vat_amount": float(amounts['vat_amount']),
            "upload_date": datetime.now(timezone.utc).isoformat(),
            "uploaded_by": current_user.id
        }
        
        # Add to project's invoice list
        await db.projects.update_one(
            {"id": project_id},
            {"$push": {"invoice_uploads": invoice_data}}
        )
        
        # Update project costs (both excl and incl VAT)
        invoice_excl_vat = float(amounts['total_excl_vat'])
        invoice_incl_vat = float(amounts['total_incl_vat'])
        
        # Get current costs
        current_material_costs = project.get("material_costs", 0)
        current_material_costs_incl_vat = project.get("material_costs_incl_vat", 0)
        current_labor_costs = project.get("labor_hours", 0) * project.get("labor_cost_per_hour", 0)
        current_other_costs = project.get("other_costs", 0)
        
        # Calculate new costs
        new_material_costs = current_material_costs + invoice_excl_vat
        new_material_costs_incl_vat = current_material_costs_incl_vat + invoice_incl_vat
        new_total_costs = current_labor_costs + new_material_costs + current_other_costs
        new_total_costs_incl_vat = current_labor_costs + new_material_costs_incl_vat + current_other_costs
        
        await db.projects.update_one(
            {"id": project_id},
            {"$set": {
                "material_costs": new_material_costs,
                "material_costs_incl_vat": new_material_costs_incl_vat,
                "total_costs": new_total_costs,
                "total_costs_incl_vat": new_total_costs_incl_vat
            }}
        )
        
        # Recalculate profit
        quote = await db.quotes.find_one({"id": project["quote_id"]})
        if quote:
            revenue = quote.get("total_incl_vat", quote.get("total_price", 0))
            profit = revenue - new_total_costs_incl_vat
            margin = (profit / revenue * 100) if revenue > 0 else 0
            
            await db.projects.update_one(
                {"id": project_id},
                {"$set": {
                    "profit": profit,
                    "profit_margin": margin
                }}
            )
        
        logger.info(f"Invoice uploaded for project {project_id}: {file.filename}")
        
        return {
            "success": True,
            "invoice": invoice_data,
            "extracted_amounts": {
                "total_excl_vat": float(amounts['total_excl_vat']),
                "total_incl_vat": float(amounts['total_incl_vat']),
                "vat_amount": float(amounts['vat_amount'])
            },
            "project_updated": True
        }
        
    except Exception as e:
        logger.error(f"Error processing invoice: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process invoice: {str(e)}"
        )
    
    finally:
        # Clean up temp file
        if temp_file and os.path.exists(temp_file):
            os.unlink(temp_file)

@api_router.get("/projects/{project_id}/invoices")
async def get_project_invoices(
    project_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get all uploaded invoices for a project."""
    project = await db.projects.find_one({"id": project_id, "user_id": current_user.id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    invoices = project.get("invoice_uploads", [])
    return {"invoices": invoices}

@api_router.delete("/projects/{project_id}/invoices/{invoice_index}")
async def delete_project_invoice(
    project_id: str,
    invoice_index: int,
    current_user: User = Depends(get_current_user)
):
    """Delete an invoice and revert the cost changes."""
    project = await db.projects.find_one({"id": project_id, "user_id": current_user.id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    invoices = project.get("invoice_uploads", [])
    if invoice_index < 0 or invoice_index >= len(invoices):
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Get the invoice to delete
    invoice_to_delete = invoices[invoice_index]
    invoice_excl_vat = invoice_to_delete.get("total_excl_vat", 0)
    invoice_incl_vat = invoice_to_delete.get("total_incl_vat", 0)
    
    # Remove invoice from list
    invoices.pop(invoice_index)
    
    # Recalculate costs (subtract the invoice amounts)
    current_material_costs = project.get("material_costs", 0)
    current_material_costs_incl_vat = project.get("material_costs_incl_vat", 0)
    current_labor_costs = project.get("labor_hours", 0) * project.get("labor_cost_per_hour", 0)
    current_other_costs = project.get("other_costs", 0)
    
    new_material_costs = max(0, current_material_costs - invoice_excl_vat)
    new_material_costs_incl_vat = max(0, current_material_costs_incl_vat - invoice_incl_vat)
    new_total_costs = current_labor_costs + new_material_costs + current_other_costs
    new_total_costs_incl_vat = current_labor_costs + new_material_costs_incl_vat + current_other_costs
    
    # Update project
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {
            "invoice_uploads": invoices,
            "material_costs": new_material_costs,
            "material_costs_incl_vat": new_material_costs_incl_vat,
            "total_costs": new_total_costs,
            "total_costs_incl_vat": new_total_costs_incl_vat
        }}
    )
    
    # Recalculate profit
    quote = await db.quotes.find_one({"id": project["quote_id"]})
    if quote:
        revenue = quote.get("total_incl_vat", quote.get("total_price", 0))
        profit = revenue - new_total_costs_incl_vat
        margin = (profit / revenue * 100) if revenue > 0 else 0
        
        await db.projects.update_one(
            {"id": project_id},
            {"$set": {
                "profit": profit,
                "profit_margin": margin
            }}
        )
    
    logger.info(f"Invoice {invoice_index} deleted from project {project_id}")
    
    return {
        "success": True,
        "message": "Invoice deleted and costs adjusted",
        "deleted_invoice": invoice_to_delete
    }

# ============= WERKBON / DAILY REPORT ROUTES =============

@api_router.get("/projects/{project_id}/quote-materials")
async def get_project_quote_materials(project_id: str, current_user: User = Depends(get_current_user)):
    """Get materials from project's quote ONLY (NO PRICES - for werkbon)"""
    # Verify project exists
    project = await db.projects.find_one({"id": project_id, "user_id": current_user.id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    materials = []
    
    # Get quote
    quote_id = project.get("quote_id")
    if not quote_id:
        return materials
    
    quote = await db.quotes.find_one({"id": quote_id, "user_id": current_user.id}, {"_id": 0})
    if not quote:
        return materials
    
    # Get line items from quote (both from catalog and manually entered)
    line_items = await db.line_items.find({"quote_id": quote_id}, {"_id": 0}).to_list(1000)
    
    for item in line_items:
        if item.get("item_type") == "materiaal":
            materials.append({
                "id": item.get("id"),
                "description_nl": item.get("description", ""),
                "description_uk": item.get("description", ""),
                "quantity_quoted": item.get("quantity", 0),
                "unit": "stuks"
            })
    
    return materials


@api_router.post("/projects/{project_id}/calculate-labor-costs")
async def calculate_project_labor_costs(project_id: str, current_user: User = Depends(get_current_user)):
    """Calculate total labor costs from all work slips and update project"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can calculate costs")
    
    # Get project
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Get all work slips for this project
    work_slips = await db.work_slips.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    # Calculate total labor hours and cost
    total_hours = 0.0
    total_labor_cost = 0.0
    
    for slip in work_slips:
        hours = slip.get("hours_worked", 0) or 0
        workers = slip.get("number_of_workers", 1) or 1
        rate = slip.get("hourly_rate", 34.0) or 34.0
        
        slip_hours = hours * workers
        slip_cost = hours * workers * rate
        
        total_hours += slip_hours
        total_labor_cost += slip_cost
    
    # Update project with calculated costs
    update_data = {
        "labor_hours": total_hours,
        "labor_cost_per_hour": 34.0,  # Fixed rate
    }
    
    # Calculate total costs
    material_costs = project.get("material_costs", 0.0)
    other_costs = project.get("other_costs", 0.0)
    total_costs = total_labor_cost + material_costs + other_costs
    
    update_data["total_costs"] = total_costs
    
    # Calculate profit if quote exists
    if project.get("quote_id"):
        quote = await db.quotes.find_one({"id": project["quote_id"]})
        if quote:
            revenue = quote.get("total_incl_vat", quote.get("total_price", 0.0))
            profit = revenue - total_costs
            profit_margin = (profit / revenue * 100) if revenue > 0 else 0
            
            update_data["profit"] = profit
            update_data["profit_margin"] = profit_margin
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": update_data}
    )
    
    return {
        "total_hours": total_hours,
        "total_labor_cost": total_labor_cost,
        "total_costs": total_costs,
        "work_slips_count": len(work_slips)
    }


@api_router.post("/projects/{project_id}/work-slips", response_model=DailyReport)
async def create_work_slip(project_id: str, report: DailyReportCreate, current_user: User = Depends(get_current_user)):
    """Create a new daily report (werkbon) for a project"""
    # Verify project exists - allow all users to create work slips for visible projects
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Set date to now if not provided
    report_data = report.model_dump()
    if report_data.get("date") is None:
        report_data["date"] = datetime.now(timezone.utc)
    
    # Calculate labor cost: hours_worked * number_of_workers * hourly_rate
    hours_worked = report_data.get("hours_worked") or 0
    number_of_workers = report_data.get("number_of_workers") or 1
    hourly_rate = report_data.get("hourly_rate") or 34.0
    labor_cost = hours_worked * number_of_workers * hourly_rate
    report_data["labor_cost"] = labor_cost
    
    # Create report
    report_obj = DailyReport(**report_data, user_id=current_user.id)
    report_doc = report_obj.model_dump()
    report_doc["date"] = report_doc["date"].isoformat()
    report_doc["created_at"] = report_doc["created_at"].isoformat()
    report_doc["updated_at"] = report_doc["updated_at"].isoformat()
    
    await db.work_slips.insert_one(report_doc)
    
    # Update project labor costs - sum all work slips labor costs
    await recalculate_project_labor_from_workslips(project_id)
    
    return report_obj

async def recalculate_project_labor_from_workslips(project_id: str):
    """Recalculate project labor costs from all work slips"""
    # Get all work slips for this project
    work_slips = await db.work_slips.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    # Sum up all labor costs and hours
    total_labor_cost = sum(slip.get("labor_cost", 0) or 0 for slip in work_slips)
    total_hours = sum((slip.get("hours_worked", 0) or 0) * (slip.get("number_of_workers", 1) or 1) for slip in work_slips)
    
    # Get current project
    project = await db.projects.find_one({"id": project_id})
    if not project:
        return
    
    # Calculate total costs
    material_costs = project.get("material_costs", 0) or 0
    other_costs = project.get("other_costs", 0) or 0
    total_costs = total_labor_cost + material_costs + other_costs
    
    # Update project with new labor costs from work slips
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {
            "labor_hours": total_hours,  # Total man-hours
            "labor_cost_from_workslips": total_labor_cost,  # Labor cost calculated from work slips
            "total_costs": total_costs
        }}
    )

@api_router.get("/projects/{project_id}/work-slips", response_model=List[DailyReport])
async def get_work_slips(project_id: str, current_user: User = Depends(get_current_user)):
    """Get all daily reports for a project"""
    # Verify project exists - allow workers to see visible projects
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    reports = await db.work_slips.find({"project_id": project_id}, {"_id": 0}).sort("date", -1).to_list(1000)
    
    # Convert date strings back to datetime
    for report in reports:
        for field in ["date", "created_at", "updated_at"]:
            if field in report and isinstance(report[field], str):
                report[field] = datetime.fromisoformat(report[field])
        
        # Hide financial info from workers
        if current_user.role == "worker":
            report["hourly_rate"] = None
            report["labor_cost"] = None
    
    return reports

@api_router.put("/projects/{project_id}/work-slips/{slip_id}", response_model=DailyReport)
async def update_work_slip(project_id: str, slip_id: str, report_update: DailyReportUpdate, current_user: User = Depends(get_current_user)):
    """Update a daily report"""
    # Verify project exists
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Find existing report
    existing = await db.work_slips.find_one({"id": slip_id, "project_id": project_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Work slip not found")
    
    # Update
    update_data = {k: v for k, v in report_update.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Recalculate labor cost if hours or workers changed
    hours_worked = update_data.get("hours_worked", existing.get("hours_worked", 0)) or 0
    number_of_workers = update_data.get("number_of_workers", existing.get("number_of_workers", 1)) or 1
    hourly_rate = update_data.get("hourly_rate", existing.get("hourly_rate", 34.0)) or 34.0
    update_data["labor_cost"] = hours_worked * number_of_workers * hourly_rate
    
    await db.work_slips.update_one(
        {"id": slip_id, "project_id": project_id},
        {"$set": update_data}
    )
    
    # Recalculate project labor costs
    await recalculate_project_labor_from_workslips(project_id)
    
    # Fetch updated report
    updated = await db.work_slips.find_one({"id": slip_id}, {"_id": 0})
    
    # Convert dates
    for field in ["date", "created_at", "updated_at"]:
        if field in updated and isinstance(updated[field], str):
            updated[field] = datetime.fromisoformat(updated[field])
    
    return DailyReport(**updated)

@api_router.delete("/projects/{project_id}/work-slips/{slip_id}")
async def delete_work_slip(project_id: str, slip_id: str, current_user: User = Depends(get_current_user)):
    """Delete a daily report"""
    # Verify project exists
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    result = await db.work_slips.delete_one({"id": slip_id, "project_id": project_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Work slip not found")
    
    # Recalculate project labor costs after deletion
    await recalculate_project_labor_from_workslips(project_id)
    
    return {"message": "Work slip deleted"}

@api_router.post("/projects/{project_id}/work-slips/{slip_id}/photos")
async def upload_work_slip_photo(
    project_id: str, 
    slip_id: str, 
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload a photo to a work slip"""
    # All admins can upload photos to any project
    if current_user.role == "admin":
        project = await db.projects.find_one({"id": project_id})
    else:
        project = await db.projects.find_one({"id": project_id, "user_id": current_user.id})
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Verify work slip exists
    slip = await db.work_slips.find_one({"id": slip_id, "project_id": project_id})
    if not slip:
        raise HTTPException(status_code=404, detail="Work slip not found")
    
    # Read file content
    content = await file.read()
    file_extension = Path(file.filename).suffix.lower().replace(".", "")
    
    # Convert HEIC/HEIF to JPEG for browser compatibility
    if file_extension in ["heic", "heif"]:
        try:
            from PIL import Image
            import pillow_heif
            import io
            
            pillow_heif.register_heif_opener()
            heic_image = Image.open(io.BytesIO(content))
            if heic_image.mode in ('RGBA', 'P'):
                heic_image = heic_image.convert('RGB')
            
            output_buffer = io.BytesIO()
            heic_image.save(output_buffer, format='JPEG', quality=85)
            content = output_buffer.getvalue()
            file_extension = "jpg"
            logger.info(f"Converted HEIC work slip photo to JPEG")
        except Exception as e:
            logger.error(f"Failed to convert HEIC: {str(e)}")
    
    # Save file to disk
    upload_dir = Path(__file__).parent / "uploads" / "work_slips" / project_id
    upload_dir.mkdir(parents=True, exist_ok=True)
    
    unique_filename = f"{slip_id}_{uuid.uuid4().hex[:8]}.{file_extension}"
    file_path = upload_dir / unique_filename
    
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Add to work slip photos array
    photo_url = f"/api/static/work_slips/{project_id}/{unique_filename}"
    await db.work_slips.update_one(
        {"id": slip_id},
        {"$push": {"photos": photo_url}}
    )
    
    return {"url": photo_url, "filename": unique_filename}

# ============= INVOICE / FACTUUR ROUTES =============

async def generate_invoice_number():
    """Generate next invoice number in format FACT-2025-001"""
    current_year = datetime.now(timezone.utc).year
    
    # Find last invoice of current year
    last_invoice = await db.invoices.find_one(
        {"invoice_number": {"$regex": f"^FACT-{current_year}-"}},
        sort=[("invoice_number", -1)]
    )
    
    if last_invoice:
        # Extract number and increment
        last_num = int(last_invoice["invoice_number"].split("-")[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f"FACT-{current_year}-{new_num:03d}"

def generate_ogm_reference(invoice_number: str):
    """
    Generate Belgian OGM (gestructureerde mededeling) payment reference
    Format: +++123/4567/89012+++
    Last 2 digits are modulo 97 checksum
    """
    # Extract numeric part from invoice number (e.g., FACT-2025-001 -> 2025001)
    import re
    numbers = re.sub(r'[^0-9]', '', invoice_number)
    
    # Ensure we have at least 10 digits, pad if needed
    if len(numbers) < 10:
        numbers = numbers.zfill(10)
    else:
        numbers = numbers[:10]  # Take first 10 digits
    
    # Calculate modulo 97 checksum
    check_number = int(numbers) % 97
    if check_number == 0:
        check_number = 97
    
    # Format as OGM: +++nnn/nnnn/nnncc+++
    formatted = f"+++{numbers[:3]}/{numbers[3:7]}/{numbers[7:10]}{check_number:02d}+++"
    
    return formatted

def hash_password(password: str) -> str:
    """Hash password using SHA-256"""
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password: str, password_hash: str) -> bool:
    """Verify password against hash"""
    return hash_password(password) == password_hash

@api_router.post("/projects/{project_id}/invoices/create")
async def create_invoice(
    project_id: str,
    request: dict,
    current_user: User = Depends(get_current_user)
):
    """Create a milestone-based customer invoice for a project - based on ALL approved quotes"""
    logger.info(f"Creating invoice - received data: {request}")
    
    # Validate request data
    try:
        invoice_data = InvoiceCreate(**request)
    except Exception as e:
        logger.error(f"Validation error: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    
    logger.info(f"Creating invoice for project {project_id}, milestone: {invoice_data.milestone}, percentage: {invoice_data.milestone_percentage}")
    
    # Verify project exists
    project = await db.projects.find_one({"id": project_id, "user_id": current_user.id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Get ALL approved quotes for this project's lead
    lead_id = project.get("lead_id")
    if not lead_id:
        raise HTTPException(status_code=400, detail="Project has no lead_id")
    
    approved_quotes = await db.quotes.find({
        "lead_id": lead_id,
        "status": "goedgekeurd"
    }, {"_id": 0}).to_list(1000)
    
    if not approved_quotes:
        raise HTTPException(status_code=404, detail="Geen goedgekeurde offertes gevonden")
    
    # Sum up all approved quotes
    total_subtotal_labor = sum(q.get("subtotal_labor", 0) for q in approved_quotes)
    total_subtotal_material = sum(q.get("subtotal_material", 0) for q in approved_quotes)
    total_excl_vat_all = sum(q.get("total_excl_vat", 0) for q in approved_quotes)
    total_vat_all = sum(q.get("total_vat", 0) for q in approved_quotes)
    total_incl_vat_all = sum(q.get("total_incl_vat", 0) for q in approved_quotes)
    
    # Combine VAT breakdowns from all quotes
    combined_vat_breakdown = {}
    for quote in approved_quotes:
        for vat_rate, amount in quote.get("vat_breakdown", {}).items():
            if vat_rate not in combined_vat_breakdown:
                combined_vat_breakdown[vat_rate] = 0.0
            combined_vat_breakdown[vat_rate] += amount
    
    # Get all line items from all approved quotes
    all_line_items = []
    for quote in approved_quotes:
        items = await db.line_items.find({"quote_id": quote["id"]}, {"_id": 0}).to_list(1000)
        all_line_items.extend(items)
    
    # Calculate invoice amount based on milestone percentage
    percentage = invoice_data.milestone_percentage / 100.0
    
    invoice_subtotal_labor = total_subtotal_labor * percentage
    invoice_subtotal_material = total_subtotal_material * percentage
    invoice_total_excl_vat = total_excl_vat_all * percentage
    invoice_total_vat = total_vat_all * percentage
    invoice_total_incl_vat = total_incl_vat_all * percentage
    
    # Calculate VAT breakdown for invoice
    vat_breakdown = {}
    for vat_rate, amount in combined_vat_breakdown.items():
        vat_breakdown[str(vat_rate)] = amount * percentage
    
    # Generate invoice number
    invoice_number = await generate_invoice_number()
    
    # Generate OGM payment reference
    payment_reference = generate_ogm_reference(invoice_number)
    
    # Calculate due date (7 days from now)
    due_date = datetime.now(timezone.utc) + timedelta(days=7)
    
    # Create invoice - quote_id references all approved quotes combined
    combined_quote_ids = ",".join([q["id"] for q in approved_quotes])
    
    invoice = Invoice(
        invoice_number=invoice_number,
        project_id=project_id,
        quote_id=combined_quote_ids,  # Store all quote IDs
        milestone=invoice_data.milestone,
        milestone_percentage=invoice_data.milestone_percentage,
        line_items=all_line_items,
        subtotal_labor=invoice_subtotal_labor,
        subtotal_material=invoice_subtotal_material,
        total_excl_vat=invoice_total_excl_vat,
        vat_breakdown=vat_breakdown,
        total_vat=invoice_total_vat,
        total_incl_vat=invoice_total_incl_vat,
        payment_status="unpaid",
        payment_term_days=7,
        payment_reference=payment_reference,
        due_date=due_date,
        user_id=current_user.id
    )
    
    # Save to database
    invoice_doc = invoice.model_dump()
    invoice_doc["invoice_date"] = invoice_doc["invoice_date"].isoformat()
    invoice_doc["created_at"] = invoice_doc["created_at"].isoformat()
    invoice_doc["due_date"] = invoice_doc["due_date"].isoformat()
    
    await db.invoices.insert_one(invoice_doc)
    
    logger.info(f"Invoice {invoice_number} created for project {project_id}, milestone: {invoice_data.milestone}")
    
    return invoice

@api_router.get("/projects/{project_id}/customer-invoices", response_model=List[Invoice])
async def get_project_customer_invoices(
    project_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get all customer invoices for a project (all admins can access)"""
    # All admins can see all project invoices
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    invoices = await db.invoices.find({"project_id": project_id}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Convert dates
    for invoice in invoices:
        for field in ["invoice_date", "created_at", "due_date"]:
            if field in invoice and isinstance(invoice[field], str):
                invoice[field] = datetime.fromisoformat(invoice[field])
        if "paid_date" in invoice and invoice["paid_date"] and isinstance(invoice["paid_date"], str):
            invoice["paid_date"] = datetime.fromisoformat(invoice["paid_date"])
    
    return invoices

# ============= MANUAL INVOICE ENTRIES (Phased Invoicing) =============

@api_router.post("/projects/{project_id}/manual-invoices")
async def create_manual_invoice_entry(
    project_id: str,
    entry: ManualInvoiceEntryCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a manual invoice entry for phased invoicing"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can create manual invoice entries")
    
    # Verify project exists
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    # Parse invoice date
    try:
        invoice_date = datetime.strptime(entry.invoice_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(status_code=400, detail="Ongeldige datum formaat. Gebruik YYYY-MM-DD")
    
    # Create the manual entry
    new_entry = ManualInvoiceEntry(
        project_id=project_id,
        amount=entry.amount,
        description=entry.description,
        invoice_date=invoice_date,
        send_via_billit=entry.send_via_billit,
        user_id=current_user.id
    )
    
    billit_invoice_id = None
    
    # If send_via_billit is True, create and send a real invoice
    if entry.send_via_billit:
        try:
            # Get lead for customer info
            lead = await db.leads.find_one({"id": project.get("lead_id")})
            if not lead:
                raise HTTPException(status_code=400, detail="Geen klantgegevens gevonden voor dit project")
            
            # Create invoice in our system
            invoice_number = f"FAC-{datetime.now(timezone.utc).year}-{str(uuid.uuid4())[:6].upper()}"
            
            # Calculate VAT (assuming 21%)
            vat_rate = 0.21
            amount_excl_vat = entry.amount / (1 + vat_rate)
            vat_amount = entry.amount - amount_excl_vat
            
            invoice_data = {
                "id": f"INV-{str(uuid.uuid4())[:8].upper()}",
                "project_id": project_id,
                "invoice_number": invoice_number,
                "invoice_date": invoice_date,
                "due_date": invoice_date + timedelta(days=30),
                "customer_name": lead.get("name", ""),
                "customer_email": lead.get("email", ""),
                "customer_address": lead.get("address", ""),
                "customer_vat": lead.get("vat_number", ""),
                "line_items": [{
                    "description": entry.description or "Deelfacturatie",
                    "quantity": 1,
                    "unit_price": amount_excl_vat,
                    "vat_rate": 21,
                    "total": entry.amount
                }],
                "subtotal": amount_excl_vat,
                "vat_amount": vat_amount,
                "total_incl_vat": entry.amount,
                "milestone": "manual",
                "milestone_percentage": 0,
                "payment_status": "pending",
                "peppol_status": "not_sent",
                "created_at": datetime.now(timezone.utc),
                "user_id": current_user.id
            }
            
            await db.invoices.insert_one(invoice_data)
            billit_invoice_id = invoice_data["id"]
            new_entry.billit_invoice_id = billit_invoice_id
            
            logger.info(f"Created invoice {invoice_number} for manual entry {new_entry.id}")
        except Exception as e:
            logger.error(f"Failed to create Billit invoice: {str(e)}")
            # Continue without Billit - just save the manual entry
            new_entry.send_via_billit = False
    
    # Save the manual entry
    entry_dict = new_entry.model_dump()
    entry_dict["invoice_date"] = entry_dict["invoice_date"].isoformat()
    entry_dict["created_at"] = entry_dict["created_at"].isoformat()
    await db.manual_invoice_entries.insert_one(entry_dict)
    
    # Return without _id
    entry_dict.pop("_id", None)
    
    return {
        "success": True,
        "entry": entry_dict,
        "billit_invoice_created": billit_invoice_id is not None,
        "message": f"Facturatieregistratie van €{entry.amount:.2f} toegevoegd" + 
                   (f" en factuur aangemaakt" if billit_invoice_id else "")
    }

@api_router.get("/projects/{project_id}/manual-invoices")
async def get_manual_invoice_entries(
    project_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get all manual invoice entries for a project"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view manual invoice entries")
    
    entries = await db.manual_invoice_entries.find(
        {"project_id": project_id}, 
        {"_id": 0}
    ).sort("invoice_date", -1).to_list(1000)
    
    # Convert dates
    for entry in entries:
        if isinstance(entry.get("invoice_date"), str):
            entry["invoice_date"] = datetime.fromisoformat(entry["invoice_date"])
        if isinstance(entry.get("created_at"), str):
            entry["created_at"] = datetime.fromisoformat(entry["created_at"])
    
    return entries

@api_router.delete("/projects/{project_id}/manual-invoices/{entry_id}")
async def delete_manual_invoice_entry(
    project_id: str,
    entry_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete a manual invoice entry"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete manual invoice entries")
    
    result = await db.manual_invoice_entries.delete_one({
        "id": entry_id,
        "project_id": project_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Registratie niet gevonden")
    
    return {"success": True, "message": "Registratie verwijderd"}

@api_router.get("/all-manual-invoices")
async def get_all_manual_invoice_entries(
    current_user: User = Depends(get_current_user)
):
    """Get all manual invoice entries for financial reporting"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view manual invoice entries")
    
    entries = await db.manual_invoice_entries.find({}, {"_id": 0}).to_list(10000)
    
    # Convert dates
    for entry in entries:
        if isinstance(entry.get("invoice_date"), str):
            entry["invoice_date"] = datetime.fromisoformat(entry["invoice_date"])
        if isinstance(entry.get("created_at"), str):
            entry["created_at"] = datetime.fromisoformat(entry["created_at"])
    
    return entries

@api_router.put("/invoices/{invoice_id}", response_model=Invoice)
async def update_invoice(
    invoice_id: str,
    invoice_update: InvoiceUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update invoice payment status"""
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Verify user owns the project
    project = await db.projects.find_one({"id": invoice["project_id"], "user_id": current_user.id})
    if not project:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update_data = {k: v for k, v in invoice_update.model_dump().items() if v is not None}
    
    if "paid_date" in update_data:
        update_data["paid_date"] = update_data["paid_date"].isoformat()
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": update_data}
    )
    
    updated_invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    
    # Convert dates
    for field in ["invoice_date", "created_at", "due_date"]:
        if field in updated_invoice and isinstance(updated_invoice[field], str):
            updated_invoice[field] = datetime.fromisoformat(updated_invoice[field])
    if "paid_date" in updated_invoice and updated_invoice["paid_date"]:
        updated_invoice["paid_date"] = datetime.fromisoformat(updated_invoice["paid_date"])
    
    return Invoice(**updated_invoice)

@api_router.get("/invoices/{invoice_id}/pdf")
async def export_invoice_pdf(invoice_id: str, current_user: User = Depends(get_current_user)):
    """Generate and download invoice PDF"""
    # Get invoice
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Verify user owns the project
    project = await db.projects.find_one({"id": invoice["project_id"], "user_id": current_user.id})
    if not project:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get quote and lead for customer info
    # quote_id can contain multiple IDs separated by comma (when multiple approved quotes)
    quote_ids = invoice["quote_id"].split(",")
    first_quote_id = quote_ids[0].strip()
    quote = await db.quotes.find_one({"id": first_quote_id})
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    lead = await db.leads.find_one({"id": quote["lead_id"]})
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Logo
    logo_path = ROOT_DIR / "qtechnics_logo.png"
    if logo_path.exists():
        img = Image(str(logo_path), width=100, height=50)
        elements.append(img)
        elements.append(Spacer(1, 12))
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1E40AF'),
        spaceAfter=30
    )
    elements.append(Paragraph(f"<b>FACTUUR {invoice['invoice_number']}</b>", title_style))
    elements.append(Spacer(1, 20))
    
    # Invoice info table
    invoice_date_str = datetime.fromisoformat(invoice["invoice_date"]).strftime('%d-%m-%Y') if isinstance(invoice["invoice_date"], str) else invoice["invoice_date"].strftime('%d-%m-%Y')
    due_date_str = datetime.fromisoformat(invoice["due_date"]).strftime('%d-%m-%Y') if isinstance(invoice["due_date"], str) else invoice["due_date"].strftime('%d-%m-%Y')
    
    info_data = [
        ['Factuurdatum:', invoice_date_str],
        ['Vervaldatum:', due_date_str],
        ['Betaaltermijn:', f"{invoice['payment_term_days']} dagen"],
        ['Status:', 'BETAALD' if invoice['payment_status'] == 'paid' else 'ONBETAALD']
    ]
    
    info_table = Table(info_data, colWidths=[100, 150])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Customer info
    elements.append(Paragraph("<b>Factuur aan:</b>", styles['Heading2']))
    customer_info = f"""
    {lead.get('name', 'N/A')}<br/>
    {lead.get('address', 'N/A')}<br/>
    {lead.get('email', 'N/A')}<br/>
    {lead.get('phone', 'N/A')}
    """
    elements.append(Paragraph(customer_info, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Milestone info
    milestone_names = {
        "10_approval": "10% bij akkoord offerte",
        "40_before_start": "40% een week voor start werken",
        "40_completion": "40% bij oplevering",
        "10_satisfaction": "10% bij tevredenheid klant"
    }
    milestone_text = milestone_names.get(invoice["milestone"], invoice["milestone"])
    elements.append(Paragraph(f"<b>Deelfactuur:</b> {milestone_text} ({invoice['milestone_percentage']}%)", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Line items with VAT - Bundle labor items
    elements.append(Paragraph("<b>Specificatie</b>", styles['Heading2']))
    elements.append(Spacer(1, 10))
    
    # Adjust amounts based on milestone percentage
    percentage = invoice["milestone_percentage"] / 100.0
    
    # Bundle all labor items
    labor_items = [item for item in invoice["line_items"] if item.get("item_type") == "arbeid"]
    material_items = [item for item in invoice["line_items"] if item.get("item_type") == "materiaal"]
    
    # === ARBEID SECTIE MET DETAILS MAAR ZONDER EENHEIDSPRIJZEN ===
    if labor_items:
        elements.append(Paragraph("<b>Arbeid</b>", styles['Heading3']))
        elements.append(Spacer(1, 5))
        
        labor_total_excl = sum(item.get("total_excl_vat", 0) for item in labor_items) * percentage
        labor_total_incl = sum(item.get("total_incl_vat", 0) for item in labor_items) * percentage
        labor_vat_rate = labor_items[0].get("vat_rate", 6) if labor_items else 6
        labor_vat = labor_total_excl * (labor_vat_rate / 100)
        
        # Create labor table showing items with quantity and unit (NO unit price)
        labor_table_data = [['Omschrijving', 'Hoeveelheid', 'Eenheid']]
        
        for item in labor_items:
            unit = item.get('unit', 'm²')
            if unit in ['m2', 'vierkante meter']:
                unit = 'm²'
            elif unit in ['lm', 'lopende meter']:
                unit = 'm'
            
            qty = item.get("quantity", 0) * percentage
            labor_table_data.append([
                item.get('description', '')[:50],
                f"{qty:.2f}" if qty else '-',
                unit
            ])
        
        # Add labor total rows
        labor_table_data.append([
            'Subtotaal Arbeid',
            '',
            f"€{labor_total_excl:.2f} excl. BTW"
        ])
        labor_table_data.append([
            f'BTW {labor_vat_rate}%',
            '',
            f"€{labor_vat:.2f}"
        ])
        labor_table_data.append([
            'Totaal Arbeid incl. BTW',
            '',
            f"€{labor_total_incl:.2f}"
        ])
        
        labor_table = Table(labor_table_data, colWidths=[250, 80, 100])
        labor_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E5E7EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('BACKGROUND', (0, -3), (-1, -1), colors.HexColor('#D1FAE5')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB'))
        ]))
        elements.append(labor_table)
        elements.append(Spacer(1, 15))
    
    # === MATERIALEN SECTIE (met prijzen) ===
    if material_items:
        elements.append(Paragraph("<b>Materialen</b>", styles['Heading3']))
        elements.append(Spacer(1, 5))
        
        table_data = [['Omschrijving', 'Aantal', 'Prijs excl.', 'BTW%', 'Totaal excl.', 'Totaal incl.']]
        
        for item in material_items:
            qty = item.get("quantity", 0) * percentage
            price = item.get("unit_price", 0)
            vat_rate = item.get("vat_rate", 21)
            total_excl = item.get("total_excl_vat", 0) * percentage
            total_incl = item.get("total_incl_vat", 0) * percentage
            
            table_data.append([
                item.get("description", "")[:35],
                f"{qty:.1f}",
                f'€{price:.2f}',
                f'{vat_rate}%',
                f'€{total_excl:.2f}',
                f'€{total_incl:.2f}'
            ])
        
        table = Table(table_data, colWidths=[180, 40, 60, 40, 70, 70])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
    
    # Totals with VAT breakdown
    elements.append(Paragraph("<b>Totalen</b>", styles['Heading2']))
    elements.append(Spacer(1, 10))
    
    story = []
    story.append(Paragraph(f"<b>Totaal excl. BTW:</b> €{invoice['total_excl_vat']:.2f}", styles['Normal']))
    
    # VAT breakdown by rate
    for vat_rate, vat_amount in invoice.get("vat_breakdown", {}).items():
        story.append(Paragraph(f"<b>BTW {vat_rate}%:</b> €{vat_amount:.2f}", styles['Normal']))
    
    story.append(Paragraph(f"<b>Totaal BTW:</b> €{invoice['total_vat']:.2f}", styles['Normal']))
    story.append(Spacer(1, 10))
    
    total_style = ParagraphStyle(
        'TotalStyle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#1E40AF'),
        spaceAfter=6
    )
    story.append(Paragraph(f"<b>TE BETALEN: €{invoice['total_incl_vat']:.2f}</b>", total_style))
    
    for item in story:
        elements.append(item)
    
    elements.append(Spacer(1, 30))
    
    # Payment info with OGM reference
    elements.append(Paragraph("<b>Betalingsinformatie</b>", styles['Heading2']))
    
    # Highlight OGM reference
    ogm_style = ParagraphStyle(
        'OGMStyle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#1E40AF'),
        fontName='Helvetica-Bold',
        spaceAfter=10
    )
    
    payment_info = f"""
    Gelieve het bedrag van €{invoice['total_incl_vat']:.2f} over te maken binnen {invoice['payment_term_days']} dagen.<br/>
    """
    elements.append(Paragraph(payment_info, styles['Normal']))
    
    # OGM reference in prominent box
    if invoice.get('payment_reference'):
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"<b>Gestructureerde mededeling:</b>", styles['Normal']))
        elements.append(Spacer(1, 5))
        elements.append(Paragraph(f"{invoice['payment_reference']}", ogm_style))
        elements.append(Spacer(1, 10))
    
    bank_info = f"""
    <b>Bankgegevens:</b><br/>
    [IBAN nummer hier invoeren]<br/>
    [Bank naam]
    """
    elements.append(Paragraph(bank_info, styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=factuur_{invoice['invoice_number']}.pdf"
        }
    )

# ============= WORKER MANAGEMENT ROUTES =============

@api_router.post("/workers")
async def create_worker(worker_data: WorkerCreate, current_user: User = Depends(get_current_user)):
    """Create a new worker account (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can create workers")
    
    # Check if username already exists
    existing_worker = await db.workers.find_one({"username": worker_data.username}, {"_id": 0})
    if existing_worker:
        raise HTTPException(status_code=400, detail="Gebruikersnaam is al in gebruik")
    
    # Check if username exists as admin
    existing_user = await db.users.find_one({"username": worker_data.username}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Gebruikersnaam is al geregistreerd als beheerder")
    
    # Create worker
    worker = Worker(
        username=worker_data.username,
        name=worker_data.name,
        password_hash=hash_password(worker_data.password),
        created_by=current_user.id
    )
    
    worker_doc = worker.model_dump()
    worker_doc["created_at"] = worker_doc["created_at"].isoformat()
    
    await db.workers.insert_one(worker_doc)
    
    # Remove password_hash and _id from response
    worker_doc.pop("password_hash", None)
    worker_doc.pop("_id", None)
    
    return worker_doc

@api_router.get("/workers")
async def get_workers(current_user: User = Depends(get_current_user)):
    """Get all workers (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view workers")
    
    workers = await db.workers.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return workers

@api_router.delete("/workers/{worker_id}")
async def delete_worker(worker_id: str, current_user: User = Depends(get_current_user)):
    """Delete a worker (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete workers")
    
    result = await db.workers.delete_one({"id": worker_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    return {"message": "Worker deleted successfully"}

@api_router.post("/workers/{worker_id}/toggle")
async def toggle_worker_status(worker_id: str, current_user: User = Depends(get_current_user)):
    """Toggle worker active status (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can modify workers")
    
    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    new_status = not worker.get("is_active", True)
    await db.workers.update_one({"id": worker_id}, {"$set": {"is_active": new_status}})
    
    return {"is_active": new_status}

@api_router.post("/workers/{worker_id}/reset-password")
async def reset_worker_password(worker_id: str, new_password: str = Query(..., min_length=6), current_user: User = Depends(get_current_user)):
    """Reset worker password (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can reset passwords")
    
    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Hash new password
    new_hash = hash_password(new_password)
    await db.workers.update_one({"id": worker_id}, {"$set": {"password_hash": new_hash}})
    
    # Invalidate existing sessions for this worker
    await db.sessions.delete_many({"user_id": worker_id})
    
    logger.info(f"Password reset for worker {worker_id} by admin {current_user.id}")
    return {"message": "Password reset successfully", "worker_name": worker.get("name", "")}

@api_router.get("/auth/test-mobile")
async def test_mobile(request: Request):
    """Simple test endpoint to verify mobile connectivity"""
    user_agent = request.headers.get("user-agent", "unknown")
    return {
        "status": "ok",
        "message": "Backend is bereikbaar!",
        "user_agent": user_agent[:100],
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

@api_router.post("/auth/simple-login")
async def simple_login(request: Request, response: Response):
    """Ultra-simple login endpoint for debugging - accepts ANY format"""
    try:
        # Try to get credentials from ANY source
        body_raw = await request.body()
        body_text = body_raw.decode('utf-8') if body_raw else ""
        
        username = None
        password = None
        
        # Try JSON
        try:
            import json
            data = json.loads(body_text)
            username = data.get("username", "").strip()
            password = data.get("password", "").strip()
        except:
            pass
        
        # Log everything for debugging
        logger.info(f"SIMPLE LOGIN - Raw body: {body_text[:200]}")
        logger.info(f"SIMPLE LOGIN - Parsed: username='{username}', password='{password}'")
        
        if not username or not password:
            return {"error": "Missing credentials", "received_body": body_text[:100]}
        
        # Check for Liam - super simple check
        if username.lower() == "liam" and password in ["Liammail123.", "Liammail123", "liammail123.", "liammail123"]:
            session_token = secrets.token_urlsafe(32)
            
            # Store session
            await db.user_sessions.insert_one({
                "user_id": "ADMIN-LIAM",
                "session_token": session_token,
                "expires_at": (datetime.now(timezone.utc) + timedelta(days=30)).isoformat(),
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            
            return {
                "success": True,
                "user": {
                    "id": "ADMIN-LIAM",
                    "username": "Liam",
                    "email": "liam.waerzeggers@qtechnics.be",
                    "name": "Liam",
                    "role": "admin"
                },
                "session_token": session_token
            }
        else:
            return {
                "error": "Invalid credentials",
                "debug": {
                    "username_received": username,
                    "username_length": len(username) if username else 0,
                    "password_length": len(password) if password else 0,
                    "username_match": username.lower() == "liam" if username else False
                }
            }
    except Exception as e:
        logger.error(f"Simple login error: {str(e)}")
        return {"error": str(e)}

@api_router.post("/auth/admin/login")
async def admin_login(
    request: Request,
    response: Response,
    login_data: Optional[AdminLoginRequest] = None,
    username: Optional[str] = None,
    password: Optional[str] = None
):
    """Admin login with username/password - accepts both JSON body and query params"""
    # Log request details for debugging mobile issues
    user_agent = request.headers.get("user-agent", "unknown")
    content_type = request.headers.get("content-type", "none")
    logger.info(f"Login request - User-Agent: {user_agent[:50]}... Content-Type: {content_type}")
    
    # Support both JSON body and query params (query params for backwards compatibility)
    actual_username = None
    actual_password = None
    
    if login_data:
        actual_username = login_data.username
        actual_password = login_data.password
        logger.info(f"Login via JSON body for user: {actual_username}")
    elif username and password:
        actual_username = username
        actual_password = password
        logger.info(f"Login via query params for user: {actual_username}")
    else:
        # Try to parse body manually if pydantic didn't catch it
        try:
            body = await request.json()
            actual_username = body.get("username")
            actual_password = body.get("password")
            logger.info(f"Login via manual JSON parse for user: {actual_username}")
        except:
            pass
    
    if not actual_username or not actual_password:
        logger.warning(f"Login failed - missing credentials. Has login_data: {login_data is not None}, Has query params: {username is not None}")
        raise HTTPException(status_code=400, detail="Gebruikersnaam en wachtwoord zijn verplicht")
    
    # Strip whitespace that mobile keyboards might add
    actual_username = actual_username.strip()
    actual_password = actual_password.strip()
    
    logger.info(f"Admin login attempt for username: '{actual_username}' (length: {len(actual_username)})")
    
    # HARDCODED LOGIN FOR LIAM - always works regardless of database
    # Log exact credentials for debugging (will be removed later)
    logger.info(f"LOGIN ATTEMPT: username='{actual_username}' (len={len(actual_username)}), password='{actual_password}' (len={len(actual_password)})")
    
    # Accept ANY case variation of liam and common password variants
    username_match = actual_username.lower().strip() == "liam"
    password_variants = ["Liammail123.", "Liammail123", "liammail123.", "liammail123", "LiamMail123.", "LiamMail123"]
    password_match = actual_password.strip() in password_variants
    
    logger.info(f"Username match: {username_match}, Password match: {password_match}")
    
    if username_match and password_match:
        logger.info("Liam login via hardcoded credentials")
        session_token = secrets.token_urlsafe(32)
        
        # Set cookie
        response.set_cookie(
            key="session_token",
            value=session_token,
            httponly=True,
            secure=True,
            samesite="none",
            max_age=30 * 24 * 60 * 60
        )
        
        # Store session
        await db.user_sessions.insert_one({
            "user_id": "ADMIN-LIAM",
            "session_token": session_token,
            "expires_at": (datetime.now(timezone.utc) + timedelta(days=30)).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        # Ensure Liam exists in database
        existing = await db.users.find_one({"username": {"$regex": "^liam$", "$options": "i"}})
        if not existing:
            await db.users.insert_one({
                "id": "ADMIN-LIAM",
                "username": "Liam",
                "email": "liam.waerzeggers@qtechnics.be",
                "name": "Liam",
                "role": "admin",
                "password_hash": hash_password("Liammail123."),
                "created_at": datetime.now(timezone.utc).isoformat()
            })
        
        return {
            "user": {
                "id": "ADMIN-LIAM",
                "username": "Liam",
                "email": "liam.waerzeggers@qtechnics.be",
                "name": "Liam",
                "role": "admin"
            },
            "session_token": session_token
        }
    
    # Check if admin exists - try exact match first, then case-insensitive
    admin = await db.users.find_one({"username": actual_username, "role": "admin"})
    
    # If not found, try case-insensitive search (mobile keyboards often capitalize)
    if not admin:
        admin = await db.users.find_one({
            "username": {"$regex": f"^{actual_username}$", "$options": "i"},
            "role": "admin"
        })
        if admin:
            logger.info(f"Found admin via case-insensitive match: {admin.get('username')}")
    
    logger.info(f"Admin found: {admin is not None}")
    if admin:
        logger.info(f"Has password_hash: {admin.get('password_hash') is not None}")
    
    if not admin or not admin.get("password_hash"):
        logger.warning(f"Login failed - admin not found or no password_hash for: {actual_username}")
        raise HTTPException(status_code=401, detail="Ongeldige inloggegevens")
    
    password_valid = verify_password(actual_password, admin["password_hash"])
    logger.info(f"Password valid: {password_valid}")
    
    if not password_valid:
        logger.warning(f"Login failed - invalid password for: {actual_username}")
        raise HTTPException(status_code=401, detail="Ongeldige inloggegevens")
    
    # Create session for admin
    session_token = secrets.token_urlsafe(32)
    expires_at = datetime.now(timezone.utc) + timedelta(days=30)
    
    # Use id field or convert _id to string
    user_id = admin.get("id") or str(admin.get("_id", ""))
    
    session = {
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.user_sessions.insert_one(session)
    
    # Set cookie for session token
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=30 * 24 * 60 * 60  # 30 days
    )
    
    # Return admin user data
    admin_data = {
        "id": user_id,
        "username": admin["username"],
        "email": admin.get("email", ""),
        "name": admin.get("name", admin["username"]),
        "role": "admin",
        "created_at": admin.get("created_at", datetime.now(timezone.utc).isoformat())
    }
    
    return {
        "user": admin_data,
        "session_token": session_token
    }

@api_router.post("/auth/worker/login")
async def worker_login(
    response: Response,
    login_data: Optional[WorkerLoginRequest] = None,
    username: Optional[str] = None,
    password: Optional[str] = None
):
    """Worker login with username/password - ONLY for workers"""
    # Support both JSON body and query params
    actual_username = login_data.username if login_data else username
    actual_password = login_data.password if login_data else password
    
    if not actual_username or not actual_password:
        raise HTTPException(status_code=400, detail="Gebruikersnaam en wachtwoord zijn verplicht")
    
    # Check if worker exists
    worker = await db.workers.find_one({"username": actual_username}, {"_id": 0})
    
    if not worker:
        raise HTTPException(status_code=401, detail="Ongeldige inloggegevens")
    
    if not worker.get("is_active", True):
        raise HTTPException(status_code=403, detail="Account is gedeactiveerd / Обліковий запис деактивовано")
    
    if not verify_password(actual_password, worker["password_hash"]):
        raise HTTPException(status_code=401, detail="Ongeldige inloggegevens")
    
    # Create session for worker
    session_token = secrets.token_urlsafe(32)
    expires_at = datetime.now(timezone.utc) + timedelta(days=30)
    
    session = {
        "user_id": worker["id"],
        "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.sessions.insert_one(session)
    
    # Set cookie for session token
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=30 * 24 * 60 * 60  # 30 days
    )
    
    # Return worker as user with role=worker
    user_data = {
        "id": worker["id"],
        "username": worker["username"],
        "name": worker["name"],
        "role": "worker",
        "created_at": worker["created_at"]
    }
    
    return {
        "user": user_data,
        "session_token": session_token
    }

# ============= PROJECT EERSTE BEZOEK & 3D ONTWERPEN ROUTES =============

@api_router.post("/projects/{project_id}/first-visit/photos")
async def upload_first_visit_photo(
    project_id: str,
    file: UploadFile = File(...),
    room: Optional[str] = Query(default="Algemeen", description="Kamer/map voor deze foto"),
    current_user: User = Depends(get_current_user)
):
    """Upload photo from first visit with optional room assignment"""
    # Allow any admin to upload photos for any project
    if current_user.role == "admin":
        project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    else:
        project = await db.projects.find_one({"id": project_id, "user_id": current_user.id}, {"_id": 0})
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Read file content
    import base64
    file_content = await file.read()
    
    file_extension = file.filename.split(".")[-1].lower()
    
    # Convert HEIC/HEIF to JPEG for browser compatibility
    if file_extension in ["heic", "heif"]:
        try:
            from PIL import Image
            import pillow_heif
            import io
            
            # Register HEIF opener with Pillow
            pillow_heif.register_heif_opener()
            
            # Open HEIC image and convert to JPEG
            heic_image = Image.open(io.BytesIO(file_content))
            
            # Convert to RGB if necessary (HEIC can have alpha channel)
            if heic_image.mode in ('RGBA', 'P'):
                heic_image = heic_image.convert('RGB')
            
            # Save as JPEG
            output_buffer = io.BytesIO()
            heic_image.save(output_buffer, format='JPEG', quality=85)
            file_content = output_buffer.getvalue()
            file_extension = "jpg"
            
            logger.info(f"Converted HEIC image to JPEG for project {project_id}")
        except ImportError:
            logger.warning("pillow-heif not installed, storing HEIC as-is")
        except Exception as e:
            logger.error(f"Failed to convert HEIC: {str(e)}, storing as-is")
    
    base64_data = base64.b64encode(file_content).decode('utf-8')
    unique_filename = f"{uuid.uuid4()}.{file_extension}"
    
    # Determine content type
    content_types = {
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png",
        "gif": "image/gif",
        "webp": "image/webp",
        "heic": "image/heic",
        "heif": "image/heif"
    }
    content_type = content_types.get(file_extension, "image/jpeg")
    
    # Create photo record with base64 data for persistent storage
    photo_record = {
        "filename": unique_filename,
        "original_filename": file.filename,
        "url": f"/api/photos/first_visit/{project_id}/{unique_filename}",
        "room": room or "Algemeen",
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "base64_data": base64_data,
        "content_type": content_type
    }
    
    # Also save to file system for backward compatibility (preview)
    photos_dir = ROOT_DIR / "uploads" / "first_visit" / project_id
    photos_dir.mkdir(parents=True, exist_ok=True)
    file_path = photos_dir / unique_filename
    with open(file_path, "wb") as buffer:
        buffer.write(file_content)
    
    # Update project - store as object with base64 data
    await db.projects.update_one(
        {"id": project_id},
        {"$push": {"first_visit_photos": photo_record}}
    )
    
    # Return record without base64 data (too large for response)
    return_record = {k: v for k, v in photo_record.items() if k != "base64_data"}
    return return_record

@api_router.delete("/projects/{project_id}/first-visit/photos/{photo_name}")
async def delete_first_visit_photo(
    project_id: str,
    photo_name: str,
    current_user: User = Depends(get_current_user)
):
    """Delete first visit photo (supports both old and new format)"""
    # Allow any admin to delete photos for any project
    if current_user.role == "admin":
        project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    else:
        project = await db.projects.find_one({"id": project_id, "user_id": current_user.id}, {"_id": 0})
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    photo_url = f"/api/static/first_visit/{project_id}/{photo_name}"
    
    # Delete file
    file_path = ROOT_DIR / "uploads" / "first_visit" / project_id / photo_name
    if file_path.exists():
        file_path.unlink()
    
    # Update project - handle both old format (string) and new format (object)
    existing_photos = project.get("first_visit_photos", [])
    
    # Filter out the photo (works for both formats)
    updated_photos = []
    for photo in existing_photos:
        if isinstance(photo, str):
            if photo != photo_url:
                updated_photos.append(photo)
        elif isinstance(photo, dict):
            if photo.get("filename") != photo_name and photo.get("url") != photo_url:
                updated_photos.append(photo)
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"first_visit_photos": updated_photos}}
    )
    
    return {"message": "Photo deleted"}

# Serve photos from database (persistent storage)
@api_router.get("/photos/{file_type}/{project_id}/{filename}")
async def serve_photo_from_db(file_type: str, project_id: str, filename: str):
    """Serve photos from database (base64 storage) - persistent across deployments"""
    import base64
    from fastapi.responses import Response
    
    if file_type not in ["first_visit", "designs", "work_slips"]:
        raise HTTPException(status_code=404, detail="Invalid file type")
    
    # Find the project
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Search for the photo in the appropriate field
    photo_data = None
    if file_type == "first_visit":
        photos = project.get("first_visit_photos", [])
        for photo in photos:
            if isinstance(photo, dict) and photo.get("filename") == filename:
                photo_data = photo
                break
    elif file_type == "designs":
        designs = project.get("design_3d_files", [])
        for design in designs:
            if isinstance(design, dict) and design.get("filename") == filename:
                photo_data = design
                break
    
    # If found in database with base64 data, serve it
    if photo_data and photo_data.get("base64_data"):
        content_type = photo_data.get("content_type", "image/jpeg")
        image_bytes = base64.b64decode(photo_data["base64_data"])
        return Response(content=image_bytes, media_type=content_type)
    
    # Fallback: try to serve from file system (for old photos)
    file_path = ROOT_DIR / "uploads" / file_type / project_id / filename
    if file_path.exists():
        suffix = file_path.suffix.lower()
        content_types = {
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".gif": "image/gif",
            ".webp": "image/webp",
        }
        content_type = content_types.get(suffix, "image/jpeg")
        return FileResponse(file_path, media_type=content_type)
    
    raise HTTPException(status_code=404, detail="Photo not found")

# Static file serving endpoint for uploads (workaround for Kubernetes ingress)
@api_router.get("/static/{file_type}/{project_id}/{filename}")
async def serve_static_file(file_type: str, project_id: str, filename: str):
    """Serve static files (photos, designs) via API route"""
    # Validate file_type
    if file_type not in ["first_visit", "designs", "work_slips"]:
        raise HTTPException(status_code=404, detail="Invalid file type")
    
    file_path = ROOT_DIR / "uploads" / file_type / project_id / filename
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    # Determine content type
    suffix = file_path.suffix.lower()
    content_types = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".pdf": "application/pdf",
        ".dxf": "application/dxf",
        ".dwg": "application/dwg",
    }
    content_type = content_types.get(suffix, "application/octet-stream")
    
    return FileResponse(file_path, media_type=content_type)

@api_router.get("/static/materials/{filename}")
async def serve_material_image(filename: str):
    """Serve material/product images"""
    file_path = ROOT_DIR / "uploads" / "materials" / filename
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    suffix = file_path.suffix.lower()
    content_types = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
    }
    content_type = content_types.get(suffix, "image/jpeg")
    
    return FileResponse(file_path, media_type=content_type)

@api_router.get("/static/catalog/{filename}")
async def serve_catalog_image(filename: str):
    """Serve material catalog images"""
    file_path = ROOT_DIR / "uploads" / "catalog" / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    suffix = file_path.suffix.lower()
    content_types = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".gif": "image/gif", ".webp": "image/webp"}
    return FileResponse(file_path, media_type=content_types.get(suffix, "image/jpeg"))

# Alternative route for uploads (some files use /uploads instead of /static)
@api_router.get("/uploads/{file_type}/{project_id}/{filename}")
async def serve_upload_file(file_type: str, project_id: str, filename: str):
    """Serve uploaded files via /uploads route (alternative to /static)"""
    # Validate file_type
    if file_type not in ["first_visit", "designs", "work_slips"]:
        raise HTTPException(status_code=404, detail="Invalid file type")
    
    file_path = ROOT_DIR / "uploads" / file_type / project_id / filename
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    # Determine content type
    suffix = file_path.suffix.lower()
    content_types = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".pdf": "application/pdf",
        ".dxf": "application/dxf",
        ".dwg": "application/dwg",
    }
    content_type = content_types.get(suffix, "application/octet-stream")
    
    return FileResponse(file_path, media_type=content_type)

@api_router.put("/projects/{project_id}/first-visit/notes")
async def update_first_visit_notes(
    project_id: str,
    notes: str,
    current_user: User = Depends(get_current_user)
):
    """Update first visit notes"""
    # Admins can update any project's notes
    if current_user.role == "admin":
        project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    else:
        project = await db.projects.find_one({"id": project_id, "user_id": current_user.id}, {"_id": 0})
    
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"first_visit_notes": notes}}
    )
    
    return {"message": "Notities opgeslagen"}

# ============= PROJECT NOTES & WORKER TASKS =============

@api_router.get("/projects/{project_id}/notes")
async def get_project_notes(project_id: str, current_user: User = Depends(get_current_user)):
    """Get all notes for a project (both first visit and general notes)"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    return {
        "first_visit_notes": project.get("first_visit_notes", ""),
        "project_notes": project.get("project_notes", [])
    }

@api_router.post("/projects/{project_id}/notes")
async def add_project_note(project_id: str, note: dict, current_user: User = Depends(get_current_user)):
    """Add a new note to a project"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen notities toevoegen")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    new_note = {
        "id": str(uuid.uuid4())[:8],
        "text": note.get("text", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.id,
        "created_by_name": current_user.name or current_user.username,
        "is_task": False,
        "assigned_to": None,
        "task_completed": False,
        "task_completed_at": None
    }
    
    await db.projects.update_one(
        {"id": project_id},
        {"$push": {"project_notes": new_note}}
    )
    
    return new_note

@api_router.put("/projects/{project_id}/notes/{note_id}")
async def update_project_note(project_id: str, note_id: str, note: dict, current_user: User = Depends(get_current_user)):
    """Update a project note"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen notities bewerken")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    project_notes = project.get("project_notes", [])
    updated = False
    for i, n in enumerate(project_notes):
        if n.get("id") == note_id:
            project_notes[i] = {**n, **note}
            updated = True
            break
    
    if not updated:
        raise HTTPException(status_code=404, detail="Notitie niet gevonden")
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"project_notes": project_notes}}
    )
    
    return {"message": "Notitie bijgewerkt"}

@api_router.delete("/projects/{project_id}/notes/{note_id}")
async def delete_project_note(project_id: str, note_id: str, current_user: User = Depends(get_current_user)):
    """Delete a project note"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen notities verwijderen")
    
    await db.projects.update_one(
        {"id": project_id},
        {"$pull": {"project_notes": {"id": note_id}}}
    )
    
    # Also delete any associated worker task
    await db.worker_tasks.delete_many({"note_id": note_id})
    
    return {"message": "Notitie verwijderd"}

@api_router.post("/projects/{project_id}/notes/{note_id}/assign")
async def assign_note_as_task(project_id: str, note_id: str, assignment: dict, current_user: User = Depends(get_current_user)):
    """Assign a project note as a task to an admin"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen taken toewijzen")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    # Find the note
    project_notes = project.get("project_notes", [])
    note = None
    for n in project_notes:
        if n.get("id") == note_id:
            note = n
            break
    
    if not note:
        raise HTTPException(status_code=404, detail="Notitie niet gevonden")
    
    # Get admin info - check both admin_id and worker_id for compatibility
    admin_id = assignment.get("admin_id") or assignment.get("worker_id")
    
    # First check in hardcoded admins
    admin_name = None
    if admin_id in HARDCODED_ADMINS:
        admin_info = HARDCODED_ADMINS[admin_id]
        admin_name = admin_info.get("name", admin_info.get("username", ""))
    else:
        # Check in database - try multiple query patterns
        admin = await db.users.find_one({"id": admin_id}, {"_id": 0})
        if not admin:
            admin = await db.users.find_one({"_id": admin_id})
        if admin:
            admin_name = admin.get("name", admin.get("username", ""))
    
    if not admin_name:
        # Last resort: use the admin_id as name (it might be a display name)
        logger.warning(f"Admin not found in DB: {admin_id}, using ID as fallback")
        admin_name = admin_id
    
    # Create task
    task = {
        "id": f"TASK-{str(uuid.uuid4())[:8].upper()}",
        "project_id": project_id,
        "project_name": project.get("name", ""),
        "note_id": note_id,
        "text": note.get("text", ""),
        "assigned_to": admin_id,
        "assigned_to_name": admin_name,
        "assigned_by": current_user.id,
        "assigned_by_name": current_user.name or current_user.username,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "completed": False,
        "completed_at": None,
        "seen": False
    }
    
    # Make a copy for the response (insert_one adds _id to the original dict)
    task_response = {**task}
    
    await db.worker_tasks.insert_one(task)
    
    # Update the note to mark it as a task
    for i, n in enumerate(project_notes):
        if n.get("id") == note_id:
            project_notes[i]["is_task"] = True
            project_notes[i]["assigned_to"] = admin_id
            project_notes[i]["assigned_to_name"] = admin_name
            project_notes[i]["task_id"] = task_response["id"]
            break
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"project_notes": project_notes}}
    )
    
    return task_response

# Worker Task endpoints
@api_router.get("/worker-tasks/my")
async def get_my_worker_tasks(current_user: User = Depends(get_current_user)):
    """Get all tasks assigned to the current worker"""
    tasks = await db.worker_tasks.find(
        {"assigned_to": current_user.id, "completed": False},
        {"_id": 0}
    ).to_list(100)
    return tasks

@api_router.get("/worker-tasks/pending")
async def get_pending_worker_tasks(current_user: User = Depends(get_current_user)):
    """Get all pending (unseen) tasks for the current worker"""
    tasks = await db.worker_tasks.find(
        {"assigned_to": current_user.id, "completed": False},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return tasks

@api_router.put("/worker-tasks/{task_id}/seen")
async def mark_task_seen(task_id: str, current_user: User = Depends(get_current_user)):
    """Mark a task as seen by the worker"""
    await db.worker_tasks.update_one(
        {"id": task_id, "assigned_to": current_user.id},
        {"$set": {"seen": True}}
    )
    return {"message": "Taak gezien"}

@api_router.put("/worker-tasks/{task_id}/complete")
async def complete_worker_task(task_id: str, current_user: User = Depends(get_current_user)):
    """Mark a worker task as completed"""
    task = await db.worker_tasks.find_one({"id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Taak niet gevonden")
    
    # Only assigned worker or admin can complete
    if task.get("assigned_to") != current_user.id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen de toegewezen medewerker kan deze taak voltooien")
    
    completed_at = datetime.now(timezone.utc).isoformat()
    
    # Update worker task
    await db.worker_tasks.update_one(
        {"id": task_id},
        {"$set": {"completed": True, "completed_at": completed_at}}
    )
    
    # Update project note
    project_id = task.get("project_id")
    note_id = task.get("note_id")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if project:
        project_notes = project.get("project_notes", [])
        for i, n in enumerate(project_notes):
            if n.get("id") == note_id:
                project_notes[i]["task_completed"] = True
                project_notes[i]["task_completed_at"] = completed_at
                break
        
        await db.projects.update_one(
            {"id": project_id},
            {"$set": {"project_notes": project_notes}}
        )
    
    return {"message": "Taak voltooid!", "completed_at": completed_at}

@api_router.get("/admin/worker-tasks")
async def get_all_worker_tasks(current_user: User = Depends(get_current_user)):
    """Admin: Get all worker tasks"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins")
    
    tasks = await db.worker_tasks.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return tasks

@api_router.post("/projects/{project_id}/designs")
async def upload_3d_design(
    project_id: str, 
    file: UploadFile = File(...), 
    room: Optional[str] = Query(default="Algemeen", description="Kamer/map voor dit ontwerp"),
    current_user: User = Depends(get_current_user)
):
    """Upload een 3D design bestand (foto/render) voor een project"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen designs uploaden")
    
    # Verify project exists
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    # Validate file type
    allowed_types = ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.pdf', '.skp', '.obj', '.fbx', '.3ds']
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed_types:
        raise HTTPException(status_code=400, detail=f"Bestandstype niet toegestaan. Toegestaan: {', '.join(allowed_types)}")
    
    # Check file size (max 50MB for design files)
    contents = await file.read()
    if len(contents) > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Bestand is te groot (max 50MB)")
    
    # Convert to base64 for persistent storage (only for images, not large 3D files)
    import base64
    base64_data = None
    if ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp'] and len(contents) < 10 * 1024 * 1024:  # Max 10MB for base64
        base64_data = base64.b64encode(contents).decode('utf-8')
    
    # Determine content type
    content_types = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".pdf": "application/pdf"
    }
    content_type = content_types.get(ext, "application/octet-stream")
    
    # Create designs directory for backward compatibility
    designs_dir = ROOT_DIR / "uploads" / "designs" / project_id
    designs_dir.mkdir(parents=True, exist_ok=True)
    
    # Generate unique filename
    unique_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename.replace(' ', '_')}"
    file_path = designs_dir / unique_filename
    
    # Save file to disk (for preview/fallback)
    with open(file_path, "wb") as f:
        f.write(contents)
    
    # Create design record with base64 data for persistent storage
    file_url = f"/api/photos/designs/{project_id}/{unique_filename}"
    design_record = {
        "filename": unique_filename,
        "original_filename": file.filename,
        "url": file_url,
        "size": len(contents),
        "type": ext,
        "room": room or "Algemeen",
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "content_type": content_type
    }
    
    # Add base64 data if available (for persistent storage)
    if base64_data:
        design_record["base64_data"] = base64_data
    
    # Update project with new design file
    existing_designs = project.get("design_3d_files", [])
    existing_designs.append(design_record)
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"design_3d_files": existing_designs}}
    )
    
    logger.info(f"3D design uploaded for project {project_id}: {unique_filename} (room: {room})")
    
    # Return record without base64 data
    return_record = {k: v for k, v in design_record.items() if k != "base64_data"}
    return return_record


# ===== PROJECT MEASUREMENTS ENDPOINTS =====
@api_router.post("/projects/{project_id}/measurements")
async def add_measurement(project_id: str, measurement: dict, current_user: User = Depends(get_current_user)):
    """Add work item measurement for quote generation"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can add measurements")
    
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Add measurement with ID
    measurement['id'] = str(uuid.uuid4())
    measurements = project.get('measurements', [])
    measurements.append(measurement)
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"measurements": measurements}}
    )
    
    return {"message": "Measurement added", "id": measurement['id']}

@api_router.delete("/projects/{project_id}/measurements/{measurement_id}")
async def delete_measurement(project_id: str, measurement_id: str, current_user: User = Depends(get_current_user)):
    """Delete a measurement"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete measurements")
    
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    measurements = [m for m in project.get('measurements', []) if m.get('id') != measurement_id]
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"measurements": measurements}}
    )
    
    return {"message": "Measurement deleted"}

@api_router.post("/projects/{project_id}/generate-quote")
async def generate_quote_from_measurements(project_id: str, current_user: User = Depends(get_current_user)):
    """Generate quote from project measurements - creates individual editable line items"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can generate quotes")
    
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    measurements = project.get('measurements', [])
    if not measurements:
        raise HTTPException(status_code=400, detail="No measurements to generate quote from")
    
    # Check if lead exists
    lead_id = project.get('lead_id')
    if not lead_id:
        raise HTTPException(status_code=400, detail="Project must have a lead to generate quote")
    
    # Create quote first
    quote_id = f"OFF-{datetime.now(timezone.utc).year}-{str(uuid.uuid4())[:6].upper()}"
    
    # Create line items from measurements - store in SEPARATE collection for editing
    line_items_to_insert = []
    
    for m in measurements:
        quantity = float(m.get('quantity', 0))
        unit_price = float(m.get('price', 0))
        vat_rate = int(m.get('vat_rate', 21))
        item_type = m.get('item_type', 'arbeid')
        
        total_excl = quantity * unit_price
        vat_amount = total_excl * (vat_rate / 100)
        total_incl = total_excl + vat_amount
        
        line_item_doc = {
            "id": str(uuid.uuid4()),
            "quote_id": quote_id,  # Link to quote
            "description": m.get('title', ''),
            "quantity": quantity,
            "unit": m.get('unit', ''),
            "unit_price": unit_price,
            "vat_rate": float(vat_rate),
            "item_type": item_type,
            "total_excl_vat": total_excl,
            "vat_amount": vat_amount,
            "total_incl_vat": total_incl,
            "total": total_incl,  # Backwards compatibility
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        line_items_to_insert.append(line_item_doc)
    
    # Insert all line items into the separate collection
    if line_items_to_insert:
        await db.line_items.insert_many(line_items_to_insert)
    
    # Calculate totals from inserted items
    subtotal_labor = sum(item['total_excl_vat'] for item in line_items_to_insert if item['item_type'] == 'arbeid')
    subtotal_material = sum(item['total_excl_vat'] for item in line_items_to_insert if item['item_type'] != 'arbeid')
    total_excl_vat = subtotal_labor + subtotal_material
    
    # VAT breakdown
    vat_breakdown = {}
    for item in line_items_to_insert:
        vat_key = str(int(item['vat_rate']))
        if vat_key not in vat_breakdown:
            vat_breakdown[vat_key] = 0.0
        vat_breakdown[vat_key] += item['vat_amount']
    
    total_vat = sum(vat_breakdown.values())
    total_incl_vat = total_excl_vat + total_vat
    
    # Create quote document (line_items empty - they're in separate collection)
    quote = {
        "id": quote_id,
        "lead_id": lead_id,
        "project_id": project_id,
        "quote_number": quote_id,
        "date": datetime.now(timezone.utc).isoformat(),
        "status": "concept",
        "line_items": [],  # Empty - items are in separate collection for editing
        "subtotal_labor": subtotal_labor,
        "subtotal_material": subtotal_material,
        "total_excl_vat": total_excl_vat,
        "vat_breakdown": vat_breakdown,
        "total_vat": total_vat,
        "total_incl_vat": total_incl_vat,
        "total_price": total_incl_vat,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "user_id": current_user.id
    }
    
    await db.quotes.insert_one(quote)
    
    # Update project with quote_id and CLEAR measurements
    await db.projects.update_one(
        {"id": project_id},
        {
            "$set": {"quote_id": quote_id},
            "$unset": {"measurements": ""}  # Clear measurements after generating quote
        }
    )
    
    return {
        "message": "Quote generated successfully",
        "quote_id": quote_id,
        "total_incl_vat": total_incl_vat,
        "line_items_count": len(line_items_to_insert),
        "measurements_cleared": True
    }

@api_router.post("/projects/{project_id}/generate-quote-from-analysis/{analysis_id}")
async def generate_quote_from_floor_plan_analysis(
    project_id: str, 
    analysis_id: str, 
    current_user: User = Depends(get_current_user)
):
    """Generate quote from floor plan analysis - creates individual editable line items"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can generate quotes")
    
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Find the specific floor plan analysis
    floor_plan_analyses = project.get('floor_plan_analyses', [])
    analysis = next((a for a in floor_plan_analyses if a.get('id') == analysis_id), None)
    
    if not analysis:
        raise HTTPException(status_code=404, detail="Floor plan analysis not found")
    
    surfaces = analysis.get('surfaces', [])
    if not surfaces:
        raise HTTPException(status_code=400, detail="No surfaces in analysis to generate quote from")
    
    # Check if any surface has work items
    has_work_items = any(s.get('work_items', []) for s in surfaces)
    if not has_work_items:
        raise HTTPException(status_code=400, detail="No work items in analysis to generate quote from")
    
    # Check if lead exists
    lead_id = project.get('lead_id')
    if not lead_id:
        raise HTTPException(status_code=400, detail="Project must have a lead to generate quote")
    
    # Create quote first
    quote_id = f"OFF-{datetime.now(timezone.utc).year}-{str(uuid.uuid4())[:6].upper()}"
    
    # Create line items from floor plan analysis
    line_items_to_insert = []
    
    for surface in surfaces:
        surface_area = float(surface.get('net_area_m2') or surface.get('area_m2') or 0)
        surface_title = surface.get('title', 'Oppervlak')
        
        for wi in surface.get('work_items', []):
            # Use custom_area if set, otherwise use surface area
            work_area = float(wi.get('custom_area', surface_area))
            unit_price = float(wi.get('price', 0))
            vat_rate = int(wi.get('vat_rate', 6))
            
            total_excl = work_area * unit_price
            vat_amount = total_excl * (vat_rate / 100)
            total_incl = total_excl + vat_amount
            
            # Create description with surface info
            description = f"{wi.get('title', '')} - {surface_title}"
            
            line_item_doc = {
                "id": str(uuid.uuid4()),
                "quote_id": quote_id,
                "description": description,
                "quantity": work_area,
                "unit": wi.get('unit', 'm²'),
                "unit_price": unit_price,
                "vat_rate": float(vat_rate),
                "item_type": "arbeid",
                "total_excl_vat": total_excl,
                "vat_amount": vat_amount,
                "total_incl_vat": total_incl,
                "total": total_incl,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            line_items_to_insert.append(line_item_doc)
    
    if not line_items_to_insert:
        raise HTTPException(status_code=400, detail="No work items to add to quote")
    
    # Insert all line items into the separate collection
    await db.line_items.insert_many(line_items_to_insert)
    
    # Calculate totals
    total_excl_vat = sum(item['total_excl_vat'] for item in line_items_to_insert)
    
    # VAT breakdown
    vat_breakdown = {}
    for item in line_items_to_insert:
        vat_key = str(int(item['vat_rate']))
        if vat_key not in vat_breakdown:
            vat_breakdown[vat_key] = 0.0
        vat_breakdown[vat_key] += item['vat_amount']
    
    total_vat = sum(vat_breakdown.values())
    total_incl_vat = total_excl_vat + total_vat
    
    # Get analysis title for quote
    analysis_title = analysis.get('analysis_result', {}).get('room_name') or 'Grondplan Analyse'
    
    # Create quote document
    quote = {
        "id": quote_id,
        "lead_id": lead_id,
        "project_id": project_id,
        "quote_number": quote_id,
        "date": datetime.now(timezone.utc).isoformat(),
        "status": "concept",
        "description": f"Offerte gegenereerd vanuit: {analysis_title}",
        "line_items": [],
        "subtotal_labor": total_excl_vat,
        "subtotal_material": 0,
        "total_excl_vat": total_excl_vat,
        "vat_breakdown": vat_breakdown,
        "total_vat": total_vat,
        "total_incl_vat": total_incl_vat,
        "total_price": total_incl_vat,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "user_id": current_user.id
    }
    
    await db.quotes.insert_one(quote)
    
    # Update project with quote_id
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"quote_id": quote_id}}
    )
    
    return {
        "message": "Quote generated from floor plan analysis",
        "quote_id": quote_id,
        "total_incl_vat": total_incl_vat,
        "line_items_count": len(line_items_to_insert),
        "analysis_title": analysis_title
    }

@api_router.delete("/projects/{project_id}/designs")
async def delete_design_file(
    project_id: str,
    filename: str,
    current_user: User = Depends(get_current_user)
):
    """Delete 3D design file (all admins can delete)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen designs verwijderen")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    # Find and delete file
    for design in project.get("design_3d_files", []):
        if design.get("filename") == filename or design.get("original_filename") == filename:
            # Try to delete the actual file
            file_url = design.get("url", "")
            if file_url:
                # Remove /api prefix and construct path
                relative_path = file_url.replace("/api/uploads/", "").replace("/api/static/", "")
                file_path = ROOT_DIR / "uploads" / relative_path
                if file_path.exists():
                    file_path.unlink()
                    logger.info(f"Deleted design file: {file_path}")
            
            # Update project
            await db.projects.update_one(
                {"id": project_id},
                {"$pull": {"design_3d_files": {"filename": design.get("filename")}}}
            )
            
            return {"message": "Design verwijderd"}
    
    raise HTTPException(status_code=404, detail="Design bestand niet gevonden")

# ============= ADMIN MANAGEMENT ROUTES =============

@api_router.post("/setup/first-admin")
async def create_first_admin(admin_data: AdminCreate):
    """Create the first admin account - only works if no admins exist yet"""
    # Check if any admins already exist
    admin_count = await db.users.count_documents({"role": "admin"})
    if admin_count > 0:
        raise HTTPException(status_code=403, detail="Er bestaat al een admin account. Log in om nieuwe admins toe te voegen.")
    
    # Check if username already exists
    existing_user = await db.users.find_one({"username": admin_data.username}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Gebruikersnaam is al in gebruik")
    
    # Create first admin user
    admin_id = f"ADMIN-{str(uuid.uuid4())[:8].upper()}"
    admin_doc = {
        "id": admin_id,
        "username": admin_data.username,
        "email": admin_data.email,
        "name": admin_data.name,
        "role": "admin",
        "password_hash": hash_password(admin_data.password),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one({"_id": admin_id, **admin_doc})
    
    logger.info(f"First admin account created: {admin_data.username}")
    
    # Remove password_hash from response
    admin_doc.pop("password_hash")
    
    return {"message": "Eerste admin account aangemaakt! U kunt nu inloggen.", "admin": admin_doc}

@api_router.get("/setup/emergency-create-admin")
@api_router.post("/setup/emergency-create-admin")
async def emergency_create_admin(
    username: str = Query(...),
    password: str = Query(..., min_length=6),
    email: str = Query(...),
    name: str = Query(...),
    emergency_key: str = Query(...)
):
    """Emergency admin creation when locked out"""
    EMERGENCY_KEY = os.environ.get('EMERGENCY_RESET_KEY', 'qtechnics-nood-reset-2024-Zx7pK9mN')
    
    if emergency_key != EMERGENCY_KEY:
        raise HTTPException(status_code=403, detail="Ongeldige nood-sleutel")
    
    # Check if username already exists
    existing = await db.users.find_one({"username": username})
    if existing:
        # Update existing user's password
        new_hash = hash_password(password)
        await db.users.update_one(
            {"username": username},
            {"$set": {"password_hash": new_hash, "role": "admin"}}
        )
        return {"message": f"Wachtwoord voor '{username}' is gereset", "action": "updated"}
    
    # Create new admin
    admin_doc = {
        "id": f"ADMIN-{str(uuid.uuid4())[:8].upper()}",
        "username": username,
        "email": email,
        "name": name,
        "role": "admin",
        "password_hash": hash_password(password),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(admin_doc)
    logger.info(f"Emergency admin created: {username}")
    
    return {
        "message": f"Admin '{username}' succesvol aangemaakt",
        "action": "created",
        "username": username
    }

@api_router.get("/debug/sales-check")
async def debug_sales_check():
    """DEBUG: Check why sales might be zero"""
    # Count all quotes
    total_quotes = await db.quotes.count_documents({})
    sold_quotes = await db.quotes.count_documents({"is_sold": True})
    approved_quotes = await db.quotes.count_documents({"status": "goedgekeurd"})
    
    # Count legacy documents (CORRECT collection name: legacy_documents)
    total_legacy = await db.legacy_documents.count_documents({})
    sold_legacy = await db.legacy_documents.count_documents({"is_sold": True})
    
    # Get sample of legacy docs to see their structure
    legacy_sample = await db.legacy_documents.find({}).limit(3).to_list(3)
    legacy_info = []
    for doc in legacy_sample:
        legacy_info.append({
            "name": doc.get("name"),
            "is_sold": doc.get("is_sold"),
            "total_price": doc.get("total_price"),
            "project_id": doc.get("project_id"),
            "lead_id": doc.get("lead_id"),
            "has_is_sold_field": "is_sold" in doc
        })
    
    # Calculate what SHOULD be the total
    sold_quotes_data = await db.quotes.find({"is_sold": True}, {"total_incl_vat": 1}).to_list(1000)
    quotes_total = sum(q.get("total_incl_vat", 0) for q in sold_quotes_data)
    
    sold_legacy_data = await db.legacy_documents.find({"is_sold": True}, {"total_price": 1}).to_list(1000)
    legacy_total = sum(d.get("total_price", 0) for d in sold_legacy_data)
    
    return {
        "quotes": {
            "total": total_quotes,
            "sold": sold_quotes,
            "approved": approved_quotes,
            "sold_value": quotes_total
        },
        "legacy_documents": {
            "total": total_legacy,
            "sold": sold_legacy,
            "sold_value": legacy_total,
            "sample": legacy_info
        },
        "total_sales_should_be": quotes_total + legacy_total
    }

@api_router.get("/setup/status")
async def check_setup_status():
    """Check if the system has been set up (has admins)"""
    admin_count = await db.users.count_documents({"role": "admin"})
    return {
        "has_admins": admin_count > 0,
        "admin_count": admin_count,
        "needs_setup": admin_count == 0
    }

@api_router.get("/test-login/{username}/{password}")
async def test_login_debug(username: str, password: str):
    """DEBUG: Test login credentials - returns detailed info about what's happening"""
    username = username.strip()
    password = password.strip()
    
    # Find admin exact match
    admin_exact = await db.users.find_one({"username": username, "role": "admin"})
    
    # Find admin case-insensitive
    admin_case_insensitive = await db.users.find_one({
        "username": {"$regex": f"^{username}$", "$options": "i"},
        "role": "admin"
    })
    
    admin = admin_exact or admin_case_insensitive
    
    result = {
        "username_received": username,
        "username_length": len(username),
        "password_length": len(password),
        "admin_found_exact": admin_exact is not None,
        "admin_found_case_insensitive": admin_case_insensitive is not None,
        "has_password_hash": admin.get("password_hash") is not None if admin else False,
    }
    
    if admin and admin.get("password_hash"):
        result["password_valid"] = verify_password(password, admin["password_hash"])
        result["stored_username"] = admin.get("username")
    else:
        result["password_valid"] = False
        result["stored_username"] = None
    
    return result

@api_router.post("/quotes/mark-all-approved-as-sold")
async def mark_all_approved_quotes_as_sold(current_user: User = Depends(get_current_user)):
    """Mark all approved quotes as sold (one-time migration)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can do this")
    
    # Find all approved quotes that are not yet marked as sold
    result = await db.quotes.update_many(
        {
            "status": "goedgekeurd",
            "$or": [{"is_sold": False}, {"is_sold": {"$exists": False}}]
        },
        {"$set": {"is_sold": True, "sold_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Also update legacy documents
    legacy_result = await db.legacy_documents.update_many(
        {
            "status": "goedgekeurd",
            "$or": [{"is_sold": False}, {"is_sold": {"$exists": False}}]
        },
        {"$set": {"is_sold": True, "sold_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    logger.info(f"Marked {result.modified_count} quotes and {legacy_result.modified_count} legacy docs as sold")
    
    return {
        "message": f"Alle goedgekeurde offertes zijn nu als verkocht gemarkeerd",
        "quotes_updated": result.modified_count,
        "legacy_docs_updated": legacy_result.modified_count
    }

# Emergency password reset with secret key (for when locked out)
EMERGENCY_RESET_KEY = os.environ.get('EMERGENCY_RESET_KEY', 'qtechnics-nood-reset-2024-Zx7pK9mN')

@api_router.post("/setup/emergency-reset")
async def emergency_password_reset(
    username: str = Query(..., description="Username of the admin to reset"),
    new_password: str = Query(..., min_length=6, description="New password"),
    emergency_key: str = Query(..., description="Emergency reset key")
):
    """
    Emergency password reset for when locked out.
    Requires the emergency reset key from environment variable.
    """
    if emergency_key != EMERGENCY_RESET_KEY:
        logger.warning(f"Emergency reset attempted with invalid key for user: {username}")
        raise HTTPException(status_code=403, detail="Ongeldige nood-reset sleutel")
    
    # Find the admin user
    admin = await db.users.find_one({"username": username, "role": "admin"})
    if not admin:
        raise HTTPException(status_code=404, detail=f"Admin '{username}' niet gevonden")
    
    # Reset the password
    new_hash = hash_password(new_password)
    await db.users.update_one(
        {"username": username, "role": "admin"},
        {"$set": {"password_hash": new_hash}}
    )
    
    logger.info(f"Emergency password reset for admin: {username}")
    
    return {
        "success": True,
        "message": f"Wachtwoord voor '{username}' is gereset. U kunt nu inloggen.",
        "username": username
    }

@api_router.get("/setup/list-admins")
async def list_admin_usernames(emergency_key: str = Query(..., description="Emergency reset key")):
    """List all admin usernames (for emergency recovery)"""
    if emergency_key != EMERGENCY_RESET_KEY:
        raise HTTPException(status_code=403, detail="Ongeldige nood-reset sleutel")
    
    admins = await db.users.find(
        {"role": "admin"}, 
        {"_id": 0, "username": 1, "email": 1, "name": 1}
    ).to_list(100)
    
    return {"admins": admins}

@api_router.post("/admins")
async def create_admin(admin_data: AdminCreate, current_user: User = Depends(get_current_user)):
    """Create a new admin account with username/password (super admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can create admins")
    
    # Check if username already exists
    existing_user = await db.users.find_one({"username": admin_data.username}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Gebruikersnaam is al in gebruik")
    
    existing_worker = await db.workers.find_one({"username": admin_data.username}, {"_id": 0})
    if existing_worker:
        raise HTTPException(status_code=400, detail="Gebruikersnaam is al geregistreerd als werkman")
    
    # Check if email already exists
    existing_email = await db.users.find_one({"email": admin_data.email}, {"_id": 0})
    if existing_email:
        raise HTTPException(status_code=400, detail="Email is al in gebruik")
    
    # Create admin user with unique ID
    admin_id = f"ADMIN-{str(uuid.uuid4())[:8].upper()}"
    admin_doc = {
        "id": admin_id,
        "username": admin_data.username,
        "email": admin_data.email,
        "name": admin_data.name,
        "role": "admin",
        "password_hash": hash_password(admin_data.password),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one({"_id": admin_id, **admin_doc})
    
    # Remove password_hash from response
    admin_doc.pop("password_hash")
    
    return admin_doc

@api_router.get("/admins")
async def get_admins(current_user: User = Depends(get_current_user)):
    """Get all email/password admin accounts (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view admins")
    
    # Get admins with password_hash (email/password logins, not OAuth)
    admins_cursor = db.users.find(
        {"role": "admin", "password_hash": {"$exists": True}}
    )
    admins = await admins_cursor.to_list(1000)
    
    # Normalize the response - ensure each admin has an 'id' field
    result = []
    for admin in admins:
        admin_data = {
            "id": admin.get("id") or admin.get("_id"),
            "username": admin.get("username"),
            "email": admin.get("email"),
            "name": admin.get("name"),
            "role": "admin",
            "created_at": admin.get("created_at")
        }
        result.append(admin_data)
    
    return result

@api_router.delete("/admins/{admin_id}")
async def delete_admin(admin_id: str, current_user: User = Depends(get_current_user)):
    """Delete an admin (admin only, cannot delete self)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete admins")
    
    if admin_id == current_user.id:
        raise HTTPException(status_code=400, detail="Je kunt jezelf niet verwijderen")
    
    # Try to find admin by id field first, then by _id
    admin = await db.users.find_one({
        "$or": [
            {"id": admin_id, "role": "admin"},
            {"_id": admin_id, "role": "admin"}
        ],
        "password_hash": {"$exists": True}
    })
    
    if not admin:
        raise HTTPException(status_code=404, detail="Beheerder niet gevonden")
    
    # Check if trying to delete yourself (check both id and _id)
    admin_actual_id = admin.get("id") or admin.get("_id")
    if admin_actual_id == current_user.id:
        raise HTTPException(status_code=400, detail="Je kunt jezelf niet verwijderen")
    
    # Delete by the actual document identifier
    if admin.get("id"):
        result = await db.users.delete_one({"id": admin_id, "role": "admin"})
    else:
        result = await db.users.delete_one({"_id": admin_id, "role": "admin"})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kon beheerder niet verwijderen")
    
    # Also delete sessions for this admin
    await db.user_sessions.delete_many({"user_id": admin_id})
    await db.sessions.delete_many({"user_id": admin_id})
    
    logger.info(f"Admin {admin_id} deleted by {current_user.id}")
    return {"message": "Beheerder succesvol verwijderd"}

@api_router.post("/admins/{admin_id}/reset-password")
async def reset_admin_password(admin_id: str, new_password: str = Query(..., min_length=6), current_user: User = Depends(get_current_user)):
    """Reset admin password (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can reset passwords")
    
    # Find admin by ID (could be email-based or custom ID)
    admin = await db.users.find_one({"id": admin_id, "password_hash": {"$exists": True}})
    if not admin:
        # Try finding by _id (for email-based admins)
        admin = await db.users.find_one({"_id": admin_id, "password_hash": {"$exists": True}})
    
    if not admin:
        raise HTTPException(status_code=404, detail="Admin not found")
    
    # Hash new password
    new_hash = hash_password(new_password)
    
    # Update using whichever field matches
    if admin.get("id"):
        await db.users.update_one({"id": admin_id}, {"$set": {"password_hash": new_hash}})
    else:
        await db.users.update_one({"_id": admin_id}, {"$set": {"password_hash": new_hash}})
    
    # Invalidate existing sessions
    await db.sessions.delete_many({"user_id": admin_id})
    await db.user_sessions.delete_many({"user_id": admin_id})
    
    logger.info(f"Password reset for admin {admin_id} by admin {current_user.id}")
    return {"message": "Password reset successfully", "admin_name": admin.get("name", "")}

# ============= BILLIT / PEPPOL INTEGRATION =============
# Billit API Documentation: https://docs.billit.be/docs/create-first-invoice
# Production: https://api.billit.io | Sandbox: https://api.sandbox-billit.xyz
#
# ARCHITECTURE:
# - Billit is the ONLY master for legal invoicing and PEPPOL delivery
# - Emergent generates invoice data and business logic
# - Billit determines delivery channel: B2B with VAT → PEPPOL, Private → PDF/Email

BILLIT_API_KEY = os.environ.get("BILLIT_API_KEY", "")
BILLIT_BASE_URL = os.environ.get("BILLIT_BASE_URL", "https://api.sandbox.billit.be")
COMPANY_VAT = os.environ.get("COMPANY_VAT", "BE0891533928")
WEBHOOK_SECRET = os.environ.get("WEBHOOK_SECRET", "")

# Billit/PEPPOL Status mapping - Extended for all transport types
BILLIT_STATUS_MAP = {
    "not_sent": "Niet verzonden",
    "pending": "In wachtrij",
    "sending": "Wordt verzonden",
    "sent": "Verzonden",
    "sent_peppol": "Verzonden via PEPPOL",
    "sent_smtp": "Verzonden via E-mail",
    "sent_email": "Verzonden via E-mail",  # Alias
    "sent_post": "Verzonden via Post",
    "delivered": "Afgeleverd",
    "delivered_peppol": "Afgeleverd via PEPPOL",
    "delivered_smtp": "Afgeleverd via E-mail",
    "delivered_email": "Afgeleverd via E-mail",  # Alias
    "rejected": "Geweigerd",
    "paid": "Betaald",
    "failed": "Mislukt",
    "error": "Fout"
}

# Helper to get user-friendly status text
def get_billit_status_text(status: str) -> str:
    return BILLIT_STATUS_MAP.get(status, status or "Onbekend")

async def get_billit_order_status(order_id: int) -> dict:
    """Get the current status of an order from Billit"""
    async with httpx.AsyncClient() as client:
        headers = {
            "apiKey": BILLIT_API_KEY,  # Billit uses apiKey header, not Bearer token
            "Accept": "application/json"
        }
        
        response = await client.get(
            f"{BILLIT_BASE_URL}/v1/orders/{order_id}",
            headers=headers,
            timeout=30.0
        )
        
        if response.status_code == 200:
            return response.json()
        return None

async def create_billit_order(invoice_data: dict) -> dict:
    """Create an order/invoice in Billit via /v1/orders endpoint.
    Returns the unique OrderID on success."""
    async with httpx.AsyncClient() as client:
        headers = {
            "apiKey": BILLIT_API_KEY,  # Billit uses apiKey header, not Bearer token
            "Content-Type": "application/json",
            "Accept": "application/json"
        }
        
        logger.info(f"Creating Billit order at {BILLIT_BASE_URL}/v1/orders")
        logger.debug(f"Request data: {json.dumps(invoice_data, indent=2)}")
        
        response = await client.post(
            f"{BILLIT_BASE_URL}/v1/orders",
            json=invoice_data,
            headers=headers,
            timeout=30.0
        )
        
        logger.info(f"Billit response status: {response.status_code}")
        logger.debug(f"Billit response: {response.text}")
        
        if response.status_code not in [200, 201]:
            error_text = response.text
            try:
                error_json = response.json()
                if "errors" in error_json:
                    error_details = ", ".join([e.get("Description", e.get("Code", str(e))) for e in error_json["errors"]])
                    error_text = error_details
            except:
                pass
            raise HTTPException(
                status_code=response.status_code,
                detail=f"Billit API error: {error_text}"
            )
        
        # Response is the unique OrderID (integer)
        order_id = response.json()
        return {"orderId": order_id}

async def send_billit_command(order_id: int, transport_type: str = "Peppol") -> dict:
    """Send the invoice via specified transport using /v1/orders/commands/send endpoint.
    TransportTypes: 'Peppol', 'Email', 'Post', etc."""
    async with httpx.AsyncClient() as client:
        headers = {
            "apiKey": BILLIT_API_KEY,  # Billit uses apiKey header, not Bearer token
            "Content-Type": "application/json",
            "Accept": "application/json"
        }
        
        # Correct endpoint: /v1/orders/commands/send
        # Body format: {"Transporttype": "Peppol", "OrderIDs": [1234]}
        send_data = {
            "Transporttype": transport_type,
            "OrderIDs": [order_id]
        }
        
        logger.info(f"Sending Billit command to {BILLIT_BASE_URL}/v1/orders/commands/send")
        logger.debug(f"Send data: {send_data}")
        
        response = await client.post(
            f"{BILLIT_BASE_URL}/v1/orders/commands/send",
            json=send_data,
            headers=headers,
            timeout=30.0
        )
        
        logger.info(f"Billit send response status: {response.status_code}")
        logger.debug(f"Billit send response: {response.text}")
        
        if response.status_code not in [200, 201]:
            error_text = response.text
            try:
                error_json = response.json()
                if "errors" in error_json:
                    error_details = ", ".join([e.get("Description", e.get("Code", str(e))) for e in error_json["errors"]])
                    error_text = error_details
            except:
                pass
            raise HTTPException(
                status_code=response.status_code,
                detail=f"Billit send error: {error_text}"
            )
        
        return response.json() if response.text else {"status": "sent"}

def transform_invoice_to_billit(invoice: dict, lead: dict, project: dict) -> dict:
    """Transform Qtechnics invoice to Billit API format.
    Based on: https://docs.billit.be/docs/create-first-invoice
    """
    # Calculate dates
    invoice_date = invoice.get("invoice_date")
    if isinstance(invoice_date, str):
        invoice_date = datetime.fromisoformat(invoice_date.replace('Z', '+00:00'))
    due_date = invoice.get("due_date")
    if isinstance(due_date, str):
        due_date = datetime.fromisoformat(due_date.replace('Z', '+00:00'))
    
    # Build order lines for Billit
    order_lines = []
    for item in invoice.get("line_items", []):
        order_lines.append({
            "Quantity": float(item.get("quantity", 1)),
            "UnitPriceExcl": float(item.get("unit_price", 0)),
            "Description": item.get("description", ""),
            "VATPercentage": float(item.get("vat_rate", 21))
        })
    
    # Parse address into components (simple split by comma or newline)
    address_str = lead.get("address", "")
    address_parts = address_str.replace('\n', ',').split(',')
    street = address_parts[0].strip() if address_parts else ""
    city = address_parts[1].strip() if len(address_parts) > 1 else ""
    zipcode = ""
    
    # Try to extract zipcode from city (e.g., "9000 Gent")
    city_parts = city.split(' ', 1)
    if city_parts and city_parts[0].isdigit():
        zipcode = city_parts[0]
        city = city_parts[1] if len(city_parts) > 1 else ""
    
    # Build the Billit order object
    billit_order = {
        "OrderType": "Invoice",
        "OrderDirection": "Income",  # Sales invoice
        "OrderNumber": invoice.get("invoice_number", ""),
        "OrderDate": invoice_date.strftime("%Y-%m-%d") if invoice_date else datetime.now().strftime("%Y-%m-%d"),
        "ExpiryDate": due_date.strftime("%Y-%m-%d") if due_date else (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d"),
        "Reference": project.get("name", ""),  # Project reference
        "PaymentReference": invoice.get("payment_reference", ""),  # OGM reference
        "Currency": "EUR",
        "Customer": {
            "Name": lead.get("name", ""),
            "PartyType": "Customer",
            "Email": lead.get("email", ""),  # Email at customer level for SMTP transport
            "Addresses": [
                {
                    "AddressType": "InvoiceAddress",
                    "Name": lead.get("name", ""),
                    "Street": street,
                    "City": city,
                    "Zipcode": zipcode,
                    "Email": lead.get("email", ""),
                    "CountryCode": "BE"
                }
            ]
        },
        "OrderLines": order_lines
    }
    
    # Add VAT number if customer is a business (B2B)
    if lead.get("vat_number"):
        vat_number = lead["vat_number"].replace(" ", "").upper()
        billit_order["Customer"]["VATNumber"] = vat_number
        
        # Add Peppol identifier for B2B routing
        # For Belgium: Use "CBE" (Belgian National Company Number) or rely on VAT alone
        # Billit transforms VAT to correct Peppol identifier automatically
        if vat_number.startswith("BE"):
            enterprise_number = vat_number[2:]  # Remove 'BE' prefix
            billit_order["Customer"]["Identifiers"] = [
                {
                    "IdentifierType": "CBE",  # Belgian National Company Number (KBO/CBE)
                    "Identifier": enterprise_number
                }
            ]
    
    return billit_order

@api_router.post("/invoices/{invoice_id}/send-to-billit")
async def send_invoice_to_billit(invoice_id: str, current_user: User = Depends(get_current_user)):
    """Send an invoice to Billit - automatically determines transport type.
    
    Architecture (as per user specification):
    - Emergent generates invoice data (source of truth for business logic)
    - Billit is the legal "master" for invoicing
    - Billit automatically determines delivery: B2B with VAT → PEPPOL, B2C → Email/PDF
    
    Process:
    1. Create order in Billit via /v1/orders
    2. Billit determines the appropriate TransportType based on customer VAT
    3. Track status via webhooks
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen facturen versturen")
    
    # Get invoice from database
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Factuur niet gevonden")
    
    # Check if already sent successfully
    if invoice.get("peppol_status") in ["sent", "delivered", "sent_peppol", "sent_email"]:
        raise HTTPException(status_code=400, detail="Factuur is al verstuurd via Billit")
    
    # Get project and lead info
    project = await db.projects.find_one({"id": invoice["project_id"]}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    lead = await db.leads.find_one({"id": project.get("lead_id")}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Klant niet gevonden")
    
    # Determine transport type based on customer VAT status
    # Billit transport types: Peppol (B2B with VAT), SMTP (Email for B2C)
    has_vat = bool(lead.get("vat_number"))
    transport_type = "Peppol" if has_vat else "SMTP"  # SMTP = Email transport
    
    # Validate email for non-VAT customers
    if not has_vat and not lead.get("email"):
        raise HTTPException(
            status_code=400, 
            detail="Klant heeft geen BTW-nummer en geen e-mailadres. Voeg een e-mailadres toe om via e-mail te versturen."
        )
    
    try:
        # Transform invoice to Billit format
        billit_data = transform_invoice_to_billit(invoice, lead, project)
        
        # Update status to sending
        await db.invoices.update_one(
            {"id": invoice_id},
            {"$set": {
                "peppol_status": "sending",
                "peppol_transport_type": transport_type,
                "peppol_error": None  # Clear any previous error
            }}
        )
        
        # Step 1: Create order in Billit
        logger.info(f"Creating Billit order for invoice {invoice_id} (Transport: {transport_type})")
        billit_response = await create_billit_order(billit_data)
        billit_order_id = billit_response.get("orderId")
        
        if not billit_order_id:
            raise Exception("Billit heeft geen order ID teruggegeven")
        
        logger.info(f"Billit order created with ID: {billit_order_id}")
        
        # Step 2: Send via determined transport type
        logger.info(f"Sending order {billit_order_id} via {transport_type}")
        send_response = await send_billit_command(billit_order_id, transport_type=transport_type)
        
        # Determine final status based on transport
        # Map SMTP to email for user-friendly status
        status_transport = "email" if transport_type == "SMTP" else transport_type.lower()
        final_status = f"sent_{status_transport}"
        
        # Update invoice with Billit info
        await db.invoices.update_one(
            {"id": invoice_id},
            {"$set": {
                "peppol_status": final_status,
                "billit_order_id": billit_order_id,
                "peppol_transport_type": "Email" if transport_type == "SMTP" else transport_type,
                "peppol_sent_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        logger.info(f"Invoice {invoice_id} sent via {transport_type}, Billit Order ID: {billit_order_id}")
        
        # User-friendly message based on transport type
        if transport_type == "Peppol":
            message = f"Factuur verstuurd via PEPPOL naar {lead.get('vat_number')}"
        else:
            message = f"Factuur verstuurd via e-mail naar {lead.get('email')}"
        
        return {
            "success": True,
            "message": message,
            "billit_order_id": billit_order_id,
            "transport_type": "Email" if transport_type == "SMTP" else transport_type,  # User-friendly name
            "customer_vat": lead.get("vat_number"),
            "customer_email": lead.get("email") if transport_type == "SMTP" else None
        }
        
    except HTTPException as he:
        # Update status to failed for HTTPException from Billit API
        error_message = he.detail if hasattr(he, 'detail') else str(he)
        await db.invoices.update_one(
            {"id": invoice_id},
            {"$set": {
                "peppol_status": "failed",
                "peppol_error": error_message,
                "peppol_failed_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        logger.error(f"Billit API error for invoice {invoice_id}: {error_message}")
        raise
    except Exception as e:
        error_message = str(e)
        # Update status to failed with detailed error
        await db.invoices.update_one(
            {"id": invoice_id},
            {"$set": {
                "peppol_status": "failed",
                "peppol_error": error_message,
                "peppol_failed_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        logger.error(f"Billit send failed for invoice {invoice_id}: {error_message}")
        raise HTTPException(status_code=500, detail=f"Verzending mislukt: {error_message}")

@api_router.post("/invoices/{invoice_id}/send-peppol")
async def send_invoice_via_peppol(invoice_id: str, current_user: User = Depends(get_current_user)):
    """Legacy endpoint - redirects to new smart send endpoint.
    Kept for backwards compatibility with existing frontend.
    """
    return await send_invoice_to_billit(invoice_id, current_user)

@api_router.post("/invoices/{invoice_id}/retry-billit")
async def retry_billit_send(invoice_id: str, current_user: User = Depends(get_current_user)):
    """Retry sending a failed invoice to Billit."""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen facturen opnieuw versturen")
    
    # Get invoice
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Factuur niet gevonden")
    
    # Only allow retry for failed invoices
    if invoice.get("peppol_status") not in ["failed", "rejected"]:
        raise HTTPException(
            status_code=400, 
            detail="Alleen mislukte facturen kunnen opnieuw worden verstuurd"
        )
    
    # Reset status and retry
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "peppol_status": "not_sent",
            "peppol_error": None,
            "billit_order_id": None
        }}
    )
    
    return await send_invoice_to_billit(invoice_id, current_user)

@api_router.get("/invoices/{invoice_id}/peppol-status")
async def get_peppol_status(invoice_id: str, current_user: User = Depends(get_current_user)):
    """Get detailed Billit/Peppol status for an invoice"""
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Factuur niet gevonden")
    
    status = invoice.get("peppol_status", "not_sent")
    
    return {
        "peppol_status": status,
        "status_text": get_billit_status_text(status),
        "transport_type": invoice.get("peppol_transport_type"),
        "billit_order_id": invoice.get("billit_order_id"),
        "peppol_sent_at": invoice.get("peppol_sent_at"),
        "peppol_delivered_at": invoice.get("peppol_delivered_at"),
        "peppol_failed_at": invoice.get("peppol_failed_at"),
        "peppol_error": invoice.get("peppol_error"),
        "can_retry": status in ["failed", "rejected", "error"]
    }

@api_router.post("/webhooks/billit")
async def billit_webhook(request: Request):
    """Receive webhook updates from Billit about invoice/order status.
    
    Billit sends webhooks when invoice status changes (delivered, failed, etc.)
    Configure webhook URL in Billit dashboard: Settings > Webhooks
    """
    # Verify webhook signature if configured
    if WEBHOOK_SECRET:
        signature = request.headers.get("X-Billit-Signature", "")
        body = await request.body()
        
        expected_signature = hmac.new(
            WEBHOOK_SECRET.encode(),
            body,
            hashlib.sha256
        ).hexdigest()
        
        if not hmac.compare_digest(signature, expected_signature):
            logger.warning("Invalid webhook signature")
            raise HTTPException(status_code=401, detail="Invalid signature")
    
    try:
        data = await request.json()
        logger.info(f"Billit webhook received: {json.dumps(data)}")
        
        # Billit webhook payload can have different structures
        order_id = data.get("OrderID") or data.get("orderId") or data.get("id")
        order_status = data.get("OrderStatus") or data.get("status")
        
        if order_id:
            # Find our invoice by billit_order_id
            invoice = await db.invoices.find_one(
                {"billit_order_id": order_id},
                {"_id": 0}
            )
            
            if not invoice:
                # Try string version of order_id
                invoice = await db.invoices.find_one(
                    {"billit_order_id": str(order_id)},
                    {"_id": 0}
                )
            
            if invoice:
                update_data = {}
                
                # Map Billit/Peppol status to our status
                status_lower = str(order_status).lower() if order_status else ""
                
                if status_lower in ["delivered", "accepted", "received"]:
                    update_data["peppol_status"] = "delivered"
                    update_data["peppol_delivered_at"] = datetime.now(timezone.utc).isoformat()
                elif status_lower in ["failed", "rejected", "error"]:
                    update_data["peppol_status"] = "failed"
                    update_data["peppol_error"] = data.get("ErrorMessage") or data.get("message", "Unknown error")
                elif status_lower in ["sent", "pending", "processing"]:
                    update_data["peppol_status"] = "sent"
                
                if update_data:
                    await db.invoices.update_one(
                        {"id": invoice["id"]},
                        {"$set": update_data}
                    )
                    logger.info(f"Updated invoice {invoice['id']} status: {update_data}")
            else:
                logger.warning(f"No invoice found for Billit order ID: {order_id}")
        
        return {"status": "ok"}
        
    except Exception as e:
        logger.error(f"Webhook processing error: {str(e)}")
        return {"status": "error", "message": str(e)}

# ============= CUSTOMER PORTAL ENDPOINTS =============
# These endpoints use access tokens instead of session authentication

@api_router.post("/projects/{project_id}/generate-customer-link")
async def generate_customer_access_link(project_id: str, force_new: bool = False, current_user: User = Depends(get_current_user)):
    """Generate a unique access link for customers to view their project.
    If a link already exists, return the existing one unless force_new=True.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can generate customer links")
    
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Check if a token already exists
    existing_token = project.get("customer_access_token")
    if existing_token and not force_new:
        return {
            "token": existing_token,
            "message": "Bestaande klantportaal link opgehaald",
            "is_existing": True
        }
    
    # Generate unique token (permanent - never expires)
    import secrets
    token = secrets.token_urlsafe(32)
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"customer_access_token": token}}
    )
    
    return {
        "token": token,
        "message": "Klantportaal link gegenereerd",
        "is_existing": False
    }

@api_router.get("/customer-portal/{access_token}")
async def get_customer_portal_data(access_token: str):
    """Get project data for customer portal (no authentication required, uses token)"""
    # Find project by access token
    project = await db.projects.find_one(
        {"customer_access_token": access_token},
        {"_id": 0}
    )
    
    if not project:
        raise HTTPException(status_code=404, detail="Ongeldig of verlopen toegangslink")
    
    project_id = project["id"]
    lead_id = project.get("lead_id")
    
    # Get lead info (customer name)
    lead = None
    if lead_id:
        lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    
    # Get approved quotes for this project (customers can see prices here)
    # Status can be "approved" or "goedgekeurd" (Dutch)
    # Search by both project_id AND lead_id since quotes can be linked to either
    quotes = []
    
    # First try by project_id
    project_quotes = await db.quotes.find({
        "project_id": project_id,
        "status": {"$in": ["approved", "goedgekeurd"]}
    }, {"_id": 0}).to_list(100)
    quotes.extend(project_quotes)
    
    # Also search by lead_id
    if lead_id:
        lead_quotes = await db.quotes.find({
            "lead_id": lead_id,
            "status": {"$in": ["approved", "goedgekeurd"]}
        }, {"_id": 0}).to_list(100)
        # Add quotes not already in list (avoid duplicates)
        existing_ids = {q["id"] for q in quotes}
        for q in lead_quotes:
            if q["id"] not in existing_ids:
                quotes.append(q)
    
    # Get line items for approved quotes - FILTER OUT unit prices for customers
    for quote in quotes:
        line_items = await db.line_items.find(
            {"quote_id": quote["id"]},
            {"_id": 0}
        ).to_list(1000)
        
        # Remove sensitive pricing info from line items for customer view
        customer_line_items = []
        for item in line_items:
            customer_line_items.append({
                "id": item.get("id"),
                "description": item.get("description"),
                "quantity": item.get("quantity"),
                "unit": item.get("unit"),
                "item_type": item.get("item_type"),
                # NOT including: unit_price, total_excl_vat, vat_amount, total_incl_vat
            })
        quote["line_items"] = customer_line_items
        
        # Keep only the grand total for customers
        quote_totals = {
            "total_incl_vat": quote.get("total_incl_vat") or quote.get("total_price", 0)
        }
        # Remove detailed price breakdowns
        quote.pop("subtotal_labor", None)
        quote.pop("subtotal_material", None)
        quote.pop("total_excl_vat", None)
        quote.pop("total_vat", None)
        quote.pop("vat_breakdown", None)
        quote["total_incl_vat"] = quote_totals["total_incl_vat"]
    
    # Get work slips marked as visible to customer (NO financial info)
    work_slips = await db.work_slips.find({
        "project_id": project_id,
        "visible_to_customer": True
    }, {"_id": 0}).to_list(100)
    
    # Filter sensitive data from work slips
    customer_work_slips = []
    for slip in work_slips:
        customer_work_slips.append({
            "id": slip.get("id"),
            "date": slip.get("date"),
            "work_description_nl": slip.get("work_description_nl"),
            "photos": slip.get("photos", [])
            # Explicitly NOT including: hours_worked, labor_cost, hourly_rate, etc.
        })
    
    # Build safe project data for customer (NO financial info)
    customer_project = {
        "id": project["id"],
        "name": project.get("name", ""),
        "status": project.get("status", ""),
        "start_date": project.get("start_date"),
        "end_date": project.get("end_date"),
        # First visit photos
        "first_visit_photos": project.get("first_visit_photos", []),
        # 3D designs
        "design_3d_files": project.get("design_3d_files", []),
        # Planning
        "scheduled_days": project.get("scheduled_days", []),
        "planning_start_date": project.get("planning_start_date"),
        "planning_end_date": project.get("planning_end_date"),
        # Messages and rating
        "customer_messages": project.get("customer_messages", []),
        "customer_rating": project.get("customer_rating"),
        "customer_rating_comment": project.get("customer_rating_comment")
    }
    
    return {
        "project": customer_project,
        "customer_name": lead.get("name") if lead else "Klant",
        "approved_quotes": quotes,
        "work_updates": customer_work_slips
    }

@api_router.post("/customer-portal/{access_token}/message")
async def send_customer_message(access_token: str, message: dict):
    """Customer sends a message/question about their project"""
    project = await db.projects.find_one(
        {"customer_access_token": access_token},
        {"_id": 0}
    )
    
    if not project:
        raise HTTPException(status_code=404, detail="Ongeldig toegangslink")
    
    message_text = message.get("message", "").strip()
    if not message_text:
        raise HTTPException(status_code=400, detail="Bericht mag niet leeg zijn")
    
    new_message = {
        "id": str(uuid.uuid4()),
        "message": message_text,
        "sender": "customer",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "is_from_customer": True,
        "is_read": False
    }
    
    await db.projects.update_one(
        {"id": project["id"]},
        {"$push": {"customer_messages": new_message}}
    )
    
    return {"success": True, "message_id": new_message["id"]}

@api_router.post("/customer-portal/{access_token}/rating")
async def submit_customer_rating(access_token: str, rating_data: dict):
    """Customer submits a satisfaction rating"""
    project = await db.projects.find_one(
        {"customer_access_token": access_token},
        {"_id": 0}
    )
    
    if not project:
        raise HTTPException(status_code=404, detail="Ongeldig toegangslink")
    
    rating = rating_data.get("rating")
    
    # Convert to int if string
    if isinstance(rating, str):
        try:
            rating = int(rating)
        except ValueError:
            raise HTTPException(status_code=400, detail="Rating moet een getal zijn")
    
    if not rating or not isinstance(rating, int) or rating < 1 or rating > 5:
        raise HTTPException(status_code=400, detail="Rating moet tussen 1 en 5 zijn")
    
    comment = rating_data.get("comment", "")
    if comment:
        comment = str(comment).strip()
    
    await db.projects.update_one(
        {"id": project["id"]},
        {"$set": {
            "customer_rating": rating,
            "customer_rating_comment": comment
        }}
    )
    
    logger.info(f"Customer rating saved for project {project['id']}: {rating} stars")
    
    return {"success": True, "message": "Bedankt voor uw beoordeling!"}

@api_router.post("/projects/{project_id}/customer-messages")
async def admin_send_message_to_customer(project_id: str, message: dict, current_user: User = Depends(get_current_user)):
    """Admin sends a message to customer"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can send messages")
    
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    message_text = message.get("message", "").strip()
    if not message_text:
        raise HTTPException(status_code=400, detail="Bericht mag niet leeg zijn")
    
    new_message = {
        "id": str(uuid.uuid4()),
        "message": message_text,
        "sender": current_user.username,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "is_from_customer": False,
        "is_read": True
    }
    
    await db.projects.update_one(
        {"id": project_id},
        {"$push": {"customer_messages": new_message}}
    )
    
    return {"success": True, "message_id": new_message["id"]}

@api_router.put("/projects/{project_id}/work-slips/{slip_id}/visibility")
async def toggle_work_slip_customer_visibility(
    project_id: str, 
    slip_id: str, 
    visibility: dict,
    current_user: User = Depends(get_current_user)
):
    """Toggle whether a work slip is visible to customers"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can change visibility")
    
    visible = visibility.get("visible_to_customer", False)
    
    result = await db.work_slips.update_one(
        {"id": slip_id, "project_id": project_id},
        {"$set": {"visible_to_customer": visible}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Werkbon niet gevonden")
    
    # Send email notification when making visible to customer
    if visible:
        work_slip = await db.work_slips.find_one({"id": slip_id}, {"_id": 0})
        slip_date = work_slip.get("date", "onbekende datum") if work_slip else "onbekende datum"
        if isinstance(slip_date, datetime):
            slip_date = slip_date.strftime("%d-%m-%Y")
        await send_customer_notification(
            project_id=project_id,
            subject="Nieuwe werkbon beschikbaar - Q-Technics",
            content_description=f"Er is een nieuwe werkbon ({slip_date}) toegevoegd aan uw project. U kunt deze nu bekijken in uw klantenportaal."
        )
    
    return {"success": True, "visible_to_customer": visible}

# ============= QUICK TASKS ENDPOINTS =============

@api_router.get("/quick-tasks")
async def get_quick_tasks(current_user: User = Depends(get_current_user)):
    """Get all quick tasks for team planning"""
    tasks = await db.quick_tasks.find({"user_id": current_user.id}, {"_id": 0}).to_list(1000)
    return tasks

# ============= PROJECT ROOMS & RENOVATION CALCULATOR =============

@api_router.post("/projects/{project_id}/project-rooms")
async def add_project_room(project_id: str, room: PropertyRoomCreate, current_user: User = Depends(get_current_user)):
    """Add a room to a project"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can modify projects")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    room_obj = PropertyRoom(
        name=room.name, room_type=room.room_type,
        length=room.length, width=room.width, height=room.height,
        windows=room.windows, doors=room.doors, notes=room.notes
    )
    room_obj.floor_area = room_obj.length * room_obj.width
    room_obj.ceiling_area = room_obj.length * room_obj.width
    room_obj.wall_area = 2 * (room_obj.length + room_obj.width) * room_obj.height
    
    await db.projects.update_one(
        {"id": project_id},
        {"$push": {"rooms": room_obj.model_dump()}}
    )
    return {"message": "Kamer toegevoegd", "room_id": room_obj.id}

@api_router.delete("/projects/{project_id}/project-rooms/{room_id}")
async def delete_project_room(project_id: str, room_id: str, current_user: User = Depends(get_current_user)):
    """Delete a room from a project"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can modify projects")
    
    await db.projects.update_one(
        {"id": project_id},
        {"$pull": {"rooms": {"id": room_id}}}
    )
    return {"message": "Kamer verwijderd"}

@api_router.post("/projects/{project_id}/project-rooms/bulk")
async def add_project_rooms_bulk(project_id: str, rooms: List[PropertyRoomCreate], current_user: User = Depends(get_current_user)):
    """Add multiple rooms to a project at once"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can modify projects")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    new_rooms = []
    for room in rooms:
        room_obj = PropertyRoom(
            name=room.name, room_type=room.room_type,
            length=room.length, width=room.width, height=room.height,
            windows=room.windows, doors=room.doors, notes=room.notes
        )
        room_obj.floor_area = room_obj.length * room_obj.width
        room_obj.ceiling_area = room_obj.length * room_obj.width
        room_obj.wall_area = 2 * (room_obj.length + room_obj.width) * room_obj.height
        new_rooms.append(room_obj.model_dump())
    
    await db.projects.update_one(
        {"id": project_id},
        {"$push": {"rooms": {"$each": new_rooms}}}
    )
    return {"message": f"{len(new_rooms)} kamers toegevoegd", "rooms_added": len(new_rooms)}

@api_router.post("/projects/{project_id}/analyze-floor-plan")
async def analyze_project_floor_plan(
    project_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload and analyze a floor plan for a project to extract room dimensions"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can analyze floor plans")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    content = await file.read()
    
    floor_plans_dir = ROOT_DIR / "uploads" / "floor_plans"
    floor_plans_dir.mkdir(parents=True, exist_ok=True)
    
    ext = file.filename.split('.')[-1] if '.' in file.filename else 'png'
    saved_filename = f"{project_id}_{uuid.uuid4().hex[:8]}.{ext}"
    file_path = floor_plans_dir / saved_filename
    
    with open(file_path, 'wb') as f:
        f.write(content)
    
    floor_plan_url = f"/api/uploads/floor_plans/{saved_filename}"
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"floor_plan_url": floor_plan_url}}
    )
    
    api_key = os.environ.get('EMERGENT_LLM_KEY')
    if not api_key:
        return {"success": True, "floor_plan_url": floor_plan_url, "rooms": [], "message": "Grondplan opgeslagen maar AI analyse niet beschikbaar"}
    
    try:
        image_base64 = base64.b64encode(content).decode('utf-8')
        
        chat = LlmChat(
            api_key=api_key,
            session_id=f"floorplan-proj-{project_id}-{uuid.uuid4()}",
            system_message="""Je bent een expert in het analyseren van grondplannen, bouwtekeningen en laser meetrapporten.

Je taak is om uit het plan/tekening ALLE kamers te identificeren met hun afmetingen.

Geef je antwoord ALLEEN als een JSON object (geen andere tekst):
{
    "rooms": [
        {
            "name": "naam van de kamer (bijv. Woonkamer, Badkamer, Slaapkamer 1)",
            "room_type": "living|bedroom|bathroom|kitchen|hallway|other",
            "length": 0.0,
            "width": 0.0,
            "height": 2.7,
            "notes": "opmerkingen"
        }
    ],
    "total_area_m2": 0.0,
    "analysis_notes": "algemene opmerkingen over het plan"
}

Regels:
- Meet alles in METERS (niet cm)
- Gebruik standaard 2.7m hoogte als niet vermeld
- room_type moet exact zijn: living, bedroom, bathroom, kitchen, hallway, of other
- Geef een Nederlandse naam voor elke kamer
- Als afmetingen niet leesbaar zijn, schat dan op basis van verhoudingen en vermeld dit in notes
- Bij lasermeetplannen: gebruik de exacte gemeten waarden"""
        ).with_model("openai", "gpt-4o")
        
        user_message = UserMessage(
            text="Analyseer dit grondplan/meetrapport en identificeer alle kamers met hun afmetingen. Geef het resultaat als JSON.",
            file_contents=[ImageContent(image_base64=image_base64)]
        )
        
        response = await chat.send_message(user_message)
        response_text = response.strip()
        
        if '```json' in response_text:
            json_str = response_text.split('```json')[1].split('```')[0].strip()
        elif '```' in response_text:
            json_str = response_text.split('```')[1].split('```')[0].strip()
        elif response_text.startswith('{'):
            json_str = response_text
        else:
            json_str = response_text
        
        result = json.loads(json_str)
        
        return {
            "success": True,
            "floor_plan_url": floor_plan_url,
            "rooms": result.get("rooms", []),
            "total_area_m2": result.get("total_area_m2", 0),
            "analysis_notes": result.get("analysis_notes", ""),
            "message": f"{len(result.get('rooms', []))} kamers gedetecteerd"
        }
    except json.JSONDecodeError:
        return {"success": True, "floor_plan_url": floor_plan_url, "rooms": [], "message": "Grondplan opgeslagen maar kon kamers niet automatisch herkennen"}
    except Exception as e:
        logger.error(f"Floor plan analysis error for project: {e}")
        return {"success": True, "floor_plan_url": floor_plan_url, "rooms": [], "message": f"Grondplan opgeslagen maar analyse mislukt: {str(e)}"}

@api_router.post("/projects/{project_id}/calculate-renovation")
async def calculate_project_renovation(project_id: str, current_user: User = Depends(get_current_user)):
    """Run renovation calculator for a project using its rooms and live DB prices"""
    logger.info(f"calculate-renovation called for project {project_id} by user {current_user.id} role={current_user.role}")
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can calculate")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    rooms = project.get("rooms", [])
    if not rooms:
        raise HTTPException(status_code=400, detail="Geen kamers gevonden. Voeg eerst kamers toe.")
    
    # Use shared calculation logic
    calculation, calc_doc, room_calculations, all_work_items = await _perform_renovation_calculation(rooms)
    
    calc_doc["project_id"] = project_id
    
    # Remove old calculation if exists
    old_calc_id = project.get("renovation_calculation_id")
    if old_calc_id:
        await db.renovation_calculations.delete_one({"id": old_calc_id})
    
    await db.renovation_calculations.insert_one(calc_doc)
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"renovation_calculation_id": calculation.id, "status": "offerte in opmaak"}}
    )
    
    return {
        "calculation_id": calculation.id,
        "total_min": calculation.total_min,
        "total_recommended": calculation.total_realistic,
        "rooms_calculated": len(calculation.room_calculations)
    }

@api_router.get("/projects/{project_id}/renovation-calculation")
async def get_project_renovation_calculation(project_id: str, current_user: User = Depends(get_current_user)):
    """Get renovation calculation for a project"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    calc_id = project.get("renovation_calculation_id")
    if not calc_id:
        raise HTTPException(status_code=404, detail="Geen berekening gevonden")
    
    calc = await db.renovation_calculations.find_one({"id": calc_id}, {"_id": 0})
    if not calc:
        raise HTTPException(status_code=404, detail="Berekening niet gevonden")
    
    return calc

@api_router.put("/projects/{project_id}/renovation-calculation/items/{item_id}")
async def toggle_project_calc_item(project_id: str, item_id: str, included: bool = True, current_user: User = Depends(get_current_user)):
    """Toggle a calculation item included/excluded for a project"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    calc_id = project.get("renovation_calculation_id")
    if not calc_id:
        raise HTTPException(status_code=404, detail="Geen berekening gevonden")
    
    calc = await db.renovation_calculations.find_one({"id": calc_id}, {"_id": 0})
    if not calc:
        raise HTTPException(status_code=404, detail="Berekening niet gevonden")
    
    # Find and update the item in any room's items
    for rc in calc.get("room_calculations", []):
        for cat in ["floor_items", "wall_items", "ceiling_items", "other_items"]:
            for item in rc.get(cat, []):
                if item.get("id") == item_id:
                    item["included"] = included
    
    # Recalculate totals
    total = 0
    for rc in calc.get("room_calculations", []):
        subtotal = 0
        for cat in ["floor_items", "wall_items", "ceiling_items", "other_items"]:
            for item in rc.get(cat, []):
                if item.get("included"):
                    subtotal += item.get("total", 0)
        rc["subtotal"] = round(subtotal, 2)
        total += subtotal
    
    calc["total_min"] = round(total, 2)
    calc["total_recommended"] = round(total * 1.1, 2)
    
    await db.renovation_calculations.update_one(
        {"id": calc_id},
        {"$set": {"room_calculations": calc["room_calculations"], "total_min": calc["total_min"], "total_recommended": calc["total_recommended"]}}
    )
    
    return {"success": True, "total_min": calc["total_min"]}

@api_router.put("/projects/{project_id}/renovation-calculation/switch-option")
async def switch_project_calc_option(project_id: str, room_id: str, option_group: str, selected_item_id: str, current_user: User = Depends(get_current_user)):
    """Switch option for a project calculation (e.g., floor finish choice)"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    calc_id = project.get("renovation_calculation_id")
    calc = await db.renovation_calculations.find_one({"id": calc_id}, {"_id": 0})
    if not calc:
        raise HTTPException(status_code=404, detail="Berekening niet gevonden")
    
    for rc in calc.get("room_calculations", []):
        if rc.get("room_id") == room_id:
            for cat in ["floor_items", "wall_items", "ceiling_items", "other_items"]:
                for item in rc.get(cat, []):
                    if item.get("option_group") == option_group:
                        item["included"] = (item.get("id") == selected_item_id)
                        item["is_selected"] = (item.get("id") == selected_item_id)
    
    # Recalculate
    total = 0
    for rc in calc.get("room_calculations", []):
        subtotal = sum(item.get("total", 0) for cat in ["floor_items", "wall_items", "ceiling_items", "other_items"] for item in rc.get(cat, []) if item.get("included"))
        rc["subtotal"] = round(subtotal, 2)
        total += subtotal
    calc["total_min"] = round(total, 2)
    calc["total_recommended"] = round(total * 1.1, 2)
    
    await db.renovation_calculations.update_one(
        {"id": calc_id},
        {"$set": {"room_calculations": calc["room_calculations"], "total_min": calc["total_min"], "total_recommended": calc["total_recommended"]}}
    )
    
    updated = await db.renovation_calculations.find_one({"id": calc_id}, {"_id": 0})
    return updated

@api_router.put("/projects/{project_id}/renovation-calculation/switch-scenario")
async def switch_project_wall_scenario(project_id: str, room_id: str, scenario: str, current_user: User = Depends(get_current_user)):
    """Switch wall scenario for a project calculation"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    calc_id = project.get("renovation_calculation_id")
    calc = await db.renovation_calculations.find_one({"id": calc_id}, {"_id": 0})
    if not calc:
        raise HTTPException(status_code=404, detail="Berekening niet gevonden")
    
    scenario_map = {
        "nieuw_pleisterwerk": "muur_scenario_a",
        "egaliseren": "muur_scenario_b",
        "gyproc": "muur_scenario_c"
    }
    active_cat = scenario_map.get(scenario, "muur_scenario_a")
    
    for rc in calc.get("room_calculations", []):
        if rc.get("room_id") == room_id:
            for item in rc.get("wall_items", []):
                cat = item.get("category", "")
                if cat.startswith("muur_scenario_"):
                    item["included"] = (cat == active_cat)
                    item["is_selected"] = (cat == active_cat)
            rc["selected_wall_scenario"] = scenario
    
    total = 0
    for rc in calc.get("room_calculations", []):
        subtotal = sum(item.get("total", 0) for cat in ["floor_items", "wall_items", "ceiling_items", "other_items"] for item in rc.get(cat, []) if item.get("included"))
        rc["subtotal"] = round(subtotal, 2)
        total += subtotal
    calc["total_min"] = round(total, 2)
    calc["total_recommended"] = round(total * 1.1, 2)
    
    await db.renovation_calculations.update_one(
        {"id": calc_id},
        {"$set": {"room_calculations": calc["room_calculations"], "total_min": calc["total_min"], "total_recommended": calc["total_recommended"]}}
    )
    
    updated = await db.renovation_calculations.find_one({"id": calc_id}, {"_id": 0})
    return updated

@api_router.post("/projects/{project_id}/generate-quote-from-calculation")
async def generate_quote_from_calculation(project_id: str, current_user: User = Depends(get_current_user)):
    """Generate a quote with line items from the renovation calculation"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can create quotes")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    calc_id = project.get("renovation_calculation_id")
    if not calc_id:
        raise HTTPException(status_code=400, detail="Geen berekening gevonden. Maak eerst een berekening.")
    
    calc = await db.renovation_calculations.find_one({"id": calc_id}, {"_id": 0})
    if not calc:
        raise HTTPException(status_code=404, detail="Berekening niet gevonden")
    
    # Ensure lead exists
    lead_id = project.get("lead_id")
    if not lead_id:
        lead_obj = {
            "id": f"LEAD-{str(uuid.uuid4())[:8].upper()}",
            "name": project.get("name", "Onbekende klant"),
            "email": "geen-email@example.com",
            "phone": "0000000000",
            "address": "",
            "project_type": "Renovatie",
            "description": f"Automatisch aangemaakt voor project: {project.get('name', project_id)}",
            "status": "offerte",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "user_id": current_user.id
        }
        await db.leads.insert_one(lead_obj)
        lead_id = lead_obj["id"]
        await db.projects.update_one({"id": project_id}, {"$set": {"lead_id": lead_id}})
    
    # Create quote
    quote_id = f"OFF-{datetime.now(timezone.utc).year}-{str(uuid.uuid4())[:6].upper()}"
    quote_doc = {
        "id": quote_id,
        "lead_id": lead_id,
        "quote_number": quote_id,
        "date": datetime.now(timezone.utc).isoformat(),
        "status": "concept",
        "line_items": [],
        "subtotal_labor": 0,
        "subtotal_material": 0,
        "total_excl_vat": 0,
        "vat_breakdown": {},
        "total_vat": 0,
        "total_incl_vat": 0,
        "total_price": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "user_id": current_user.id
    }
    await db.quotes.insert_one(quote_doc)
    
    # Create line items from all included calculation items
    line_items = []
    total_excl = 0
    
    for rc in calc.get("room_calculations", []):
        room_name = rc.get("room_name", "")
        for cat_key in ["floor_items", "wall_items", "ceiling_items", "other_items"]:
            for item in rc.get(cat_key, []):
                if not item.get("included"):
                    continue
                
                li_id = str(uuid.uuid4())
                quantity = item.get("quantity", 1)
                unit_price = item.get("unit_price", 0)
                subtotal = round(quantity * unit_price, 2)
                vat_rate = 21.0
                vat_amount = round(subtotal * vat_rate / 100, 2)
                
                line_item = {
                    "id": li_id,
                    "quote_id": quote_id,
                    "description": f"{room_name}: {item.get('title', '?')} ({item.get('quantity', 0)} {item.get('unit', 'stuk')})",
                    "quantity": quantity,
                    "unit_price": unit_price,
                    "item_type": "arbeid",
                    "vat_rate": vat_rate,
                    "total_excl_vat": subtotal,
                    "vat_amount": vat_amount,
                    "total_incl_vat": round(subtotal + vat_amount, 2),
                    "total": subtotal,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                line_items.append(line_item)
                total_excl += subtotal
    
    if line_items:
        await db.line_items.insert_many(line_items)
    
    # Update quote totals
    total_vat = round(total_excl * 0.21, 2)
    await db.quotes.update_one(
        {"id": quote_id},
        {"$set": {
            "subtotal_labor": round(total_excl, 2),
            "total_excl_vat": round(total_excl, 2),
            "total_vat": total_vat,
            "total_incl_vat": round(total_excl + total_vat, 2),
            "total_price": round(total_excl + total_vat, 2)
        }}
    )
    
    return {
        "quote_id": quote_id,
        "line_items_count": len(line_items),
        "total_excl_vat": round(total_excl, 2),
        "total_incl_vat": round(total_excl + total_vat, 2),
        "message": f"Offerte aangemaakt met {len(line_items)} regelposten"
    }

@api_router.post("/work-items/auto-save")
async def auto_save_work_item(data: dict, current_user: User = Depends(get_current_user)):
    """Auto-save a manually entered line item as a new work item for future reuse"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can save work items")
    
    title = data.get("title", "").strip()
    price = data.get("price", 0)
    unit = data.get("unit", "stuk")
    component_label = data.get("component_label", None)
    
    if not title:
        raise HTTPException(status_code=400, detail="Titel is verplicht")
    
    # Check if item already exists (case-insensitive)
    existing = await db.work_items.find_one(
        {"title": {"$regex": f"^{title}$", "$options": "i"}},
        {"_id": 0}
    )
    
    if existing:
        return {"message": "Item bestaat al", "work_item_id": existing.get("id"), "is_new": False}
    
    work_item = {
        "id": str(uuid.uuid4()),
        "title": title,
        "price": float(price),
        "unit": unit,
        "component_label": component_label,
        "room_types": ["all"],
        "auto_saved": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.id
    }
    
    await db.work_items.insert_one(work_item)
    
    return {"message": "Nieuw werk item opgeslagen", "work_item_id": work_item["id"], "is_new": True}

@api_router.post("/quick-tasks")
async def create_quick_task(task: QuickTaskCreate, current_user: User = Depends(get_current_user)):
    """Create a new quick task"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can create quick tasks")
    
    new_task = QuickTask(
        title=task.title,
        description=task.description,
        start_date=task.start_date,
        end_date=task.end_date,
        team_name=task.team_name,
        user_id=current_user.id
    )
    
    await db.quick_tasks.insert_one(new_task.model_dump())
    return new_task

@api_router.put("/quick-tasks/{task_id}")
async def update_quick_task(task_id: str, task_update: QuickTaskUpdate, current_user: User = Depends(get_current_user)):
    """Update a quick task (including team assignment)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update quick tasks")
    
    update_data = {k: v for k, v in task_update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    result = await db.quick_tasks.update_one(
        {"id": task_id, "user_id": current_user.id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Taak niet gevonden")
    
    updated_task = await db.quick_tasks.find_one({"id": task_id}, {"_id": 0})
    return updated_task

@api_router.delete("/quick-tasks/{task_id}")
async def delete_quick_task(task_id: str, current_user: User = Depends(get_current_user)):
    """Delete a quick task"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete quick tasks")
    
    result = await db.quick_tasks.delete_one({"id": task_id, "user_id": current_user.id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Taak niet gevonden")
    
    return {"success": True}

# Mount static files for uploads at /api/uploads BEFORE including router
# Note: Kubernetes ingress routes /api/* to backend, so this will be accessible
uploads_dir = ROOT_DIR / "uploads"
uploads_dir.mkdir(exist_ok=True)
app.mount("/api/uploads", StaticFiles(directory=str(uploads_dir)), name="uploads")

# ============= CELEBRATIONS ENDPOINTS =============

@api_router.get("/celebrations/pending")
async def get_pending_celebrations(current_user: User = Depends(get_current_user)):
    """Get celebrations that the current user hasn't seen yet"""
    if current_user.role != "admin":
        return []
    
    # Find celebrations where current user is not in seen_by
    celebrations = await db.celebrations.find(
        {"seen_by": {"$ne": current_user.id}},
        {"_id": 0}
    ).sort("sold_at", -1).to_list(10)
    
    return celebrations

@api_router.post("/celebrations/{celebration_id}/mark-seen")
async def mark_celebration_seen(celebration_id: str, current_user: User = Depends(get_current_user)):
    """Mark a celebration as seen by the current user"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can mark celebrations")
    
    result = await db.celebrations.update_one(
        {"id": celebration_id},
        {"$addToSet": {"seen_by": current_user.id}}
    )
    
    return {"success": True}

# ============= WORK PERIOD MATERIALS ENDPOINTS =============

@api_router.post("/projects/{project_id}/scheduled-days/{period_id}/materials")
async def add_material_to_work_period(
    project_id: str,
    period_id: str,
    material_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Add a material to a scheduled work period"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can add materials")
    
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    scheduled_days = project.get("scheduled_days", [])
    period_found = False
    
    for period in scheduled_days:
        if period.get("id") == period_id:
            period_found = True
            if "materials" not in period:
                period["materials"] = []
            
            # Add material with unique ID and order reminder date
            new_material = {
                "id": f"MAT-{str(uuid.uuid4())[:8].upper()}",
                "name": material_data.get("name", ""),
                "quantity": material_data.get("quantity", 1),
                "unit": material_data.get("unit", "stuk"),
                "notes": material_data.get("notes", ""),
                "from_catalog": material_data.get("from_catalog", False),
                "catalog_id": material_data.get("catalog_id"),
                "order_reminder_date": material_data.get("order_reminder_date"),  # When to order
                "is_ordered": False,  # Track if already ordered
                "added_at": datetime.now(timezone.utc).isoformat()
            }
            period["materials"].append(new_material)
            break
    
    if not period_found:
        raise HTTPException(status_code=404, detail="Werkperiode niet gevonden")
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"scheduled_days": scheduled_days}}
    )
    
    return {"success": True, "material": new_material}

@api_router.delete("/projects/{project_id}/scheduled-days/{period_id}/materials/{material_id}")
async def remove_material_from_work_period(
    project_id: str,
    period_id: str,
    material_id: str,
    current_user: User = Depends(get_current_user)
):
    """Remove a material from a scheduled work period"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can remove materials")
    
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    scheduled_days = project.get("scheduled_days", [])
    
    for period in scheduled_days:
        if period.get("id") == period_id:
            materials = period.get("materials", [])
            period["materials"] = [m for m in materials if m.get("id") != material_id]
            break
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"scheduled_days": scheduled_days}}
    )
    
    return {"success": True}

@api_router.put("/projects/{project_id}/scheduled-days/{period_id}/materials/{material_id}")
async def update_material_in_work_period(
    project_id: str,
    period_id: str,
    material_id: str,
    material_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Update a material in a scheduled work period (e.g., mark as ordered)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update materials")
    
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project niet gevonden")
    
    scheduled_days = project.get("scheduled_days", [])
    material_found = False
    
    for period in scheduled_days:
        if period.get("id") == period_id:
            materials = period.get("materials", [])
            for mat in materials:
                if mat.get("id") == material_id:
                    material_found = True
                    # Update allowed fields
                    if "is_ordered" in material_data:
                        mat["is_ordered"] = material_data["is_ordered"]
                    if "order_reminder_date" in material_data:
                        mat["order_reminder_date"] = material_data["order_reminder_date"]
                    if "notes" in material_data:
                        mat["notes"] = material_data["notes"]
                    break
            break
    
    if not material_found:
        raise HTTPException(status_code=404, detail="Materiaal niet gevonden")
    
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {"scheduled_days": scheduled_days}}
    )
    
    return {"success": True}

@api_router.get("/dashboard/material-reminders")
async def get_material_reminders(current_user: User = Depends(get_current_user)):
    """Get materials that need to be ordered (based on order_reminder_date)"""
    if current_user.role != "admin":
        return []
    
    now = datetime.now(timezone.utc)
    today_str = now.strftime("%Y-%m-%d")
    
    # Find all projects
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    
    reminders = []
    for project in projects:
        scheduled_days = project.get("scheduled_days", [])
        for period in scheduled_days:
            materials = period.get("materials", [])
            if not materials:
                continue
            
            # Filter materials that need ordering
            materials_to_order = []
            for mat in materials:
                # Skip if already ordered
                if mat.get("is_ordered"):
                    continue
                
                order_date = mat.get("order_reminder_date")
                if not order_date:
                    continue
                
                # Check if order date is today or in the past
                if order_date <= today_str:
                    materials_to_order.append({
                        **mat,
                        "days_overdue": (now - datetime.strptime(order_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)).days
                    })
            
            if materials_to_order:
                reminders.append({
                    "project_id": project["id"],
                    "project_name": project.get("name", "Onbekend project"),
                    "period_id": period.get("id"),
                    "period_description": period.get("description", "Geen beschrijving"),
                    "work_start_date": period.get("start_date"),
                    "materials": materials_to_order
                })
    
    # Sort by most overdue first
    reminders.sort(key=lambda x: max((m.get("days_overdue", 0) for m in x["materials"]), default=0), reverse=True)
    
    return reminders

# ============= MULTI-TENANT PLATFORM ENDPOINTS =============

# --- Helper function for tenant filtering ---
def get_tenant_filter(current_user: User, entity_type: str) -> dict:
    """Get MongoDB filter based on user role and entity type"""
    if current_user.role == "admin":
        return {}  # Admin ziet alles
    
    if entity_type == "property":
        if current_user.role == "realtor":
            return {"owner_id": current_user.id, "owner_type": "realtor"}
        if current_user.role == "investor":
            return {
                "$or": [
                    {"owner_id": current_user.id},
                    {"shared_with": current_user.id}
                ]
            }
    
    if entity_type == "subcontractor_price":
        if current_user.role == "subcontractor":
            subcontractor_id = current_user.subcontractor_id
            if subcontractor_id:
                return {"subcontractor_id": subcontractor_id}
            return {"subcontractor_id": "__none__"}  # Return nothing
    
    # Default: alleen eigen data
    return {"owner_id": current_user.id}

# --- Realtor Management (Admin only) ---

@api_router.post("/realtors")
async def create_realtor(realtor: RealtorCreate, current_user: User = Depends(get_current_user)):
    """Create a new realtor account (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can create realtors")
    
    # Check if username exists
    existing = await db.users.find_one({"username": realtor.username})
    if existing:
        raise HTTPException(status_code=400, detail="Gebruikersnaam bestaat al")
    
    # Create user account
    user_id = f"realtor-{realtor.username}"
    password_hash_value = hash_password(realtor.password)
    
    user_doc = {
        "_id": user_id,
        "id": user_id,
        "username": realtor.username,
        "email": realtor.email,
        "name": realtor.contact_name,
        "role": "realtor",
        "password_hash": password_hash_value,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user_doc)
    
    # Create realtor profile
    profile = RealtorProfile(
        user_id=user_id,
        company_name=realtor.company_name,
        contact_name=realtor.contact_name,
        email=realtor.email,
        phone=realtor.phone
    )
    
    profile_doc = profile.model_dump()
    profile_doc["created_at"] = profile_doc["created_at"].isoformat()
    
    await db.realtor_profiles.insert_one(profile_doc)
    
    # Update user with realtor_id
    await db.users.update_one({"_id": user_id}, {"$set": {"realtor_id": profile.id}})
    
    logger.info(f"Created realtor: {realtor.company_name} ({realtor.username})")
    
    return {"message": "Makelaar aangemaakt", "realtor_id": profile.id, "username": realtor.username}

@api_router.get("/realtors")
async def get_realtors(current_user: User = Depends(get_current_user)):
    """Get all realtors (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view realtors")
    
    realtors = await db.realtor_profiles.find({}, {"_id": 0}).to_list(1000)
    return realtors

@api_router.delete("/realtors/{realtor_id}")
async def delete_realtor(realtor_id: str, current_user: User = Depends(get_current_user)):
    """Delete a realtor (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete realtors")
    
    profile = await db.realtor_profiles.find_one({"id": realtor_id})
    if not profile:
        raise HTTPException(status_code=404, detail="Makelaar niet gevonden")
    
    # Delete user account
    await db.users.delete_one({"_id": profile["user_id"]})
    # Delete profile
    await db.realtor_profiles.delete_one({"id": realtor_id})
    # Delete sessions
    await db.sessions.delete_many({"user_id": profile["user_id"]})
    
    return {"message": "Makelaar verwijderd"}

# --- Investor Management (Admin only) ---

@api_router.post("/investors")
async def create_investor(investor: InvestorCreate, current_user: User = Depends(get_current_user)):
    """Create a new investor account (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can create investors")
    
    # Check if username exists
    existing = await db.users.find_one({"username": investor.username})
    if existing:
        raise HTTPException(status_code=400, detail="Gebruikersnaam bestaat al")
    
    # Create user account
    user_id = f"investor-{investor.username}"
    password_hash_value = hash_password(investor.password)
    
    user_doc = {
        "_id": user_id,
        "id": user_id,
        "username": investor.username,
        "email": investor.email,
        "name": investor.name,
        "role": "investor",
        "password_hash": password_hash_value,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user_doc)
    
    # Create investor profile
    profile = InvestorProfile(
        user_id=user_id,
        name=investor.name,
        email=investor.email,
        phone=investor.phone,
        target_roi=investor.target_roi
    )
    
    profile_doc = profile.model_dump()
    profile_doc["created_at"] = profile_doc["created_at"].isoformat()
    
    await db.investor_profiles.insert_one(profile_doc)
    
    # Update user with investor_id
    await db.users.update_one({"_id": user_id}, {"$set": {"investor_id": profile.id}})
    
    logger.info(f"Created investor: {investor.name} ({investor.username})")
    
    return {"message": "Investeerder aangemaakt", "investor_id": profile.id, "username": investor.username}

@api_router.get("/investors")
async def get_investors(current_user: User = Depends(get_current_user)):
    """Get all investors (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view investors")
    
    investors = await db.investor_profiles.find({}, {"_id": 0}).to_list(1000)
    return investors

# --- Subcontractor Management (Admin only) ---

@api_router.post("/subcontractors")
async def create_subcontractor(subcontractor: SubcontractorCreate, current_user: User = Depends(get_current_user)):
    """Create a new subcontractor (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can create subcontractors")
    
    # Create subcontractor profile
    sub = Subcontractor(
        company_name=subcontractor.company_name,
        contact_name=subcontractor.contact_name,
        email=subcontractor.email,
        phone=subcontractor.phone,
        vat_number=subcontractor.vat_number,
        category=subcontractor.category
    )
    
    # If password provided, create user account
    if subcontractor.password:
        username = subcontractor.email.split("@")[0].lower()
        user_id = f"sub-{username}"
        password_hash_value = hash_password(subcontractor.password)
        
        user_doc = {
            "_id": user_id,
            "id": user_id,
            "username": username,
            "email": subcontractor.email,
            "name": subcontractor.contact_name,
            "role": "subcontractor",
            "password_hash": password_hash_value,
            "subcontractor_id": sub.id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.users.insert_one(user_doc)
        sub.user_id = user_id
    
    sub_doc = sub.model_dump()
    sub_doc["created_at"] = sub_doc["created_at"].isoformat()
    
    await db.subcontractors.insert_one(sub_doc)
    
    logger.info(f"Created subcontractor: {subcontractor.company_name} ({subcontractor.category})")
    
    return {"message": "Onderaannemer aangemaakt", "subcontractor_id": sub.id}

@api_router.get("/subcontractors")
async def get_subcontractors(current_user: User = Depends(get_current_user)):
    """Get all subcontractors (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view subcontractors")
    
    subcontractors = await db.subcontractors.find({}, {"_id": 0}).to_list(1000)
    return subcontractors

@api_router.post("/subcontractors/{subcontractor_id}/prices")
async def add_subcontractor_price(subcontractor_id: str, price: SubcontractorPriceCreate, current_user: User = Depends(get_current_user)):
    """Add a price to a subcontractor (admin or subcontractor owner)"""
    # Check access
    if current_user.role == "subcontractor":
        if current_user.subcontractor_id != subcontractor_id:
            raise HTTPException(status_code=403, detail="Geen toegang tot deze onderaannemer")
    elif current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    # Check subcontractor exists
    sub = await db.subcontractors.find_one({"id": subcontractor_id})
    if not sub:
        raise HTTPException(status_code=404, detail="Onderaannemer niet gevonden")
    
    price_obj = SubcontractorPrice(
        subcontractor_id=subcontractor_id,
        title=price.title,
        category=price.category,
        price_type=price.price_type,
        price=price.price,
        price_min=price.price_min,
        price_max=price.price_max
    )
    
    price_doc = price_obj.model_dump()
    price_doc["created_at"] = price_doc["created_at"].isoformat()
    price_doc["updated_at"] = price_doc["updated_at"].isoformat()
    
    await db.subcontractor_prices.insert_one(price_doc)
    
    return {"message": "Prijs toegevoegd", "price_id": price_obj.id}

@api_router.get("/subcontractors/{subcontractor_id}/prices")
async def get_subcontractor_prices(subcontractor_id: str, current_user: User = Depends(get_current_user)):
    """Get prices for a subcontractor"""
    # Check access
    if current_user.role == "subcontractor":
        if current_user.subcontractor_id != subcontractor_id:
            raise HTTPException(status_code=403, detail="Geen toegang tot deze onderaannemer")
    elif current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    prices = await db.subcontractor_prices.find({"subcontractor_id": subcontractor_id}, {"_id": 0}).to_list(1000)
    return prices

# --- Property Management (Realtors & Investors) ---

@api_router.post("/properties")
async def create_property(prop: PropertyCreate, current_user: User = Depends(get_current_user)):
    """Create a new property"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang tot panden")
    
    # Determine owner type
    owner_type = current_user.role if current_user.role in ["realtor", "investor"] else "admin"
    
    # Check property limit for realtors
    if current_user.role == "realtor":
        realtor_profile = await db.realtor_profiles.find_one({"user_id": current_user.id})
        if realtor_profile:
            if realtor_profile.get("properties_used", 0) >= realtor_profile.get("property_limit", 5):
                raise HTTPException(status_code=403, detail="Pand limiet bereikt. Upgrade naar een hoger abonnement.")
    
    # Process rooms - calculate areas
    rooms = []
    for room_data in prop.rooms:
        room = PropertyRoom(
            name=room_data.name,
            room_type=room_data.room_type,
            length=room_data.length,
            width=room_data.width,
            height=room_data.height,
            windows=room_data.windows,
            doors=room_data.doors,
            notes=room_data.notes
        )
        # Calculate areas
        room.floor_area = room.length * room.width
        room.ceiling_area = room.length * room.width
        room.wall_area = 2 * (room.length + room.width) * room.height
        rooms.append(room)
    
    property_obj = Property(
        owner_type=owner_type,
        owner_id=current_user.id,
        source_url=prop.source_url,
        source_platform="manual",
        address=prop.address,
        postal_code=prop.postal_code,
        city=prop.city,
        living_area=prop.living_area,
        plot_area=prop.plot_area,
        bedrooms=prop.bedrooms,
        bathrooms=prop.bathrooms,
        construction_year=prop.construction_year,
        epc_score=prop.epc_score,
        epc_value=prop.epc_value,
        asking_price=prop.asking_price,
        rooms=rooms,
        status="imported"
    )
    
    prop_doc = property_obj.model_dump()
    prop_doc["created_at"] = prop_doc["created_at"].isoformat()
    prop_doc["updated_at"] = prop_doc["updated_at"].isoformat()
    # Convert rooms to dicts
    prop_doc["rooms"] = [r.model_dump() for r in rooms]
    
    await db.properties.insert_one(prop_doc)
    
    # Update realtor property count
    if current_user.role == "realtor":
        await db.realtor_profiles.update_one(
            {"user_id": current_user.id},
            {"$inc": {"properties_used": 1}}
        )
    
    logger.info(f"Created property: {prop.address} by {current_user.id}")
    
    return {"message": "Pand aangemaakt", "property_id": property_obj.id}

@api_router.get("/properties")
async def get_properties(current_user: User = Depends(get_current_user)):
    """Get properties based on user role"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang tot panden")
    
    filter_query = get_tenant_filter(current_user, "property")
    
    properties = await db.properties.find(filter_query, {"_id": 0}).to_list(1000)
    
    return properties

@api_router.get("/properties/{property_id}")
async def get_property(property_id: str, current_user: User = Depends(get_current_user)):
    """Get a specific property"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang tot panden")
    
    base_filter = {"id": property_id}
    tenant_filter = get_tenant_filter(current_user, "property")
    
    # Combine filters
    if tenant_filter:
        filter_query = {"$and": [base_filter, tenant_filter]}
    else:
        filter_query = base_filter
    
    prop = await db.properties.find_one(filter_query, {"_id": 0})
    
    if not prop:
        raise HTTPException(status_code=404, detail="Pand niet gevonden")
    
    return prop

@api_router.put("/properties/{property_id}")
async def update_property(property_id: str, update: PropertyUpdate, current_user: User = Depends(get_current_user)):
    """Update a property"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang tot panden")
    
    # Check ownership
    tenant_filter = get_tenant_filter(current_user, "property")
    filter_query = {"id": property_id}
    if tenant_filter:
        filter_query = {"$and": [{"id": property_id}, tenant_filter]}
    
    prop = await db.properties.find_one(filter_query)
    if not prop:
        raise HTTPException(status_code=404, detail="Pand niet gevonden")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.properties.update_one({"id": property_id}, {"$set": update_data})
    
    return {"message": "Pand bijgewerkt"}

@api_router.delete("/properties/{property_id}")
async def delete_property(property_id: str, current_user: User = Depends(get_current_user)):
    """Delete a property"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang tot panden")
    
    # Check ownership
    tenant_filter = get_tenant_filter(current_user, "property")
    filter_query = {"id": property_id}
    if tenant_filter:
        filter_query = {"$and": [{"id": property_id}, tenant_filter]}
    
    result = await db.properties.delete_one(filter_query)
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pand niet gevonden")
    
    # Update realtor property count
    if current_user.role == "realtor":
        await db.realtor_profiles.update_one(
            {"user_id": current_user.id},
            {"$inc": {"properties_used": -1}}
        )
    
    # Delete associated calculations
    await db.renovation_calculations.delete_many({"property_id": property_id})
    
    return {"message": "Pand verwijderd"}

@api_router.post("/properties/{property_id}/rooms")
async def add_property_room(property_id: str, room: PropertyRoomCreate, current_user: User = Depends(get_current_user)):
    """Add a room to a property"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    # Check ownership
    tenant_filter = get_tenant_filter(current_user, "property")
    filter_query = {"id": property_id}
    if tenant_filter:
        filter_query = {"$and": [{"id": property_id}, tenant_filter]}
    
    prop = await db.properties.find_one(filter_query)
    if not prop:
        raise HTTPException(status_code=404, detail="Pand niet gevonden")
    
    # Create room with calculated areas
    room_obj = PropertyRoom(
        name=room.name,
        room_type=room.room_type,
        length=room.length,
        width=room.width,
        height=room.height,
        windows=room.windows,
        doors=room.doors,
        notes=room.notes
    )
    room_obj.floor_area = room_obj.length * room_obj.width
    room_obj.ceiling_area = room_obj.length * room_obj.width
    room_obj.wall_area = 2 * (room_obj.length + room_obj.width) * room_obj.height
    
    await db.properties.update_one(
        {"id": property_id},
        {
            "$push": {"rooms": room_obj.model_dump()},
            "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
        }
    )
    
    return {"message": "Kamer toegevoegd", "room_id": room_obj.id}

@api_router.put("/properties/{property_id}/rooms/{room_id}")
async def update_property_room(property_id: str, room_id: str, room: PropertyRoomCreate, current_user: User = Depends(get_current_user)):
    """Update a room in a property"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    # Check ownership
    tenant_filter = get_tenant_filter(current_user, "property")
    filter_query = {"id": property_id}
    if tenant_filter:
        filter_query = {"$and": [{"id": property_id}, tenant_filter]}
    
    prop = await db.properties.find_one(filter_query)
    if not prop:
        raise HTTPException(status_code=404, detail="Pand niet gevonden")
    
    # Find and update room
    rooms = prop.get("rooms", [])
    room_found = False
    for i, r in enumerate(rooms):
        if r.get("id") == room_id:
            rooms[i] = {
                "id": room_id,
                "name": room.name,
                "room_type": room.room_type,
                "length": room.length,
                "width": room.width,
                "height": room.height,
                "floor_area": room.length * room.width,
                "ceiling_area": room.length * room.width,
                "wall_area": 2 * (room.length + room.width) * room.height,
                "windows": room.windows,
                "doors": room.doors,
                "notes": room.notes
            }
            room_found = True
            break
    
    if not room_found:
        raise HTTPException(status_code=404, detail="Kamer niet gevonden")
    
    await db.properties.update_one(
        {"id": property_id},
        {
            "$set": {
                "rooms": rooms,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    return {"message": "Kamer bijgewerkt"}

@api_router.delete("/properties/{property_id}/rooms/{room_id}")
async def delete_property_room(property_id: str, room_id: str, current_user: User = Depends(get_current_user)):
    """Delete a room from a property"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    await db.properties.update_one(
        {"id": property_id},
        {
            "$pull": {"rooms": {"id": room_id}},
            "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
        }
    )
    
    return {"message": "Kamer verwijderd"}

@api_router.post("/properties/{property_id}/rooms/bulk")
async def add_property_rooms_bulk(property_id: str, rooms: List[PropertyRoomCreate], current_user: User = Depends(get_current_user)):
    """Add multiple rooms to a property at once (e.g. from floor plan analysis)"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    tenant_filter = get_tenant_filter(current_user, "property")
    filter_query = {"id": property_id}
    if tenant_filter:
        filter_query = {"$and": [{"id": property_id}, tenant_filter]}
    
    prop = await db.properties.find_one(filter_query)
    if not prop:
        raise HTTPException(status_code=404, detail="Pand niet gevonden")
    
    new_rooms = []
    for room in rooms:
        room_obj = PropertyRoom(
            name=room.name,
            room_type=room.room_type,
            length=room.length,
            width=room.width,
            height=room.height,
            windows=room.windows,
            doors=room.doors,
            notes=room.notes
        )
        room_obj.floor_area = room_obj.length * room_obj.width
        room_obj.ceiling_area = room_obj.length * room_obj.width
        room_obj.wall_area = 2 * (room_obj.length + room_obj.width) * room_obj.height
        new_rooms.append(room_obj.model_dump())
    
    await db.properties.update_one(
        {"id": property_id},
        {
            "$push": {"rooms": {"$each": new_rooms}},
            "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
        }
    )
    
    return {"message": f"{len(new_rooms)} kamers toegevoegd", "rooms_added": len(new_rooms)}

@api_router.post("/properties/{property_id}/analyze-floor-plan")
async def analyze_property_floor_plan(
    property_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload and analyze a floor plan to extract room dimensions using GPT-4o Vision"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    tenant_filter = get_tenant_filter(current_user, "property")
    filter_query = {"id": property_id}
    if tenant_filter:
        filter_query = {"$and": [{"id": property_id}, tenant_filter]}
    
    prop = await db.properties.find_one(filter_query)
    if not prop:
        raise HTTPException(status_code=404, detail="Pand niet gevonden")
    
    # Save the floor plan file
    content = await file.read()
    
    floor_plans_dir = ROOT_DIR / "uploads" / "floor_plans"
    floor_plans_dir.mkdir(parents=True, exist_ok=True)
    
    ext = file.filename.split('.')[-1] if '.' in file.filename else 'png'
    saved_filename = f"{property_id}_{uuid.uuid4().hex[:8]}.{ext}"
    file_path = floor_plans_dir / saved_filename
    
    with open(file_path, 'wb') as f:
        f.write(content)
    
    floor_plan_url = f"/api/uploads/floor_plans/{saved_filename}"
    
    # Save URL to property
    await db.properties.update_one(
        {"id": property_id},
        {"$set": {"floor_plan_url": floor_plan_url, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Analyze with GPT-4o Vision
    api_key = os.environ.get('EMERGENT_LLM_KEY')
    if not api_key:
        return {"success": True, "floor_plan_url": floor_plan_url, "rooms": [], "message": "Grondplan opgeslagen maar AI analyse niet beschikbaar (API key ontbreekt)"}
    
    try:
        image_base64 = base64.b64encode(content).decode('utf-8')
        
        chat = LlmChat(
            api_key=api_key,
            session_id=f"floorplan-prop-{property_id}-{uuid.uuid4()}",
            system_message="""Je bent een expert in het analyseren van grondplannen, bouwtekeningen en laser meetrapporten.

Je taak is om uit het plan/tekening ALLE kamers te identificeren met hun afmetingen.

Geef je antwoord ALLEEN als een JSON object (geen andere tekst):
{
    "rooms": [
        {
            "name": "naam van de kamer (bijv. Woonkamer, Badkamer, Slaapkamer 1)",
            "room_type": "living|bedroom|bathroom|kitchen|hallway|other",
            "length": 0.0,
            "width": 0.0,
            "height": 2.7,
            "notes": "opmerkingen"
        }
    ],
    "total_area_m2": 0.0,
    "analysis_notes": "algemene opmerkingen over het plan"
}

Regels:
- Meet alles in METERS (niet cm)
- Gebruik standaard 2.7m hoogte als niet vermeld
- room_type moet exact zijn: living, bedroom, bathroom, kitchen, hallway, of other
- Geef een Nederlandse naam voor elke kamer
- Als afmetingen niet leesbaar zijn, schat dan op basis van verhoudingen en vermeld dit in notes
- Bij lasermeetplannen: gebruik de exacte gemeten waarden"""
        ).with_model("openai", "gpt-4o")
        
        user_message = UserMessage(
            text="Analyseer dit grondplan/meetrapport en identificeer alle kamers met hun afmetingen. Geef het resultaat als JSON.",
            file_contents=[ImageContent(image_base64=image_base64)]
        )
        
        response = await chat.send_message(user_message)
        response_text = response.strip()
        
        if '```json' in response_text:
            json_str = response_text.split('```json')[1].split('```')[0].strip()
        elif '```' in response_text:
            json_str = response_text.split('```')[1].split('```')[0].strip()
        elif response_text.startswith('{'):
            json_str = response_text
        else:
            json_str = response_text
        
        result = json.loads(json_str)
        
        return {
            "success": True,
            "floor_plan_url": floor_plan_url,
            "rooms": result.get("rooms", []),
            "total_area_m2": result.get("total_area_m2", 0),
            "analysis_notes": result.get("analysis_notes", ""),
            "message": f"{len(result.get('rooms', []))} kamers gedetecteerd"
        }
        
    except json.JSONDecodeError as e:
        logger.error(f"Floor plan JSON parse error: {e}")
        return {
            "success": True,
            "floor_plan_url": floor_plan_url,
            "rooms": [],
            "message": "Grondplan opgeslagen maar kon kamers niet automatisch herkennen. Voeg ze handmatig toe."
        }
    except Exception as e:
        logger.error(f"Floor plan analysis error for property: {e}")
        return {
            "success": True,
            "floor_plan_url": floor_plan_url,
            "rooms": [],
            "message": f"Grondplan opgeslagen maar analyse mislukt: {str(e)}"
        }


async def share_property(property_id: str, user_id: str = Query(...), current_user: User = Depends(get_current_user)):
    """Share a property with another user (admin or owner only)"""
    if current_user.role not in ["admin", "realtor"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    # Check ownership
    if current_user.role == "realtor":
        prop = await db.properties.find_one({"id": property_id, "owner_id": current_user.id})
    else:
        prop = await db.properties.find_one({"id": property_id})
    
    if not prop:
        raise HTTPException(status_code=404, detail="Pand niet gevonden")
    
    # Add user to shared_with
    await db.properties.update_one(
        {"id": property_id},
        {
            "$addToSet": {"shared_with": user_id},
            "$set": {
                "status": "shared",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    return {"message": "Pand gedeeld"}

# --- Property Scraping ---

class ScrapedPropertyData(BaseModel):
    address: str = ""
    postal_code: str = ""
    city: str = ""
    living_area: float = 0.0
    plot_area: float = 0.0
    bedrooms: int = 0
    bathrooms: int = 0
    construction_year: Optional[int] = None
    epc_score: Optional[str] = None
    epc_value: Optional[float] = None
    asking_price: float = 0.0
    photos: List[str] = []
    source_platform: str = "unknown"
    raw_description: str = ""

async def scrape_immoweb(url: str) -> ScrapedPropertyData:
    """Scrape property data from Immoweb.be"""
    data = ScrapedPropertyData(source_platform="immoweb")
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "nl-BE,nl;q=0.9,en;q=0.8"
    }
    
    async with httpx.AsyncClient(follow_redirects=True, timeout=30.0) as client:
        response = await client.get(url, headers=headers)
        
        if response.status_code != 200:
            logger.warning(f"Immoweb scrape failed: {response.status_code}")
            return data
        
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        
        # Try to find JSON-LD data first (most reliable)
        json_ld_scripts = soup.find_all('script', type='application/ld+json')
        for script in json_ld_scripts:
            try:
                json_data = json.loads(script.string)
                if isinstance(json_data, dict):
                    if json_data.get('@type') == 'Product' or json_data.get('@type') == 'RealEstateListing':
                        # Extract from JSON-LD
                        if 'offers' in json_data:
                            price_str = str(json_data['offers'].get('price', '0'))
                            data.asking_price = float(re.sub(r'[^\d.]', '', price_str) or 0)
                        if 'name' in json_data:
                            data.address = json_data['name']
            except:
                pass
        
        # Extract from window.classified data (Immoweb specific)
        scripts = soup.find_all('script')
        for script in scripts:
            if script.string and 'window.classified' in str(script.string):
                try:
                    # Find JSON in script
                    match = re.search(r'window\.classified\s*=\s*({.*?});', script.string, re.DOTALL)
                    if match:
                        classified_data = json.loads(match.group(1))
                        
                        # Property details
                        prop = classified_data.get('property', {})
                        data.living_area = float(prop.get('netHabitableSurface', 0) or 0)
                        data.plot_area = float(prop.get('land', {}).get('surface', 0) or 0)
                        data.bedrooms = int(prop.get('bedroomCount', 0) or 0)
                        data.bathrooms = int(prop.get('bathroomCount', 0) or 0)
                        data.construction_year = prop.get('building', {}).get('constructionYear')
                        
                        # Location
                        location = prop.get('location', {})
                        data.postal_code = str(location.get('postalCode', ''))
                        data.city = location.get('locality', '')
                        street = location.get('street', '')
                        number = location.get('number', '')
                        data.address = f"{street} {number}".strip()
                        
                        # EPC
                        certificates = prop.get('certificates', {})
                        epc_data = certificates.get('epcScore')
                        if epc_data:
                            data.epc_score = epc_data
                        
                        # Price
                        transaction = classified_data.get('transaction', {})
                        if transaction.get('sale', {}).get('price'):
                            data.asking_price = float(transaction['sale']['price'])
                        
                        # Photos
                        media = classified_data.get('media', {})
                        photos = media.get('pictures', [])
                        data.photos = [p.get('largeUrl') or p.get('mediumUrl') or p.get('smallUrl') for p in photos[:10] if p.get('largeUrl') or p.get('mediumUrl')]
                        
                except Exception as e:
                    logger.warning(f"Error parsing Immoweb data: {e}")
        
        # Fallback: parse HTML directly
        if not data.address:
            title = soup.find('h1', class_='classified__title')
            if title:
                data.address = title.get_text(strip=True)
        
        if not data.asking_price:
            price_elem = soup.find('p', class_='classified__price')
            if price_elem:
                price_text = price_elem.get_text()
                price_match = re.search(r'[\d\s,.]+', price_text)
                if price_match:
                    data.asking_price = float(re.sub(r'[^\d]', '', price_match.group()) or 0)
    
    return data

async def scrape_zimmo(url: str) -> ScrapedPropertyData:
    """Scrape property data from Zimmo.be"""
    data = ScrapedPropertyData(source_platform="zimmo")
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    }
    
    async with httpx.AsyncClient(follow_redirects=True, timeout=30.0) as client:
        response = await client.get(url, headers=headers)
        
        if response.status_code != 200:
            return data
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Title/Address
        title = soup.find('h1')
        if title:
            data.address = title.get_text(strip=True)
        
        # Price
        price_elem = soup.find('span', class_='price')
        if price_elem:
            price_text = price_elem.get_text()
            data.asking_price = float(re.sub(r'[^\d]', '', price_text) or 0)
        
        # Features - look for key-value pairs
        features = soup.find_all(['dt', 'dd', 'li'])
        for i, elem in enumerate(features):
            text = elem.get_text(strip=True).lower()
            
            if 'bewoonbare' in text or 'opp' in text:
                next_elem = features[i+1] if i+1 < len(features) else None
                if next_elem:
                    match = re.search(r'(\d+)', next_elem.get_text())
                    if match:
                        data.living_area = float(match.group(1))
            
            if 'slaapkamer' in text:
                match = re.search(r'(\d+)', text)
                if match:
                    data.bedrooms = int(match.group(1))
            
            if 'badkamer' in text:
                match = re.search(r'(\d+)', text)
                if match:
                    data.bathrooms = int(match.group(1))
            
            if 'epc' in text:
                match = re.search(r'([A-G])', text.upper())
                if match:
                    data.epc_score = match.group(1)
        
        # Photos
        img_tags = soup.find_all('img', src=True)
        for img in img_tags[:10]:
            src = img.get('src', '')
            if 'zimmo' in src and ('property' in src or 'photo' in src):
                data.photos.append(src)
    
    return data

async def scrape_immoscoop(url: str) -> ScrapedPropertyData:
    """Scrape property data from Immoscoop.be"""
    data = ScrapedPropertyData(source_platform="immoscoop")
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }
    
    async with httpx.AsyncClient(follow_redirects=True, timeout=30.0) as client:
        response = await client.get(url, headers=headers)
        
        if response.status_code != 200:
            return data
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Similar parsing logic as Zimmo
        title = soup.find('h1')
        if title:
            data.address = title.get_text(strip=True)
        
        # Look for price
        price_patterns = soup.find_all(text=re.compile(r'€\s*[\d\s.,]+'))
        for price_text in price_patterns:
            match = re.search(r'€\s*([\d\s.,]+)', str(price_text))
            if match:
                price_str = re.sub(r'[^\d]', '', match.group(1))
                if price_str and len(price_str) > 4:  # Reasonable price
                    data.asking_price = float(price_str)
                    break
    
    return data

async def scrape_generic_website(url: str) -> ScrapedPropertyData:
    """Generic scraper for any real estate website (realtor websites, etc.)"""
    data = ScrapedPropertyData(source_platform="generic")
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "nl-BE,nl;q=0.9,en;q=0.8"
    }
    
    try:
        async with httpx.AsyncClient(follow_redirects=True, timeout=30.0) as client:
            response = await client.get(url, headers=headers)
            
            if response.status_code != 200:
                logger.warning(f"Generic scrape failed: {response.status_code} for {url}")
                return data
            
            soup = BeautifulSoup(response.text, 'html.parser')
            text = soup.get_text(separator=' ')
            
            # --- Extract Title/Address ---
            # Try h1 first
            h1 = soup.find('h1')
            if h1:
                data.address = h1.get_text(strip=True)[:200]
            
            # Look for address patterns in meta tags
            meta_desc = soup.find('meta', {'name': 'description'})
            if meta_desc and meta_desc.get('content'):
                data.raw_description = meta_desc['content']
            
            # --- Extract Price ---
            price_patterns = [
                r'€\s*([\d\s.,]+)',
                r'([\d\s.,]+)\s*€',
                r'prijs[:\s]*([\d\s.,]+)',
                r'vraagprijs[:\s]*([\d\s.,]+)',
                r'price[:\s]*([\d\s.,]+)',
            ]
            
            for pattern in price_patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                for match in matches:
                    clean = re.sub(r'[^\d]', '', str(match))
                    if clean:
                        try:
                            price = int(clean)
                            # Valid price range for Belgian real estate
                            if 30000 < price < 15000000:
                                data.asking_price = float(price)
                                break
                        except:
                            pass
                if data.asking_price > 0:
                    break
            
            # --- Extract Living Area ---
            area_patterns = [
                r'bewoonbare?\s*(?:opp(?:ervlakte)?)?[:\s]*(\d+)\s*m[²2]',
                r'(\d+)\s*m[²2]\s*bewoonbaar',
                r'woonoppervlakte[:\s]*(\d+)',
                r'living\s*area[:\s]*(\d+)',
                r'(\d+)\s*m[²2]',  # Generic fallback
            ]
            
            for pattern in area_patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                for match in matches:
                    try:
                        area = int(match)
                        # Valid living area range
                        if 20 < area < 2000:
                            data.living_area = float(area)
                            break
                    except:
                        pass
                if data.living_area > 0:
                    break
            
            # --- Extract Plot Area ---
            plot_patterns = [
                r'grond(?:opp(?:ervlakte)?)?[:\s]*(\d+)\s*m[²2]',
                r'perceel[:\s]*(\d+)\s*m[²2]',
                r'terrein[:\s]*(\d+)\s*m[²2]',
                r'tuin[:\s]*(\d+)\s*m[²2]',
            ]
            
            for pattern in plot_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    try:
                        plot = int(match.group(1))
                        if 10 < plot < 100000:
                            data.plot_area = float(plot)
                            break
                    except:
                        pass
            
            # --- Extract Bedrooms ---
            bedroom_patterns = [
                r'(\d+)\s*slaapkamer',
                r'slaapkamers?[:\s]*(\d+)',
                r'(\d+)\s*bedroom',
                r'bedrooms?[:\s]*(\d+)',
                r'(\d+)\s*slpk',
            ]
            
            for pattern in bedroom_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    try:
                        bedrooms = int(match.group(1))
                        if 0 < bedrooms < 20:
                            data.bedrooms = bedrooms
                            break
                    except:
                        pass
            
            # --- Extract Bathrooms ---
            bathroom_patterns = [
                r'(\d+)\s*badkamer',
                r'badkamers?[:\s]*(\d+)',
                r'(\d+)\s*bathroom',
                r'bathrooms?[:\s]*(\d+)',
            ]
            
            for pattern in bathroom_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    try:
                        bathrooms = int(match.group(1))
                        if 0 < bathrooms < 10:
                            data.bathrooms = bathrooms
                            break
                    except:
                        pass
            
            # --- Extract EPC ---
            epc_patterns = [
                r'epc[:\s-]*([A-Ga-g])\b',
                r'energielabel[:\s]*([A-Ga-g])\b',
                r'energie[:\s]*([A-Ga-g])\b',
            ]
            
            for pattern in epc_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    data.epc_score = match.group(1).upper()
                    break
            
            # --- Extract EPC Value ---
            epc_value_patterns = [
                r'(\d+)\s*kwh/m[²2]',
                r'epc[:\s]*(\d+)\s*kwh',
            ]
            
            for pattern in epc_value_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    try:
                        epc_val = int(match.group(1))
                        if 0 < epc_val < 1000:
                            data.epc_value = float(epc_val)
                            break
                    except:
                        pass
            
            # --- Extract Construction Year ---
            year_patterns = [
                r'bouwjaar[:\s]*(\d{4})',
                r'gebouwd\s*(?:in)?\s*(\d{4})',
                r'construction[:\s]*(\d{4})',
                r'jaar[:\s]*(\d{4})',
            ]
            
            for pattern in year_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    try:
                        year = int(match.group(1))
                        if 1800 < year < 2030:
                            data.construction_year = year
                            break
                    except:
                        pass
            
            # --- Extract Postal Code and City ---
            # Belgian postal code pattern
            postal_patterns = [
                r'\b(\d{4})\s+([A-Za-z-]+)\b',  # 9000 Gent
                r'([A-Za-z-]+)\s+(\d{4})\b',     # Gent 9000
            ]
            
            for pattern in postal_patterns:
                matches = re.findall(pattern, text)
                for match in matches:
                    if pattern == postal_patterns[0]:
                        postal, city = match
                    else:
                        city, postal = match
                    
                    try:
                        postal_int = int(postal)
                        if 1000 <= postal_int <= 9999:  # Belgian postal code range
                            if not data.postal_code:
                                data.postal_code = postal
                            if not data.city and len(city) > 2:
                                data.city = city.strip()
                            break
                    except:
                        pass
            
            # --- Extract Photos ---
            img_tags = soup.find_all('img', src=True)
            seen_urls = set()
            for img in img_tags:
                src = img.get('src', '') or img.get('data-src', '')
                if not src:
                    continue
                    
                # Make absolute URL
                if src.startswith('//'):
                    src = 'https:' + src
                elif src.startswith('/'):
                    from urllib.parse import urlparse
                    parsed = urlparse(url)
                    src = f"{parsed.scheme}://{parsed.netloc}{src}"
                
                # Filter for likely property photos
                src_lower = src.lower()
                if any(x in src_lower for x in ['photo', 'image', 'property', 'pand', 'woning', 'house', 'thumb', 'gallery']):
                    if src not in seen_urls and len(data.photos) < 10:
                        seen_urls.add(src)
                        data.photos.append(src)
            
            logger.info(f"Generic scrape results: price={data.asking_price}, area={data.living_area}, bedrooms={data.bedrooms}")
            
    except Exception as e:
        logger.error(f"Generic scrape error: {e}")
    
    return data

@api_router.post("/properties/scrape")
async def scrape_property_url(url: str = Query(...), current_user: User = Depends(get_current_user)):
    """Scrape property data from a real estate URL (Immoweb, Zimmo, Immoscoop, or any realtor website)"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    logger.info(f"Scraping property URL: {url}")
    
    # Determine which scraper to use
    url_lower = url.lower()
    
    try:
        if 'immoweb.be' in url_lower:
            data = await scrape_immoweb(url)
            platform_name = "Immoweb"
        elif 'zimmo.be' in url_lower:
            data = await scrape_zimmo(url)
            platform_name = "Zimmo"
        elif 'immoscoop.be' in url_lower:
            data = await scrape_immoscoop(url)
            platform_name = "Immoscoop"
        else:
            # Use generic scraper for any other website (realtor websites, etc.)
            data = await scrape_generic_website(url)
            platform_name = "website"
        
        # Log what we found
        logger.info(f"Scraped from {platform_name}: address={data.address}, price={data.asking_price}, area={data.living_area}")
        
        # Check if we got meaningful data
        has_data = bool(data.address or data.asking_price > 0 or data.living_area > 0 or data.bedrooms > 0)
        
        if has_data:
            # Count how many fields we found
            fields_found = sum([
                bool(data.address),
                data.asking_price > 0,
                data.living_area > 0,
                data.bedrooms > 0,
                data.bathrooms > 0,
                bool(data.postal_code),
                bool(data.city),
                bool(data.epc_score),
                data.construction_year is not None
            ])
            
            if fields_found >= 3:
                message = f"✅ {fields_found} gegevens opgehaald van {platform_name}! Controleer en vul aan."
            else:
                message = f"⚠️ Slechts {fields_found} gegevens gevonden. Vul de rest handmatig aan."
            
            return {
                "success": True,
                "data": data.model_dump(),
                "message": message,
                "fields_found": fields_found
            }
        else:
            # No data found - might be blocked or page structure changed
            if 'immoweb.be' in url_lower or 'zimmo.be' in url_lower:
                return {
                    "success": False,
                    "data": data.model_dump(),
                    "message": f"{platform_name} blokkeert automatisch ophalen (anti-bot beveiliging). Kopieer de gegevens handmatig.",
                    "hint": "💡 Tip: Probeer de makelaar's eigen website in plaats van Immoweb/Zimmo"
                }
            else:
                return {
                    "success": False,
                    "data": data.model_dump(),
                    "message": "Kon geen gegevens vinden op deze pagina. Mogelijk is de pagina leeg of heeft een andere structuur.",
                    "hint": "💡 Vul de gegevens handmatig in"
                }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Scraping error: {e}")
        return {
            "success": False,
            "data": ScrapedPropertyData().model_dump(),
            "message": f"Fout bij ophalen: {str(e)[:100]}"
        }

# --- Work Item Labels (voor renovatiecalculator) ---

@api_router.put("/work-items/{work_item_id}/label")
async def update_work_item_label(
    work_item_id: str, 
    component_label: str = Query(..., description="vloer, muur, plafond, elektriciteit, sanitair, verwarming, isolatie, overig"),
    room_types: str = Query("all", description="Komma-gescheiden lijst: all, bathroom, kitchen, bedroom, living, hallway"),
    current_user: User = Depends(get_current_user)
):
    """Update component label for a work item (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update work item labels")
    
    valid_labels = ["vloer", "muur", "plafond", "elektriciteit", "sanitair", "verwarming", "isolatie", "overig"]
    if component_label not in valid_labels:
        raise HTTPException(status_code=400, detail=f"Ongeldig label. Kies uit: {', '.join(valid_labels)}")
    
    room_types_list = [rt.strip() for rt in room_types.split(",")]
    
    result = await db.work_items.update_one(
        {"id": work_item_id},
        {"$set": {
            "component_label": component_label,
            "room_types": room_types_list
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Werkpost niet gevonden")
    
    return {"message": "Label bijgewerkt"}

# --- Renovation Calculator ---

# Helper function to find work item price by title pattern
def find_work_item_price(work_items, title_pattern, default_price, default_unit="m²"):
    """Find a work item by title pattern (case-insensitive partial match)"""
    pattern_lower = title_pattern.lower()
    for item in work_items:
        if pattern_lower in item.get("title", "").lower():
            return {
                "id": item.get("id"),
                "title": item.get("title"),
                "price": item.get("price", default_price),
                "unit": item.get("unit", default_unit)
            }
    return {"id": None, "title": title_pattern, "price": default_price, "unit": default_unit}

@api_router.post("/properties/{property_id}/calculate")
async def calculate_renovation(property_id: str, current_user: User = Depends(get_current_user)):
    """Generate smart renovation calculation using real work item prices from database"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    # Get property
    tenant_filter = get_tenant_filter(current_user, "property")
    filter_query = {"id": property_id}
    if tenant_filter:
        filter_query = {"$and": [{"id": property_id}, tenant_filter]}
    
    prop = await db.properties.find_one(filter_query, {"_id": 0})
    if not prop:
        raise HTTPException(status_code=404, detail="Pand niet gevonden")
    
    rooms = prop.get("rooms", [])
    if not rooms:
        raise HTTPException(status_code=400, detail="Voeg eerst kamers toe aan het pand")
    
    # Use shared calculation logic
    calculation, calc_doc, room_calculations, all_work_items = await _perform_renovation_calculation(rooms)
    
    # Remove old calculation if exists
    await db.renovation_calculations.delete_many({"property_id": property_id})
    
    calc_doc["property_id"] = property_id
    await db.renovation_calculations.insert_one(calc_doc)
    
    await db.properties.update_one(
        {"id": property_id},
        {"$set": {
            "status": "calculated",
            "renovation_calculation_id": calculation.id,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    logger.info(f"Calculated renovation for property {property_id}: €{calculation.total_min:.2f} (using {len(all_work_items)} work items)")
    
    return {
        "calculation_id": calculation.id,
        "total_min": calculation.total_min,
        "total_realistic": calculation.total_realistic,
        "total_max": calculation.total_max,
        "estimated_duration_weeks": calculation.estimated_duration_weeks,
        "rooms_calculated": len(room_calculations),
        "work_items_used": len(all_work_items)
    }

async def _perform_renovation_calculation(rooms):
    """Shared renovation calculation logic for both properties and projects"""
    # ============= FETCH ALL WORK ITEMS FROM DATABASE =============
    all_work_items = await db.work_items.find({}, {"_id": 0}).to_list(500)
    
    # Group work items by component_label
    work_items_by_label = {
        "vloer": [],
        "muur": [],
        "plafond": [],
        "elektriciteit": [],
        "sanitair": [],
        "verwarming": [],
        "isolatie": [],
        "overig": []
    }
    
    for item in all_work_items:
        label = item.get("component_label")
        if not label:
            continue  # Skip items without a label - they are not categorized yet
        if label in work_items_by_label:
            work_items_by_label[label].append(item)
        else:
            work_items_by_label["overig"].append(item)
    
    logger.info(f"Loaded work items by label: vloer={len(work_items_by_label['vloer'])}, muur={len(work_items_by_label['muur'])}, plafond={len(work_items_by_label['plafond'])}, elektriciteit={len(work_items_by_label['elektriciteit'])}, sanitair={len(work_items_by_label['sanitair'])}, overig={len(work_items_by_label['overig'])}")
    
    # ============= DEFAULT PRICES (fallback if not in database) =============
    DEFAULT_PRICES = {
        "afbraak_vloer": 27,
        "egaliseren": 35,
        "tegels_standaard": 95,
        "tegels_badkamer": 115,
        "parket": 65,
        "laminaat": 35,
        "vinyl": 68,
        "pleisterij": 60,
        "gyproc_wit": 65,
        "gyproc_groen": 75,
        "afbraak_wand": 25,
        "schilderen": 32,
        "plafond_gyproc": 150,
        "afbraak_plafond": 25,
        "spot_wit": 80,
        "spot_zwart": 92,
        "schakelaar_wit": 50,
        "stopcontact_wit": 50,
    }
    
    # STANDAARD HOOGTE
    STANDAARD_HOOGTE = 2.55
    
    # Calculate per room
    room_calculations = []
    total = 0.0
    
    for room in rooms:
        room_type = room.get("room_type", "other")
        
        # Calculate areas
        floor_area = room.get("floor_area", 0) or (room.get("length", 0) * room.get("width", 0))
        ceiling_area = room.get("ceiling_area", 0) or floor_area
        
        room_height = room.get("height", 0)
        if not room_height or room_height <= 0:
            room_height = STANDAARD_HOOGTE
            height_source = "standaard"
        else:
            height_source = "opgegeven"
        
        wall_area = room.get("wall_area", 0)
        if not wall_area and room.get("length") and room.get("width"):
            perimeter = 2 * (room.get("length", 0) + room.get("width", 0))
            wall_area = perimeter * room_height
        
        is_bathroom = room_type == "bathroom"
        is_kitchen = room_type == "kitchen"
        
        room_calc = RoomCalculation(
            room_id=room.get("id", str(uuid.uuid4())),
            room_name=room.get("name", "Onbekend"),
            room_type=room_type,
            floor_area=round(floor_area, 2),
            wall_area=round(wall_area, 2),
            ceiling_area=round(ceiling_area, 2),
            room_height=round(room_height, 2),
            height_source=height_source,
            floor_items=[],
            wall_items=[],
            ceiling_items=[],
            other_items=[],
            subtotal=0.0,
            selected_floor_option="tegels",
            selected_wall_scenario="nieuw_pleisterwerk"
        )
        
        room_subtotal = 0.0
        
        # ==================== VLOER ====================
        # Get prices from database
        vloer_items = work_items_by_label.get("vloer", [])
        
        # 1. Afbraak bestaande vloer
        afbraak_info = find_work_item_price(all_work_items, "afbraak vloer", DEFAULT_PRICES["afbraak_vloer"])
        afbraak_total = floor_area * afbraak_info["price"]
        room_calc.floor_items.append(CalculationItem(
            work_item_id=afbraak_info["id"],
            title=afbraak_info["title"],
            quantity=round(floor_area, 2),
            unit="m²",
            unit_price=afbraak_info["price"],
            total=round(afbraak_total, 2),
            included=True,
            category="vloer_voorbereiding",
            is_selected=True,
            option_group="vloer_basis"
        ))
        room_subtotal += afbraak_total
        
        # 2. Egaliseren
        egaliseer_info = find_work_item_price(all_work_items, "egaliseren", DEFAULT_PRICES["egaliseren"])
        egaliseer_total = floor_area * egaliseer_info["price"]
        room_calc.floor_items.append(CalculationItem(
            work_item_id=egaliseer_info["id"],
            title=egaliseer_info["title"],
            quantity=round(floor_area, 2),
            unit="m²",
            unit_price=egaliseer_info["price"],
            total=round(egaliseer_total, 2),
            included=True,
            category="vloer_voorbereiding",
            is_selected=True,
            option_group="vloer_basis"
        ))
        room_subtotal += egaliseer_total
        
        # 3. VLOER AFWERKING OPTIES
        # Optie A: Tegels (standaard geselecteerd)
        if is_bathroom:
            tegel_info = find_work_item_price(all_work_items, "tegelen badkamer", DEFAULT_PRICES["tegels_badkamer"])
        else:
            tegel_info = find_work_item_price(all_work_items, "tegelen: standaard", DEFAULT_PRICES["tegels_standaard"])
        tegel_total = floor_area * tegel_info["price"]
        room_calc.floor_items.append(CalculationItem(
            work_item_id=tegel_info["id"],
            title=tegel_info["title"],
            quantity=round(floor_area, 2),
            unit="m²",
            unit_price=tegel_info["price"],
            total=round(tegel_total, 2),
            included=True,
            category="vloer_afwerking",
            is_selected=True,
            option_group="vloer_afwerking_keuze"
        ))
        room_subtotal += tegel_total
        
        # Optie B: Parket
        parket_info = find_work_item_price(vloer_items, "parket", DEFAULT_PRICES["parket"])
        parket_total = floor_area * parket_info["price"]
        room_calc.floor_items.append(CalculationItem(
            work_item_id=parket_info["id"],
            title=parket_info["title"],
            quantity=round(floor_area, 2),
            unit="m²",
            unit_price=parket_info["price"],
            total=round(parket_total, 2),
            included=False,
            category="vloer_afwerking",
            is_selected=False,
            option_group="vloer_afwerking_keuze"
        ))
        
        # Optie C: Laminaat
        laminaat_info = find_work_item_price(vloer_items, "laminaat", DEFAULT_PRICES["laminaat"])
        laminaat_total = floor_area * laminaat_info["price"]
        room_calc.floor_items.append(CalculationItem(
            work_item_id=laminaat_info["id"],
            title=laminaat_info["title"],
            quantity=round(floor_area, 2),
            unit="m²",
            unit_price=laminaat_info["price"],
            total=round(laminaat_total, 2),
            included=False,
            category="vloer_afwerking",
            is_selected=False,
            option_group="vloer_afwerking_keuze"
        ))
        
        # Optie D: Vinyl
        vinyl_info = find_work_item_price(vloer_items, "vinyl", DEFAULT_PRICES["vinyl"])
        vinyl_total = floor_area * vinyl_info["price"]
        room_calc.floor_items.append(CalculationItem(
            work_item_id=vinyl_info["id"],
            title=vinyl_info["title"],
            quantity=round(floor_area, 2),
            unit="m²",
            unit_price=vinyl_info["price"],
            total=round(vinyl_total, 2),
            included=False,
            category="vloer_afwerking",
            is_selected=False,
            option_group="vloer_afwerking_keuze"
        ))
        
        # VLOER EXTRAS - remaining vloer items not already used as options
        used_vloer_titles = {"parket", "laminaat", "vinyl"}  # Already added as options
        for item in vloer_items:
            title_lower = item.get("title", "").lower()
            if any(x in title_lower for x in used_vloer_titles):
                continue
            room_types = item.get("room_types", ["all"])
            if "all" in room_types or room_type in room_types:
                if item.get("unit") == "m²":
                    item_total = floor_area * item.get("price", 0)
                    quantity = floor_area
                else:
                    item_total = item.get("price", 0)
                    quantity = 1
                room_calc.floor_items.append(CalculationItem(
                    work_item_id=item.get("id"),
                    title=f"Extra: {item.get('title', '?')}",
                    quantity=round(quantity, 2),
                    unit=item.get("unit", "stuk"),
                    unit_price=item.get("price", 0),
                    total=round(item_total, 2),
                    included=False,
                    category="vloer_extra",
                    is_selected=False,
                    option_group="vloer_extras"
                ))
        
        # ==================== MUREN ====================
        muur_items = work_items_by_label.get("muur", [])
        
        # SCENARIO A: Nieuw pleisterwerk
        afbraak_wand_info = find_work_item_price(all_work_items, "afbraak wandtegel", DEFAULT_PRICES["afbraak_wand"])
        afbraak_wand_total = wall_area * afbraak_wand_info["price"]
        room_calc.wall_items.append(CalculationItem(
            work_item_id=afbraak_wand_info["id"],
            title=afbraak_wand_info["title"],
            quantity=round(wall_area, 2),
            unit="m²",
            unit_price=afbraak_wand_info["price"],
            total=round(afbraak_wand_total, 2),
            included=True,
            category="muur_scenario_a",
            is_selected=True,
            option_group="muur_ondergrond"
        ))
        room_subtotal += afbraak_wand_total
        
        pleisterij_info = find_work_item_price(muur_items, "pleisterij", DEFAULT_PRICES["pleisterij"])
        pleisterij_total = wall_area * pleisterij_info["price"]
        room_calc.wall_items.append(CalculationItem(
            work_item_id=pleisterij_info["id"],
            title=pleisterij_info["title"],
            quantity=round(wall_area, 2),
            unit="m²",
            unit_price=pleisterij_info["price"],
            total=round(pleisterij_total, 2),
            included=True,
            category="muur_scenario_a",
            is_selected=True,
            option_group="muur_ondergrond"
        ))
        room_subtotal += pleisterij_total
        
        # SCENARIO B: Gyproc wit (egaliseren alternatief)
        gyproc_wit_info = find_work_item_price(muur_items, "gyproc wit", DEFAULT_PRICES["gyproc_wit"])
        gyproc_wit_total = wall_area * gyproc_wit_info["price"]
        room_calc.wall_items.append(CalculationItem(
            work_item_id=gyproc_wit_info["id"],
            title=gyproc_wit_info["title"],
            quantity=round(wall_area, 2),
            unit="m²",
            unit_price=gyproc_wit_info["price"],
            total=round(gyproc_wit_total, 2),
            included=False,
            category="muur_scenario_b",
            is_selected=False,
            option_group="muur_ondergrond"
        ))
        
        # SCENARIO C: Gyproc groen (badkamer)
        gyproc_groen_info = find_work_item_price(muur_items, "gyproc groen", DEFAULT_PRICES["gyproc_groen"])
        gyproc_groen_total = wall_area * gyproc_groen_info["price"]
        room_calc.wall_items.append(CalculationItem(
            work_item_id=gyproc_groen_info["id"],
            title=gyproc_groen_info["title"],
            quantity=round(wall_area, 2),
            unit="m²",
            unit_price=gyproc_groen_info["price"],
            total=round(gyproc_groen_total, 2),
            included=False,
            category="muur_scenario_c",
            is_selected=False,
            option_group="muur_ondergrond"
        ))
        
        # Schilderwerk muren
        schilder_info = find_work_item_price(all_work_items, "schilderen", DEFAULT_PRICES["schilderen"])
        schilder_total = wall_area * schilder_info["price"]
        room_calc.wall_items.append(CalculationItem(
            work_item_id=schilder_info["id"],
            title=f"⬜ {schilder_info['title']} (optioneel)",
            quantity=round(wall_area, 2),
            unit="m²",
            unit_price=schilder_info["price"],
            total=round(schilder_total, 2),
            included=True,
            category="muur_afwerking",
            is_selected=True,
            option_group="schilderwerk"
        ))
        room_subtotal += schilder_total
        
        # MUUR EXTRAS - alle muur items die niet standaard zijn gebruikt
        for item in muur_items:
            title_lower = item.get("title", "").lower()
            # Skip items we already added
            if any(x in title_lower for x in ["pleisterij", "voorzetwand gyproc wit", "scheidingswand gyproc groen"]):
                continue
            room_types = item.get("room_types", ["all"])
            if "all" in room_types or room_type in room_types:
                if item.get("unit") == "m²":
                    item_total = wall_area * item.get("price", 0)
                    quantity = wall_area
                elif item.get("unit") == "lm":
                    perimeter = 2 * (room.get("length", 0) + room.get("width", 0))
                    item_total = perimeter * item.get("price", 0)
                    quantity = perimeter
                else:
                    item_total = item.get("price", 0)
                    quantity = 1
                room_calc.wall_items.append(CalculationItem(
                    work_item_id=item.get("id"),
                    title=f"Extra: {item.get('title', '?')}",
                    quantity=round(quantity, 2),
                    unit=item.get("unit", "stuk"),
                    unit_price=item.get("price", 0),
                    total=round(item_total, 2),
                    included=False,
                    category="muur_extra",
                    is_selected=False,
                    option_group="muur_extras"
                ))
        
        # ==================== PLAFOND ====================
        plafond_items = work_items_by_label.get("plafond", [])
        
        # Afbraak plafond
        afbraak_plafond_info = find_work_item_price(all_work_items, "afbraak plafond", DEFAULT_PRICES["afbraak_plafond"])
        afbraak_plafond_total = ceiling_area * afbraak_plafond_info["price"]
        room_calc.ceiling_items.append(CalculationItem(
            work_item_id=afbraak_plafond_info["id"],
            title=afbraak_plafond_info["title"],
            quantity=round(ceiling_area, 2),
            unit="m²",
            unit_price=afbraak_plafond_info["price"],
            total=round(afbraak_plafond_total, 2),
            included=True,
            category="plafond_basis",
            is_selected=True,
            option_group="plafond_constructie"
        ))
        room_subtotal += afbraak_plafond_total
        
        # Nieuw plafond gyproc
        plafond_gyproc_info = find_work_item_price(plafond_items, "gyproc", DEFAULT_PRICES["plafond_gyproc"])
        plafond_gyproc_total = ceiling_area * plafond_gyproc_info["price"]
        room_calc.ceiling_items.append(CalculationItem(
            work_item_id=plafond_gyproc_info["id"],
            title=plafond_gyproc_info["title"],
            quantity=round(ceiling_area, 2),
            unit="m²",
            unit_price=plafond_gyproc_info["price"],
            total=round(plafond_gyproc_total, 2),
            included=True,
            category="plafond_basis",
            is_selected=True,
            option_group="plafond_constructie"
        ))
        room_subtotal += plafond_gyproc_total
        
        # Schilderwerk plafond
        schilder_plafond_total = ceiling_area * schilder_info["price"]
        room_calc.ceiling_items.append(CalculationItem(
            work_item_id=schilder_info["id"],
            title=f"⬜ {schilder_info['title']} plafond (optioneel)",
            quantity=round(ceiling_area, 2),
            unit="m²",
            unit_price=schilder_info["price"],
            total=round(schilder_plafond_total, 2),
            included=True,
            category="plafond_afwerking",
            is_selected=True,
            option_group="schilderwerk"
        ))
        room_subtotal += schilder_plafond_total
        
        # PLAFOND EXTRAS
        for item in plafond_items:
            title_lower = item.get("title", "").lower()
            if "gyproc" in title_lower:
                continue  # Already added
            room_types = item.get("room_types", ["all"])
            if "all" in room_types or room_type in room_types:
                if item.get("unit") == "m²":
                    item_total = ceiling_area * item.get("price", 0)
                    quantity = ceiling_area
                else:
                    item_total = item.get("price", 0)
                    quantity = 1
                room_calc.ceiling_items.append(CalculationItem(
                    work_item_id=item.get("id"),
                    title=f"Extra: {item.get('title', '?')}",
                    quantity=round(quantity, 2),
                    unit=item.get("unit", "stuk"),
                    unit_price=item.get("price", 0),
                    total=round(item_total, 2),
                    included=False,
                    category="plafond_extra",
                    is_selected=False,
                    option_group="plafond_extras"
                ))
        
        # ==================== ELEKTRICITEIT ====================
        elektriciteit_items = work_items_by_label.get("elektriciteit", [])
        
        # Spots (1 per 5m²)
        aantal_spots = max(1, int(floor_area / 5))
        spot_info = find_work_item_price(elektriciteit_items, "spot wit", DEFAULT_PRICES["spot_wit"])
        spot_total = aantal_spots * spot_info["price"]
        room_calc.other_items.append(CalculationItem(
            work_item_id=spot_info["id"],
            title=f"{spot_info['title']} (1 per 5m²)",
            quantity=aantal_spots,
            unit="stuk",
            unit_price=spot_info["price"],
            total=round(spot_total, 2),
            included=True,
            category="elektriciteit",
            is_selected=True,
            option_group="verlichting"
        ))
        room_subtotal += spot_total
        
        # Schakelaars (1 per 10m²)
        aantal_schakelaars = max(1, int(floor_area / 10))
        schakelaar_info = find_work_item_price(elektriciteit_items, "schakelaar wit", DEFAULT_PRICES["schakelaar_wit"])
        schakelaar_total = aantal_schakelaars * schakelaar_info["price"]
        room_calc.other_items.append(CalculationItem(
            work_item_id=schakelaar_info["id"],
            title=f"{schakelaar_info['title']} (1 per 10m²)",
            quantity=aantal_schakelaars,
            unit="stuk",
            unit_price=schakelaar_info["price"],
            total=round(schakelaar_total, 2),
            included=True,
            category="elektriciteit",
            is_selected=True,
            option_group="schakelaars"
        ))
        room_subtotal += schakelaar_total
        
        # Stopcontacten (1 per 10m²)
        aantal_stopcontacten = max(1, int(floor_area / 10))
        stopcontact_info = find_work_item_price(elektriciteit_items, "stopcontact wit", DEFAULT_PRICES["stopcontact_wit"])
        stopcontact_total = aantal_stopcontacten * stopcontact_info["price"]
        room_calc.other_items.append(CalculationItem(
            work_item_id=stopcontact_info["id"],
            title=f"{stopcontact_info['title']} (1 per 10m²)",
            quantity=aantal_stopcontacten,
            unit="stuk",
            unit_price=stopcontact_info["price"],
            total=round(stopcontact_total, 2),
            included=True,
            category="elektriciteit",
            is_selected=True,
            option_group="stopcontacten"
        ))
        room_subtotal += stopcontact_total
        
        # ELEKTRICITEIT EXTRAS
        for item in elektriciteit_items:
            title_lower = item.get("title", "").lower()
            if any(x in title_lower for x in ["spot wit", "schakelaar wit", "stopcontact wit"]):
                continue
            room_calc.other_items.append(CalculationItem(
                work_item_id=item.get("id"),
                title=f"Extra: {item.get('title', '?')}",
                quantity=1,
                unit=item.get("unit", "stuk"),
                unit_price=item.get("price", 0),
                total=round(item.get("price", 0), 2),
                included=False,
                category="elektriciteit_extra",
                is_selected=False,
                option_group="elektriciteit_extras"
            ))
        
        # ==================== SANITAIR (alleen voor badkamer/keuken) ====================
        if is_bathroom or is_kitchen:
            sanitair_items = work_items_by_label.get("sanitair", [])
            for item in sanitair_items:
                room_types = item.get("room_types", ["all"])
                if "all" in room_types or room_type in room_types:
                    room_calc.other_items.append(CalculationItem(
                        work_item_id=item.get("id"),
                        title=f"Extra: {item.get('title', '?')}",
                        quantity=1,
                        unit=item.get("unit", "stuk"),
                        unit_price=item.get("price", 0),
                        total=round(item.get("price", 0), 2),
                        included=False,
                        category="sanitair_extra",
                        is_selected=False,
                        option_group="sanitair_extras"
                    ))
        
        # ==================== OVERIG / ALGEMEEN ====================
        overig_items = work_items_by_label.get("overig", [])
        for item in overig_items:
            room_types = item.get("room_types", ["all"])
            if "all" in room_types or room_type in room_types:
                if item.get("unit") == "m²":
                    item_total = floor_area * item.get("price", 0)
                    quantity = floor_area
                else:
                    item_total = item.get("price", 0)
                    quantity = 1
                room_calc.other_items.append(CalculationItem(
                    work_item_id=item.get("id"),
                    title=f"Extra: {item.get('title', '?')}",
                    quantity=round(quantity, 2),
                    unit=item.get("unit", "stuk"),
                    unit_price=item.get("price", 0),
                    total=round(item_total, 2),
                    included=False,
                    category="overig_extra",
                    is_selected=False,
                    option_group="overig_extras"
                ))
        
        room_calc.subtotal = round(room_subtotal, 2)
        total += room_subtotal
        room_calculations.append(room_calc)
    
    # Create calculation object
    calculation = RenovationCalculation(
        calculated_by="system",
        room_calculations=room_calculations,
        total_min=round(total * 0.85, 2),
        total_realistic=round(total, 2),
        total_max=round(total * 1.20, 2),
        estimated_duration_weeks=max(2, int(total / 10000)),
        estimated_epc_improvement=""
    )
    
    calc_doc = calculation.model_dump()
    calc_doc["created_at"] = calc_doc["created_at"].isoformat()
    calc_doc["updated_at"] = calc_doc["updated_at"].isoformat()
    calc_doc["room_calculations"] = [
        {
            **rc.model_dump(),
            "floor_items": [item.model_dump() for item in rc.floor_items],
            "wall_items": [item.model_dump() for item in rc.wall_items],
            "ceiling_items": [item.model_dump() for item in rc.ceiling_items],
            "other_items": [item.model_dump() for item in rc.other_items]
        }
        for rc in room_calculations
    ]
    
    logger.info(f"Renovation calculation done: €{total:.2f} (using {len(all_work_items)} work items)")
    
    return calculation, calc_doc, room_calculations, all_work_items

@api_router.get("/properties/{property_id}/calculation")
async def get_renovation_calculation(property_id: str, current_user: User = Depends(get_current_user)):
    """Get renovation calculation for a property"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    # Check property access
    tenant_filter = get_tenant_filter(current_user, "property")
    filter_query = {"id": property_id}
    if tenant_filter:
        filter_query = {"$and": [{"id": property_id}, tenant_filter]}
    
    prop = await db.properties.find_one(filter_query)
    if not prop:
        raise HTTPException(status_code=404, detail="Pand niet gevonden")
    
    calc = await db.renovation_calculations.find_one({"property_id": property_id}, {"_id": 0})
    if not calc:
        raise HTTPException(status_code=404, detail="Geen berekening gevonden. Start eerst een berekening.")
    
    return calc

@api_router.put("/properties/{property_id}/calculation/items/{item_id}")
async def toggle_calculation_item(
    property_id: str, 
    item_id: str, 
    included: bool = Query(...),
    current_user: User = Depends(get_current_user)
):
    """Toggle a calculation item on/off and recalculate totals"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    # Get calculation
    calc = await db.renovation_calculations.find_one({"property_id": property_id})
    if not calc:
        raise HTTPException(status_code=404, detail="Berekening niet gevonden")
    
    # Find and update item
    room_calculations = calc.get("room_calculations", [])
    item_found = False
    new_total = 0.0
    
    for room_calc in room_calculations:
        room_subtotal = 0.0
        for item_list_name in ["floor_items", "wall_items", "ceiling_items", "other_items"]:
            for item in room_calc.get(item_list_name, []):
                if item.get("id") == item_id:
                    item["included"] = included
                    item_found = True
                
                if item.get("included", True):
                    room_subtotal += item.get("total", 0)
        
        room_calc["subtotal"] = round(room_subtotal, 2)
        new_total += room_subtotal
    
    if not item_found:
        raise HTTPException(status_code=404, detail="Item niet gevonden")
    
    # Update calculation
    await db.renovation_calculations.update_one(
        {"property_id": property_id},
        {"$set": {
            "room_calculations": room_calculations,
            "total_min": round(new_total * 0.85, 2),
            "total_realistic": round(new_total, 2),
            "total_max": round(new_total * 1.20, 2),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "message": "Item bijgewerkt",
        "total_realistic": round(new_total, 2)
    }

@api_router.put("/properties/{property_id}/calculation/switch-option")
async def switch_calculation_option(
    property_id: str,
    room_id: str = Query(..., description="ID van de kamer"),
    option_group: str = Query(..., description="Naam van de optie groep (bijv. vloer_afwerking_keuze)"),
    selected_item_id: str = Query(..., description="ID van het item dat geselecteerd moet worden"),
    current_user: User = Depends(get_current_user)
):
    """Switch between options within the same option_group (e.g., switch from tiles to parquet)"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    # Get calculation
    calc = await db.renovation_calculations.find_one({"property_id": property_id})
    if not calc:
        raise HTTPException(status_code=404, detail="Berekening niet gevonden")
    
    room_calculations = calc.get("room_calculations", [])
    option_switched = False
    new_total = 0.0
    
    for room_calc in room_calculations:
        if room_calc.get("room_id") != room_id:
            # Recalculate subtotal for this room
            room_subtotal = 0.0
            for item_list_name in ["floor_items", "wall_items", "ceiling_items", "other_items"]:
                for item in room_calc.get(item_list_name, []):
                    if item.get("included", True):
                        room_subtotal += item.get("total", 0)
            room_calc["subtotal"] = round(room_subtotal, 2)
            new_total += room_subtotal
            continue
        
        # This is the target room - switch options
        room_subtotal = 0.0
        
        for item_list_name in ["floor_items", "wall_items", "ceiling_items", "other_items"]:
            for item in room_calc.get(item_list_name, []):
                # Check if this item is in the target option_group
                if item.get("option_group") == option_group:
                    if item.get("id") == selected_item_id:
                        # This is the newly selected item
                        item["included"] = True
                        item["is_selected"] = True
                        option_switched = True
                    else:
                        # Deselect all other items in this group
                        item["included"] = False
                        item["is_selected"] = False
                
                # Add to subtotal if included
                if item.get("included", True):
                    room_subtotal += item.get("total", 0)
        
        room_calc["subtotal"] = round(room_subtotal, 2)
        new_total += room_subtotal
    
    if not option_switched:
        raise HTTPException(status_code=404, detail="Optie niet gevonden in deze optie groep")
    
    # Update calculation
    await db.renovation_calculations.update_one(
        {"property_id": property_id},
        {"$set": {
            "room_calculations": room_calculations,
            "total_min": round(new_total * 0.85, 2),
            "total_realistic": round(new_total, 2),
            "total_max": round(new_total * 1.20, 2),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "message": "Optie gewisseld",
        "total_realistic": round(new_total, 2)
    }

@api_router.put("/properties/{property_id}/calculation/switch-scenario")
async def switch_wall_scenario(
    property_id: str,
    room_id: str = Query(..., description="ID van de kamer"),
    scenario: str = Query(..., description="Scenario: 'nieuw_pleisterwerk', 'egaliseren', of 'gyproc'"),
    current_user: User = Depends(get_current_user)
):
    """Switch wall scenario for a room (nieuw pleisterwerk, egaliseren, or gyproc)"""
    if current_user.role not in ["admin", "realtor", "investor"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    valid_scenarios = {
        "nieuw_pleisterwerk": ["muur_scenario_a"],
        "egaliseren": ["muur_scenario_b"],
        "gyproc": ["muur_scenario_c"]
    }
    
    if scenario not in valid_scenarios:
        raise HTTPException(status_code=400, detail=f"Ongeldig scenario. Kies uit: {list(valid_scenarios.keys())}")
    
    # Get calculation
    calc = await db.renovation_calculations.find_one({"property_id": property_id})
    if not calc:
        raise HTTPException(status_code=404, detail="Berekening niet gevonden")
    
    room_calculations = calc.get("room_calculations", [])
    scenario_switched = False
    new_total = 0.0
    
    selected_categories = valid_scenarios[scenario]
    all_scenario_categories = ["muur_scenario_a", "muur_scenario_b", "muur_scenario_c"]
    
    for room_calc in room_calculations:
        room_subtotal = 0.0
        
        if room_calc.get("room_id") == room_id:
            # Switch scenario for this room
            room_calc["selected_wall_scenario"] = scenario
            
            for item in room_calc.get("wall_items", []):
                item_category = item.get("category", "")
                
                if item_category in all_scenario_categories:
                    if item_category in selected_categories:
                        item["included"] = True
                        item["is_selected"] = True
                        scenario_switched = True
                    else:
                        item["included"] = False
                        item["is_selected"] = False
                
                if item.get("included", True):
                    room_subtotal += item.get("total", 0)
            
            # Also add floor, ceiling, other items
            for item_list_name in ["floor_items", "ceiling_items", "other_items"]:
                for item in room_calc.get(item_list_name, []):
                    if item.get("included", True):
                        room_subtotal += item.get("total", 0)
        else:
            # Just recalculate subtotal for other rooms
            for item_list_name in ["floor_items", "wall_items", "ceiling_items", "other_items"]:
                for item in room_calc.get(item_list_name, []):
                    if item.get("included", True):
                        room_subtotal += item.get("total", 0)
        
        room_calc["subtotal"] = round(room_subtotal, 2)
        new_total += room_subtotal
    
    if not scenario_switched:
        raise HTTPException(status_code=404, detail="Kamer of scenario niet gevonden")
    
    # Update calculation
    await db.renovation_calculations.update_one(
        {"property_id": property_id},
        {"$set": {
            "room_calculations": room_calculations,
            "total_min": round(new_total * 0.85, 2),
            "total_realistic": round(new_total, 2),
            "total_max": round(new_total * 1.20, 2),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "message": f"Muur scenario gewisseld naar: {scenario}",
        "total_realistic": round(new_total, 2)
    }

# --- Realtor/Investor Login ---

class TenantLoginRequest(BaseModel):
    username: str
    password: str

@api_router.post("/auth/tenant/login")
async def tenant_login(login_data: TenantLoginRequest, response: Response):
    """Login for realtors and investors"""
    logger.info(f"Tenant login attempt for: {login_data.username}")
    
    # Find user
    user = await db.users.find_one({"username": login_data.username, "role": {"$in": ["realtor", "investor", "subcontractor"]}})
    
    if not user:
        logger.warning(f"Tenant login failed: user not found for {login_data.username}")
        raise HTTPException(status_code=401, detail="Ongeldige inloggegevens")
    
    if not user.get("password_hash"):
        logger.warning(f"Tenant login failed: no password_hash for {login_data.username}")
        raise HTTPException(status_code=401, detail="Ongeldige inloggegevens")
    
    if not verify_password(login_data.password, user["password_hash"]):
        logger.warning(f"Tenant login failed: wrong password for {login_data.username}")
        raise HTTPException(status_code=401, detail="Ongeldige inloggegevens")
    
    # Create session token
    session_token = secrets.token_urlsafe(32)
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    session_doc = {
        "user_id": user["_id"],
        "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.sessions.insert_one(session_doc)
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=7*24*60*60,
        path="/"
    )
    
    user_response = {
        "id": str(user["_id"]),
        "username": user.get("username"),
        "email": user.get("email"),
        "name": user.get("name"),
        "role": user.get("role"),
        "created_at": user.get("created_at")
    }
    
    logger.info(f"Tenant login successful for: {login_data.username}")
    return {"success": True, "user": user_response, "token": session_token, "role": user.get("role")}

# ============= MATERIAL REQUEST ENDPOINTS (Werkman Materiaal Aanvragen) =============

@api_router.post("/material-requests")
async def create_material_request(request: MaterialRequestCreate, current_user: User = Depends(get_current_user)):
    """Create a new material request (workers only)"""
    if current_user.role not in ["worker", "admin"]:
        raise HTTPException(status_code=403, detail="Alleen werkmannen kunnen materialen aanvragen")
    
    # Create request
    mat_request = MaterialRequest(
        title=request.title,
        quantity=request.quantity,
        needed_by=request.needed_by,
        photo_url=request.photo_url,
        notes=request.notes,
        project_id=request.project_id,
        project_name=request.project_name,
        requested_by=current_user.id,
        requested_by_name=current_user.name or current_user.username,
        status="pending",
        is_ordered=False,
        is_delivered=False
    )
    
    # Save to database
    doc = mat_request.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    if doc.get("ordered_at"):
        doc["ordered_at"] = doc["ordered_at"].isoformat()
    if doc.get("delivered_at"):
        doc["delivered_at"] = doc["delivered_at"].isoformat()
    
    await db.material_requests.insert_one(doc)
    
    logger.info(f"Material request created by {current_user.name}: {request.title}")
    
    return {
        "message": "Materiaal aanvraag verstuurd / Запит на матеріал надіслано",
        "request_id": mat_request.id
    }

@api_router.get("/material-requests")
async def get_material_requests(current_user: User = Depends(get_current_user)):
    """Get all material requests (admins see all, workers see their own)"""
    if current_user.role == "admin":
        requests = await db.material_requests.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    else:
        requests = await db.material_requests.find(
            {"requested_by": current_user.id}, 
            {"_id": 0}
        ).sort("created_at", -1).to_list(100)
    
    return requests

@api_router.get("/material-requests/pending")
async def get_pending_material_requests(current_user: User = Depends(get_current_user)):
    """Get pending material requests for admin notification banner"""
    if current_user.role != "admin":
        return []
    
    # Get requests that are not fully delivered yet
    requests = await db.material_requests.find(
        {"is_delivered": False},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return requests

@api_router.put("/material-requests/{request_id}/status")
async def update_material_request_status(
    request_id: str,
    is_ordered: Optional[bool] = None,
    is_delivered: Optional[bool] = None,
    current_user: User = Depends(get_current_user)
):
    """Update material request status (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen beheerders kunnen status bijwerken")
    
    # Find request
    request = await db.material_requests.find_one({"id": request_id})
    if not request:
        raise HTTPException(status_code=404, detail="Aanvraag niet gevonden")
    
    update_fields = {}
    
    if is_ordered is not None:
        update_fields["is_ordered"] = is_ordered
        if is_ordered:
            update_fields["ordered_at"] = datetime.now(timezone.utc).isoformat()
            update_fields["ordered_by"] = current_user.id
        else:
            update_fields["ordered_at"] = None
            update_fields["ordered_by"] = None
    
    if is_delivered is not None:
        update_fields["is_delivered"] = is_delivered
        if is_delivered:
            update_fields["delivered_at"] = datetime.now(timezone.utc).isoformat()
            update_fields["delivered_by"] = current_user.id
            update_fields["status"] = "delivered"
        else:
            update_fields["delivered_at"] = None
            update_fields["delivered_by"] = None
            update_fields["status"] = "ordered" if request.get("is_ordered") else "pending"
    
    # Determine status
    if update_fields.get("is_delivered"):
        update_fields["status"] = "delivered"
    elif update_fields.get("is_ordered") or request.get("is_ordered"):
        update_fields["status"] = "ordered"
    else:
        update_fields["status"] = "pending"
    
    await db.material_requests.update_one(
        {"id": request_id},
        {"$set": update_fields}
    )
    
    return {"message": "Status bijgewerkt"}

@api_router.delete("/material-requests/{request_id}")
async def delete_material_request(request_id: str, current_user: User = Depends(get_current_user)):
    """Delete a material request (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen beheerders kunnen aanvragen verwijderen")
    
    result = await db.material_requests.delete_one({"id": request_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Aanvraag niet gevonden")
    
    return {"message": "Aanvraag verwijderd"}

# ============= MATERIAL CATALOG ENDPOINTS (Beheerder Materialenlijst) =============

@api_router.get("/material-categories")
async def get_material_categories(current_user: User = Depends(get_current_user)):
    """Get all material categories"""
    cats = await db.material_categories.find({}, {"_id": 0}).sort("sort_order", 1).to_list(100)
    return cats

@api_router.post("/material-categories")
async def create_material_category(request: Request, current_user: User = Depends(get_current_user)):
    """Create a new material category (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen beheerders")
    body = await request.json()
    name = body.get("name", "")
    name_ua = body.get("name_ua") or None
    if not name:
        raise HTTPException(status_code=400, detail="Naam is verplicht")
    max_order = await db.material_categories.find_one(sort=[("sort_order", -1)])
    next_order = (max_order.get("sort_order", 0) + 1) if max_order else 0
    cat = MaterialCategory(name=name, name_ua=name_ua, sort_order=next_order)
    await db.material_categories.insert_one(cat.model_dump())
    result = cat.model_dump()
    result.pop("_id", None)
    return result

@api_router.put("/material-categories/{cat_id}")
async def update_material_category(cat_id: str, request: Request, current_user: User = Depends(get_current_user)):
    """Update category name"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen beheerders")
    body = await request.json()
    update_data = {}
    if "name" in body:
        update_data["name"] = body["name"]
    if "name_ua" in body:
        update_data["name_ua"] = body["name_ua"]
    if not update_data:
        raise HTTPException(status_code=400, detail="Geen wijzigingen")
    result = await db.material_categories.update_one({"id": cat_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Categorie niet gevonden")
    return {"message": "Categorie bijgewerkt"}

@api_router.delete("/material-categories/{cat_id}")
async def delete_material_category(cat_id: str, current_user: User = Depends(get_current_user)):
    """Delete a category and unlink its items"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen beheerders")
    result = await db.material_categories.delete_one({"id": cat_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Categorie niet gevonden")
    await db.material_catalog.update_many({"category_id": cat_id}, {"$set": {"category_id": None}})
    return {"message": "Categorie verwijderd"}

@api_router.get("/material-catalog")
async def get_material_catalog(current_user: User = Depends(get_current_user)):
    """Get all active catalog items"""
    query = {"active": True} if current_user.role == "worker" else {}
    items = await db.material_catalog.find(query, {"_id": 0}).sort("title", 1).to_list(500)
    return items

@api_router.post("/material-catalog")
async def create_catalog_item(item: MaterialCatalogItemCreate, current_user: User = Depends(get_current_user)):
    """Create a new catalog item (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen beheerders")
    doc = MaterialCatalogItem(title=item.title, title_ua=item.title_ua, category_id=item.category_id, description=item.description, sizes=item.sizes)
    await db.material_catalog.insert_one(doc.model_dump())
    result = doc.model_dump()
    result.pop("_id", None)
    return result

@api_router.put("/material-catalog/{item_id}")
async def update_catalog_item(item_id: str, update: MaterialCatalogItemUpdate, current_user: User = Depends(get_current_user)):
    """Update a catalog item (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen beheerders")
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Geen wijzigingen")
    result = await db.material_catalog.update_one({"id": item_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item niet gevonden")
    updated = await db.material_catalog.find_one({"id": item_id}, {"_id": 0})
    return updated

@api_router.delete("/material-catalog/{item_id}")
async def delete_catalog_item(item_id: str, current_user: User = Depends(get_current_user)):
    """Delete a catalog item (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen beheerders")
    result = await db.material_catalog.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item niet gevonden")
    return {"message": "Item verwijderd"}

@api_router.post("/material-catalog/{item_id}/upload-image")
async def upload_catalog_image(item_id: str, file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    """Upload an image for a catalog item"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen beheerders")
    item = await db.material_catalog.find_one({"id": item_id})
    if not item:
        raise HTTPException(status_code=404, detail="Item niet gevonden")
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Alleen afbeeldingen toegestaan")
    catalog_dir = ROOT_DIR / "uploads" / "catalog"
    catalog_dir.mkdir(parents=True, exist_ok=True)
    unique_filename = f"{item_id}_{uuid.uuid4().hex[:8]}_{file.filename}"
    file_path = catalog_dir / unique_filename
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    image_url = f"/api/static/catalog/{unique_filename}"
    await db.material_catalog.update_one({"id": item_id}, {"$set": {"image_url": image_url}})
    return {"image_url": image_url}

@api_router.post("/material-orders")
async def create_material_order(order: MaterialOrderCreate, current_user: User = Depends(get_current_user)):
    """Create material order from catalog (worker submits order)"""
    if current_user.role not in ["worker", "admin"]:
        raise HTTPException(status_code=403, detail="Alleen werkmannen")
    if not order.items:
        raise HTTPException(status_code=400, detail="Geen items geselecteerd")
    # Create individual material_requests for each item in the order
    created = []
    delivery_text = order.delivery_date if order.delivery_date else "Zo snel mogelijk"
    for item in order.items:
        size_text = f" ({item.get('selected_size', '')})" if item.get('selected_size') else ""
        doc = MaterialRequest(
            title=f"{item['title']}{size_text}",
            quantity=str(item.get('quantity', 1)),
            needed_by=delivery_text,
            photo_url=item.get('image_url'),
            notes=order.notes,
            project_id=order.project_id,
            project_name=order.project_name,
            requested_by=current_user.id,
            requested_by_name=current_user.name or current_user.username
        )
        await db.material_requests.insert_one(doc.model_dump())
        created.append(doc.id)
    return {"message": f"{len(created)} materialen besteld", "ids": created}

# ============= MAINTENANCE ENDPOINTS =============

MAINTENANCE_TYPES = {
    "verwarming": "Centrale Verwarming",
    "ventilatie": "Ventilatie",
    "waterfilter": "Waterfilter"
}

@api_router.post("/maintenance")
async def create_maintenance_contract(contract: MaintenanceContractCreate, current_user: User = Depends(get_current_user)):
    """Create a new maintenance contract/dossier"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Alleen admins kunnen onderhoudsdossiers aanmaken")
    
    if contract.maintenance_type not in MAINTENANCE_TYPES:
        raise HTTPException(status_code=400, detail=f"Ongeldig type. Kies uit: {', '.join(MAINTENANCE_TYPES.keys())}")
    
    # Create contract
    new_contract = MaintenanceContract(
        client_name=contract.client_name,
        client_email=contract.client_email,
        client_phone=contract.client_phone,
        client_address=contract.client_address,
        client_postal_code=contract.client_postal_code,
        client_city=contract.client_city,
        maintenance_type=contract.maintenance_type,
        description=contract.description,
        scheduled_date=contract.scheduled_date,
        frequency_months=contract.frequency_months,
        service_price=contract.service_price,
        notes=contract.notes,
        created_by=current_user.id
    )
    
    contract_doc = new_contract.model_dump()
    contract_doc["created_at"] = contract_doc["created_at"].isoformat()
    contract_doc["updated_at"] = contract_doc["updated_at"].isoformat()
    
    await db.maintenance_contracts.insert_one(contract_doc)
    
    logger.info(f"Created maintenance contract: {new_contract.id} for {contract.client_name}")
    
    return {"message": "Onderhoudsdossier aangemaakt", "contract_id": new_contract.id}

@api_router.get("/maintenance")
async def get_maintenance_contracts(
    status: Optional[str] = None,
    maintenance_type: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all maintenance contracts"""
    if current_user.role not in ["admin", "worker"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    filter_query = {}
    if status:
        filter_query["status"] = status
    if maintenance_type:
        filter_query["maintenance_type"] = maintenance_type
    
    contracts = await db.maintenance_contracts.find(filter_query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Add type labels
    for contract in contracts:
        contract["maintenance_type_label"] = MAINTENANCE_TYPES.get(contract.get("maintenance_type"), "Onbekend")
    
    return contracts

@api_router.get("/maintenance/{contract_id}")
async def get_maintenance_contract(contract_id: str, current_user: User = Depends(get_current_user)):
    """Get a specific maintenance contract with purchases"""
    if current_user.role not in ["admin", "worker"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    contract = await db.maintenance_contracts.find_one({"id": contract_id}, {"_id": 0})
    if not contract:
        raise HTTPException(status_code=404, detail="Onderhoudsdossier niet gevonden")
    
    # Get purchases
    purchases = await db.maintenance_purchases.find({"maintenance_id": contract_id}, {"_id": 0}).to_list(100)
    contract["purchases"] = purchases
    
    # Calculate total materials cost
    contract["materials_cost"] = sum(p.get("amount", 0) for p in purchases)
    
    # Get invoices
    invoices = await db.maintenance_invoices.find({"maintenance_id": contract_id}, {"_id": 0}).to_list(100)
    contract["invoices"] = invoices
    
    contract["maintenance_type_label"] = MAINTENANCE_TYPES.get(contract.get("maintenance_type"), "Onbekend")
    
    return contract

@api_router.put("/maintenance/{contract_id}")
async def update_maintenance_contract(contract_id: str, update: MaintenanceContractUpdate, current_user: User = Depends(get_current_user)):
    """Update a maintenance contract"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Alleen admins kunnen dossiers wijzigen")
    
    contract = await db.maintenance_contracts.find_one({"id": contract_id})
    if not contract:
        raise HTTPException(status_code=404, detail="Onderhoudsdossier niet gevonden")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.maintenance_contracts.update_one({"id": contract_id}, {"$set": update_data})
    
    return {"message": "Onderhoudsdossier bijgewerkt"}

@api_router.delete("/maintenance/{contract_id}")
async def delete_maintenance_contract(contract_id: str, current_user: User = Depends(get_current_user)):
    """Delete a maintenance contract"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen dossiers verwijderen")
    
    result = await db.maintenance_contracts.delete_one({"id": contract_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Onderhoudsdossier niet gevonden")
    
    # Delete associated purchases and invoices
    await db.maintenance_purchases.delete_many({"maintenance_id": contract_id})
    await db.maintenance_invoices.delete_many({"maintenance_id": contract_id})
    
    return {"message": "Onderhoudsdossier verwijderd"}

# --- Maintenance Purchases (Aankoopfacturen) ---

@api_router.post("/maintenance/{contract_id}/purchases")
async def add_maintenance_purchase(contract_id: str, purchase: MaintenancePurchaseCreate, current_user: User = Depends(get_current_user)):
    """Add a purchase/expense to a maintenance contract"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Alleen admins kunnen aankopen toevoegen")
    
    contract = await db.maintenance_contracts.find_one({"id": contract_id})
    if not contract:
        raise HTTPException(status_code=404, detail="Onderhoudsdossier niet gevonden")
    
    # Calculate total
    total = purchase.amount + purchase.vat_amount
    
    new_purchase = MaintenancePurchase(
        maintenance_id=contract_id,
        supplier=purchase.supplier,
        invoice_number=purchase.invoice_number,
        invoice_date=purchase.invoice_date,
        description=purchase.description,
        amount=purchase.amount,
        vat_amount=purchase.vat_amount,
        total_amount=total
    )
    
    purchase_doc = new_purchase.model_dump()
    purchase_doc["created_at"] = purchase_doc["created_at"].isoformat()
    
    await db.maintenance_purchases.insert_one(purchase_doc)
    
    # Update contract's materials cost
    all_purchases = await db.maintenance_purchases.find({"maintenance_id": contract_id}).to_list(100)
    total_materials = sum(p.get("amount", 0) for p in all_purchases)
    await db.maintenance_contracts.update_one(
        {"id": contract_id},
        {"$set": {"materials_cost": total_materials, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Aankoop toegevoegd", "purchase_id": new_purchase.id}

@api_router.delete("/maintenance/{contract_id}/purchases/{purchase_id}")
async def delete_maintenance_purchase(contract_id: str, purchase_id: str, current_user: User = Depends(get_current_user)):
    """Delete a purchase from a maintenance contract"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen aankopen verwijderen")
    
    result = await db.maintenance_purchases.delete_one({"id": purchase_id, "maintenance_id": contract_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Aankoop niet gevonden")
    
    # Recalculate materials cost
    all_purchases = await db.maintenance_purchases.find({"maintenance_id": contract_id}).to_list(100)
    total_materials = sum(p.get("amount", 0) for p in all_purchases)
    await db.maintenance_contracts.update_one(
        {"id": contract_id},
        {"$set": {"materials_cost": total_materials, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Aankoop verwijderd"}

# --- Maintenance Invoices (Verkoopfacturen) ---

@api_router.post("/maintenance/{contract_id}/invoices")
async def create_maintenance_invoice(contract_id: str, invoice: MaintenanceInvoiceCreate, current_user: User = Depends(get_current_user)):
    """Create an invoice for a maintenance contract"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen facturen aanmaken")
    
    contract = await db.maintenance_contracts.find_one({"id": contract_id})
    if not contract:
        raise HTTPException(status_code=404, detail="Onderhoudsdossier niet gevonden")
    
    # Calculate amounts
    subtotal = invoice.service_amount + invoice.materials_amount
    vat_amount = subtotal * (invoice.vat_rate / 100)
    total_amount = subtotal + vat_amount
    
    # Generate invoice number (MAINT-YYYY-XXXX)
    year = datetime.now().year
    count = await db.maintenance_invoices.count_documents({"invoice_number": {"$regex": f"^MAINT-{year}"}})
    invoice_number = f"MAINT-{year}-{str(count + 1).zfill(4)}"
    
    # Set dates
    invoice_date = invoice.invoice_date or datetime.now().strftime("%Y-%m-%d")
    due_date = invoice.due_date or (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")
    
    new_invoice = MaintenanceInvoice(
        maintenance_id=contract_id,
        invoice_number=invoice_number,
        invoice_date=invoice_date,
        due_date=due_date,
        service_amount=invoice.service_amount,
        materials_amount=invoice.materials_amount,
        subtotal=subtotal,
        vat_rate=invoice.vat_rate,
        vat_amount=round(vat_amount, 2),
        total_amount=round(total_amount, 2)
    )
    
    invoice_doc = new_invoice.model_dump()
    invoice_doc["created_at"] = invoice_doc["created_at"].isoformat()
    
    await db.maintenance_invoices.insert_one(invoice_doc)
    
    # Update contract status and link invoice
    await db.maintenance_contracts.update_one(
        {"id": contract_id},
        {
            "$set": {"status": "gefactureerd", "updated_at": datetime.now(timezone.utc).isoformat()},
            "$push": {"invoice_ids": new_invoice.id}
        }
    )
    
    return {
        "message": "Factuur aangemaakt",
        "invoice_id": new_invoice.id,
        "invoice_number": invoice_number,
        "total_amount": total_amount
    }

@api_router.put("/maintenance/invoices/{invoice_id}/status")
async def update_maintenance_invoice_status(invoice_id: str, status: str = Query(...), current_user: User = Depends(get_current_user)):
    """Update invoice status (verstuurd, betaald)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen factuurstatus wijzigen")
    
    valid_statuses = ["concept", "verstuurd", "betaald"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Ongeldige status. Kies uit: {', '.join(valid_statuses)}")
    
    update_data = {"status": status}
    if status == "betaald":
        update_data["paid_date"] = datetime.now().strftime("%Y-%m-%d")
    
    result = await db.maintenance_invoices.update_one({"id": invoice_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Factuur niet gevonden")
    
    return {"message": f"Factuurstatus gewijzigd naar {status}"}

# --- Maintenance Financial Overview ---

@api_router.get("/maintenance/finances/overview")
async def get_maintenance_finances(year: int = Query(default=None), month: int = Query(default=None), current_user: User = Depends(get_current_user)):
    """Get financial overview for maintenance contracts"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Alleen admins kunnen financiën bekijken")
    
    if not year:
        year = datetime.now().year
    
    # Build date filter for invoices
    if month:
        start_date = f"{year}-{str(month).zfill(2)}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{str(month + 1).zfill(2)}-01"
    else:
        start_date = f"{year}-01-01"
        end_date = f"{year + 1}-01-01"
    
    # Get all paid invoices in period
    invoices = await db.maintenance_invoices.find({
        "status": "betaald",
        "invoice_date": {"$gte": start_date, "$lt": end_date}
    }, {"_id": 0}).to_list(1000)
    
    # Get all purchases in period
    purchases = await db.maintenance_purchases.find({
        "invoice_date": {"$gte": start_date, "$lt": end_date}
    }, {"_id": 0}).to_list(1000)
    
    # Calculate totals
    total_revenue = sum(inv.get("total_amount", 0) for inv in invoices)
    total_service_revenue = sum(inv.get("service_amount", 0) for inv in invoices)
    total_materials_revenue = sum(inv.get("materials_amount", 0) for inv in invoices)
    total_costs = sum(p.get("amount", 0) for p in purchases)
    total_profit = total_revenue - total_costs
    
    # Per type breakdown
    type_breakdown = {}
    for mtype, mlabel in MAINTENANCE_TYPES.items():
        type_contracts = await db.maintenance_contracts.find({"maintenance_type": mtype}, {"_id": 0, "id": 1}).to_list(1000)
        contract_ids = [c["id"] for c in type_contracts]
        
        type_invoices = [inv for inv in invoices if inv.get("maintenance_id") in contract_ids]
        type_purchases = [p for p in purchases if p.get("maintenance_id") in contract_ids]
        
        type_revenue = sum(inv.get("total_amount", 0) for inv in type_invoices)
        type_costs = sum(p.get("amount", 0) for p in type_purchases)
        
        type_breakdown[mtype] = {
            "label": mlabel,
            "revenue": type_revenue,
            "costs": type_costs,
            "profit": type_revenue - type_costs,
            "contract_count": len(contract_ids),
            "invoice_count": len(type_invoices)
        }
    
    # Monthly breakdown (if no specific month)
    monthly_data = []
    if not month:
        for m in range(1, 13):
            m_start = f"{year}-{str(m).zfill(2)}-01"
            if m == 12:
                m_end = f"{year + 1}-01-01"
            else:
                m_end = f"{year}-{str(m + 1).zfill(2)}-01"
            
            m_invoices = [inv for inv in invoices if m_start <= inv.get("invoice_date", "") < m_end]
            m_purchases = [p for p in purchases if m_start <= p.get("invoice_date", "") < m_end]
            
            m_revenue = sum(inv.get("total_amount", 0) for inv in m_invoices)
            m_costs = sum(p.get("amount", 0) for p in m_purchases)
            
            monthly_data.append({
                "month": m,
                "month_name": ["Jan", "Feb", "Mrt", "Apr", "Mei", "Jun", "Jul", "Aug", "Sep", "Okt", "Nov", "Dec"][m-1],
                "revenue": m_revenue,
                "costs": m_costs,
                "profit": m_revenue - m_costs
            })
    
    return {
        "year": year,
        "month": month,
        "total_revenue": total_revenue,
        "total_service_revenue": total_service_revenue,
        "total_materials_revenue": total_materials_revenue,
        "total_costs": total_costs,
        "total_profit": total_profit,
        "margin_percentage": round((total_profit / total_revenue * 100) if total_revenue > 0 else 0, 1),
        "invoice_count": len(invoices),
        "type_breakdown": type_breakdown,
        "monthly_data": monthly_data
    }

# --- Mark maintenance as completed ---

@api_router.post("/maintenance/{contract_id}/complete")
async def complete_maintenance(contract_id: str, technician_notes: str = Query(default=""), current_user: User = Depends(get_current_user)):
    """Mark a maintenance contract as completed"""
    if current_user.role not in ["admin", "worker"]:
        raise HTTPException(status_code=403, detail="Geen toegang")
    
    contract = await db.maintenance_contracts.find_one({"id": contract_id})
    if not contract:
        raise HTTPException(status_code=404, detail="Onderhoudsdossier niet gevonden")
    
    today = datetime.now().strftime("%Y-%m-%d")
    frequency = contract.get("frequency_months", 12)
    next_date = (datetime.now() + timedelta(days=frequency * 30)).strftime("%Y-%m-%d")
    
    await db.maintenance_contracts.update_one(
        {"id": contract_id},
        {"$set": {
            "status": "uitgevoerd",
            "last_maintenance_date": today,
            "next_maintenance_date": next_date,
            "technician_notes": technician_notes,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Onderhoud gemarkeerd als uitgevoerd", "next_maintenance_date": next_date}

# Include routers
app.include_router(api_router)
app.include_router(auth2_router, prefix="/api")

# CORS configuration - explicit origins for mobile browser compatibility
cors_origins = os.environ.get('CORS_ORIGINS', '')
if cors_origins:
    origins_list = [o.strip() for o in cors_origins.split(',') if o.strip()]
else:
    # Default origins including common production and preview URLs
    origins_list = [
        "https://dashboard.qtechnics.be",
        "https://www.dashboard.qtechnics.be",
        "http://localhost:3000",
        "http://localhost:8001",
    ]
    # Add preview URL pattern
    import re
    
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],  # Allow all origins
    allow_origin_regex=r"https://.*\.preview\.emergentagent\.com",  # Also allow preview URLs
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============= EMAIL NOTIFICATION HELPER =============
async def send_customer_notification(project_id: str, subject: str, content_description: str):
    """
    Send email notification to customer when new content is added to their portal.
    Looks up customer email from the project's associated lead.
    """
    if not resend.api_key:
        logger.warning("RESEND_API_KEY not configured - skipping email notification")
        return None
    
    try:
        # Get project to find lead_id
        project = await db.projects.find_one({"id": project_id}, {"_id": 0})
        if not project:
            logger.warning(f"Project {project_id} not found for email notification")
            return None
        
        lead_id = project.get("lead_id")
        if not lead_id:
            logger.warning(f"No lead_id found for project {project_id}")
            return None
        
        # Get lead to find customer email
        lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
        if not lead or not lead.get("email"):
            logger.warning(f"No email found for lead {lead_id}")
            return None
        
        customer_email = lead.get("email")
        customer_name = lead.get("name", "Klant")
        project_name = project.get("name", "Uw project")
        
        # Check for customer access token for portal link
        portal_token = project.get("customer_access_token")
        portal_link = ""
        if portal_token:
            portal_link = f"<p><a href='{os.environ.get('APP_URL', 'https://renovation-calc-5.preview.emergentagent.com')}/customer/{portal_token}' style='background-color: #1E40AF; color: white; padding: 12px 24px; text-decoration: none; border-radius: 8px; display: inline-block;'>Bekijk uw project</a></p>"
        
        # Build HTML email
        html_content = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <h2 style="color: #1E40AF;">Update voor {project_name}</h2>
            <p>Beste {customer_name},</p>
            <p>{content_description}</p>
            {portal_link}
            <hr style="border: none; border-top: 1px solid #E5E7EB; margin: 20px 0;" />
            <p style="color: #6B7280; font-size: 12px;">
                Met vriendelijke groeten,<br/>
                Q-Technics
            </p>
        </div>
        """
        
        params = {
            "from": SENDER_EMAIL,
            "to": [customer_email],
            "subject": subject,
            "html": html_content
        }
        
        # Run sync SDK in thread to keep FastAPI non-blocking
        email_result = await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Email notification sent to {customer_email} for project {project_id}")
        return email_result
        
    except Exception as e:
        logger.error(f"Failed to send email notification: {str(e)}")
        return None

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()